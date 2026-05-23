import openpyxl
from collections import Counter

p = r'c:\Users\aviso\OneDrive\Desktop\trading-alerts\backtest\results\h1only_20260401_20260408\trades.xlsx'
wb = openpyxl.load_workbook(p, data_only=True)
ws = wb['Trades']
h = [c.value for c in ws[1]]
iZone=h.index('Entry Type'); iPair=h.index('Currency Pair'); iHow=h.index('How Trade Closed')
iPL=h.index('Dollar P&L'); iR=h.index('R Achieved')
iNews=h.index('News Blocked'); iIST=h.index('IST Window Blocked')
iAlert=h.index('Alert Time (UTC)')

rows=list(ws.iter_rows(min_row=2, values_only=True))

c=Counter()
for r in rows: c[r[iZone]] += 1
print('Entry Type counts:', dict(c))
print()

prox = [r for r in rows if r[iZone] and 'Proximal' in str(r[iZone])]
print(f'Total Proximal rows: {len(prox)}')
print()
print('All Proximal rows:')
for r in prox:
    print(f'  {str(r[iAlert])[:19]:<20} {str(r[iPair]):<7} {str(r[iHow]):<32} R={r[iR]!r:<8} PL={r[iPL]!r:<10} news={r[iNews]} ist={r[iIST]}')
print()
total_all_prox = sum((r[iPL] or 0) for r in prox)
prox_clean = [r for r in prox if not r[iNews] and not r[iIST]]
total_clean = sum((r[iPL] or 0) for r in prox_clean)
sumR_clean = sum((r[iR] or 0) for r in prox_clean)
print(f'Sum PL all proximal rows:                  {total_all_prox}')
print(f'Proximal rows after news=F AND ist=F:      {len(prox_clean)}')
print(f'Sum PL of those (matches your Excel sum):  {total_clean}')
print(f'Sum R of those:                            {sumR_clean}')
print(f'Sum R * $250:                              {sumR_clean*250}')
# breakdown: filled vs never_filled
filled_clean = [r for r in prox_clean if r[iHow] and 'Never filled' not in str(r[iHow])]
never_clean = [r for r in prox_clean if r[iHow] and 'Never filled' in str(r[iHow])]
print(f'  of which filled (How != Never filled): {len(filled_clean)}, sum PL={sum((r[iPL] or 0) for r in filled_clean)}, sum R={sum((r[iR] or 0) for r in filled_clean)}')
print(f'  of which never_filled:                 {len(never_clean)}, sum PL={sum((r[iPL] or 0) for r in never_clean)}, sum R={sum((r[iR] or 0) for r in never_clean)}')
