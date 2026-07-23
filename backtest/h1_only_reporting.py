"""H1-only backtest report writer.

Produces:
  results/<run_id>/trades.csv          machine-readable, full column set (both zones)
  results/<run_id>/trades.xlsx         human-readable, plain-English column names
  results/<run_id>/report_forex.html   Book A email (FX majors + Gold), proximal-only
  results/<run_id>/report_gold_nas.html Book B email (new FX + BTC), proximal-only
  results/<run_id>/raw_alerts.jsonl    OB-touch alerts before simulation
  results/<run_id>/summary.json        all metrics, machine-readable
(The combined report.html was removed 2026-06-30: it was never emailed and
rendered the per-zone block twice. The two per-group emails are the only HTML.)
"""

from __future__ import annotations

import json
import os
import time
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd

from backtest.insights import win_rate_pct as _win_rate
from backtest.insights import bootstrap_ci as _bootstrap_ci
from backtest import insights as _insights


# ---------------------------------------------------------------------------
# Global trade ID counter  (A0001 … A9999, B0000 … Z9999)
# ---------------------------------------------------------------------------
# Persists across runs in backtest/results/.trade_id_counter (one integer).
# Never resets. Each run claims the next N slots atomically via a lock file
# so parallel runs cannot collide.  Format: letter prefix (A–Z) + 4-digit
# zero-padded number.  260 000 total slots; current inventory ~2 400 rows.

_COUNTER_FILE = Path(__file__).resolve().parent / ".trade_id_counter"
_LOCK_FILE    = Path(__file__).resolve().parent / ".trade_id_counter.lock"


def _int_to_trade_id(n: int) -> str:
    """Convert a 1-based integer to A0001…Z9999 format.

    Mapping: 1→A0001, 2→A0002, …, 9999→A9999, 10000→B0000, 10001→B0001, …
    Each letter covers 10 000 slots (0000–9999). A0000 is intentionally
    skipped (counter starts at 1) so every displayed ID reads non-zero.
    """
    idx = n - 1                     # 0-based: 0→A0001 after digit adjustment
    letter = chr(ord("A") + idx // 10000)
    digits = (idx % 10000) + 1      # shift 0→1 so A-block is A0001..A9999+1
    if digits > 9999:               # A-block overflow rolls into B0000
        letter = chr(ord(letter) + 1)
        digits = 0
    return f"{letter}{digits:04d}"


def _claim_trade_ids(count: int) -> int:
    """Atomically claim `count` IDs. Returns the first integer to use (1-based).
    Cross-platform: uses an exclusive lock file (works on Windows and Linux)."""
    _COUNTER_FILE.parent.mkdir(parents=True, exist_ok=True)
    # Acquire lock: spin on O_EXCL create (atomic on all OSes).
    deadline = time.monotonic() + 10.0
    while True:
        try:
            fd = os.open(str(_LOCK_FILE), os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            os.close(fd)
            break
        except FileExistsError:
            if time.monotonic() > deadline:
                raise RuntimeError("trade_id_counter lock timeout — delete .trade_id_counter.lock and retry")
            time.sleep(0.05)
    try:
        raw = _COUNTER_FILE.read_text().strip() if _COUNTER_FILE.exists() else ""
        current = int(raw) if raw else 0
        first = current + 1
        _COUNTER_FILE.write_text(str(current + count))
        return first
    finally:
        _LOCK_FILE.unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# Aggregation helpers (used by summary.json + HTML)
# ---------------------------------------------------------------------------

# Trades to exclude from every aggregate. These are NOT real wins or losses:
#   never_filled = limit never hit, no position ever opened (r_realised = 0).
#   timeout      = filled, then force-closed at the 48-bar hold cap with neither
#                  SL nor TP touched -- a snapshot at an arbitrary clock moment.
#   window_end   = filled, then force-closed because the BACKTEST DATA RAN OUT
#                  before SL/TP/timeout -- the exit price is set by where the
#                  data file ends, not by the market. Shift the data window and
#                  the same trade gets a different P&L (or resolves for real).
# Both timeout and window_end are unresolved positions closed at an arbitrary
# close price: measurement artifacts, not edge. Folding them into the headline
# would distort expectancy for reasons that have nothing to do with the system.
# These rows REMAIN in CSV/Excel and in the exit-reason counts (so the trader
# sees how many trades didn't resolve) but never feed P&L, win rate, expectancy,
# or any reported metric. Audit-only, not hidden. (RCA #5; veteran SMC call
# 2026-06-16.) This set is the single source of truth -- every "is this a real
# resolved trade?" check in this module routes through _is_real_filled.
_EXCLUDE_REASONS = {"never_filled", "timeout", "window_end"}


def _is_real_filled(t: Dict[str, Any]) -> bool:
    return t.get("exit_reason") not in _EXCLUDE_REASONS


def _config_score_floor() -> float:
    """Live's MIN_SCORE_TO_EMAIL (config.scoring.min_score_to_email, default 4).

    Read from config.json so the backtest headline auto-tracks live: change the
    floor in one place and both move together. Live Phase 2 gates on raw score
    < floor (Phase2_Alert_Engine.MIN_SCORE_TO_EMAIL); the backtest must apply
    the identical floor or its headline reports trades live would never email.
    """
    try:
        cfg_path = Path(__file__).resolve().parent.parent / "config.json"
        with open(cfg_path) as f:
            cfg = json.load(f)
        return float(cfg.get("scoring", {}).get("min_score_to_email", 4))
    except Exception:
        return 4.0


SCORE_FLOOR = _config_score_floor()


def _below_score_floor(t: Dict[str, Any]) -> bool:
    """RETIRED (2026-06-30, trader decision): the backtest no longer filters on
    score ANYWHERE. We want the true, unfiltered P&L of every OB-touch the system
    found — score is recorded per row for analysis but never gates a trade out of
    the headline or any aggregate. This always returns False so the headline
    reflects all resolved trades regardless of score.

    Kept as a function (not deleted) so the score-bucket discovery table and the
    G1 eligibility path keep a single, stable call site; flipping the floor back
    on would be a one-line change here if ever needed."""
    return False


def _headline_exclusion(t: Dict[str, Any]) -> str:
    """THE single rule for whether a row feeds the headline -- expressed as the
    REASON it is excluded. Returns "" when the row is eligible (feeds the
    headline + every aggregate), else a short machine-stable tag naming why.

    Tags:
      unresolved:timeout / unresolved:window_end -- filled but force-closed at
          an arbitrary price (hold cap / data ran out). Audit-only, never P&L.
      never_filled -- no real position was ever held.
      ist_blocked -- alert fell outside the 09:00-24:00 IST trading window;
          live's smc_radar blackout means it could never have alerted.
      weekend_blocked -- (crypto only) fill fell inside the configured weekend
          no-trade window (Sat 00:00 -> Mon 09:00 IST for BTC). We do not trade
          crypto weekends, so these are audit-only and never feed P&L.

    This is the ONE place the eligibility rule lives. `_is_eligible` (used by the
    scoreboards and imported by the G1 gate), the CSV/Excel export, and the
    reconciliation test all route through it, so the exported file is always
    reconcilable to the headline: sum(pnl_usd where eligible) == headline.

    Score does NOT gate (2026-06-30): the backtest reports the true P&L of every
    OB-touch regardless of score. Killzone and news are display signals in live
    (they never suppress an alert), so they are audit-only here and must NOT gate.
    """
    er = t.get("exit_reason")
    if er in _EXCLUDE_REASONS:
        return f"unresolved:{er}" if er in ("timeout", "window_end") else str(er)
    if t.get("ist_blocked"):
        return "ist_blocked"
    if t.get("weekend_blocked"):
        return "weekend_blocked"
    return ""


def _is_eligible(t: Dict[str, Any]) -> bool:
    """True iff the row feeds the headline + every aggregate. Thin wrapper over
    _headline_exclusion so there is exactly one definition of eligibility."""
    return _headline_exclusion(t) == ""


# Hard pre-filter conditions: rows we structurally do not trade (IST window,
# crypto weekend). These are dropped UPFRONT before any scoreboard is built,
# unlike the fill-resolution exclusions (timeout / never_filled) which are kept
# here and filtered later by the aggregator (the never-filled audit section still
# needs them). Defined ONCE so every pre-filter site stays in lock-step with
# _headline_exclusion — the G1 reconciliation depends on these dropping the exact
# same rows. A new "could never trade" condition must be added HERE and in
# _headline_exclusion together. NOTE: score does NOT gate (2026-06-30) — the
# backtest reports the true P&L of every OB-touch regardless of score.
def _is_hard_blocked(t: Dict[str, Any]) -> bool:
    return bool(t.get("ist_blocked") or t.get("weekend_blocked"))


def _aggregate_for_exit(trades: List[Dict[str, Any]], r_col: str,
                        risk_usd: float) -> Dict[str, Any]:
    """Aggregate filled trades under a hypothetical exit policy."""
    filled = [t for t in trades if _is_real_filled(t)]
    if not filled:
        return {"trades": 0, "filled": 0}
    df = pd.DataFrame(filled)
    if r_col not in df.columns:
        return {"trades": 0, "filled": 0}
    wins   = df[df[r_col] > 0]
    losses = df[df[r_col] < 0]
    bes    = df[df[r_col] == 0]
    exp    = float(df[r_col].mean())
    avg_mfe = float(wins["mfe_r"].mean()) if (len(wins) and "mfe_r" in wins.columns) else 0
    avg_mae = float(losses["mae_r"].mean()) if (len(losses) and "mae_r" in losses.columns) else 0
    capture = round(float(wins[r_col].mean()) / avg_mfe * 100, 0) if (avg_mfe > 0 and len(wins)) else 0
    return {
        "trades":         int(len(df)),
        "wins":           int(len(wins)),
        "losses":         int(len(losses)),
        "breakevens":     int(len(bes)),
        "win_rate_pct":   _win_rate(df, r_col),
        "expectancy_r":   round(exp, 3),
        "expectancy_usd": round(exp * risk_usd, 2),
        "total_r":        round(float(df[r_col].sum()), 3),
        "total_pnl_usd":  round(float(df[r_col].sum()) * risk_usd, 2),
        "avg_win_r":      round(float(wins[r_col].mean()), 3) if len(wins) else 0.0,
        "avg_loss_r":     round(float(losses[r_col].mean()), 3) if len(losses) else 0.0,
        "avg_win_usd":    round(float(wins[r_col].mean()) * risk_usd, 0) if len(wins) else 0.0,
        "avg_loss_usd":   round(float(losses[r_col].mean()) * risk_usd, 0) if len(losses) else 0.0,
        "avg_mfe_r":      round(avg_mfe, 3),
        "avg_mae_r":      round(avg_mae, 3),
        "win_capture_pct": capture,
    }


def _per_pair_breakdown(trades: List[Dict[str, Any]], r_col: str,
                        risk_usd: float) -> List[Dict[str, Any]]:
    """Per-pair scoreboard. `verdict` reuses the SAME CI + per-quarter test as
    the headline banner (_stat_block) — a pair is only "edge"/"loser" when its
    own bootstrap 95% CI clears zero, never from win rate or raw P&L sign
    alone. This keeps the pair table honest on thin per-pair samples instead
    of asserting a driver off a handful of trades."""
    filled = [t for t in trades if _is_real_filled(t)]
    if not filled:
        return []
    df = pd.DataFrame(filled)
    out = []
    for pair, sub in df.groupby("pair"):
        vals = [float(v) for v in sub[r_col].tolist()]
        ts = sub["alert_ts"].tolist() if "alert_ts" in sub.columns else None
        stat = _stat_block(vals, ts)
        out.append({
            "pair":          pair,
            "trades":        int(len(sub)),
            "win_rate_pct":  _win_rate(sub, r_col),
            "expectancy_r":  round(float(sub[r_col].mean()), 3),
            "total_pnl_usd": round(float(sub[r_col].sum()) * risk_usd, 2),
            "verdict":       stat["verdict"],
        })
    out.sort(key=lambda r: r["total_pnl_usd"], reverse=True)
    return out


def _per_session_breakdown(trades: List[Dict[str, Any]], r_col: str,
                           risk_usd: float) -> List[Dict[str, Any]]:
    filled = [t for t in trades if _is_real_filled(t)]
    if not filled:
        return []
    df = pd.DataFrame(filled)
    if "session" not in df.columns:
        return []
    out = []
    for sess, sub in df.groupby("session"):
        out.append({
            "session":       sess,
            "trades":        int(len(sub)),
            "win_rate_pct":  _win_rate(sub, r_col),
            "expectancy_r":  round(float(sub[r_col].mean()), 3),
            "total_pnl_usd": round(float(sub[r_col].sum()) * risk_usd, 2),
        })
    session_order = {"Asia": 0, "London": 1, "NY": 2, "Other": 3}
    out.sort(key=lambda r: session_order.get(r["session"], 9))
    return out


def _fill_rate(trades: List[Dict[str, Any]], entry_zone: str) -> Dict[str, Any]:
    zone_rows = [t for t in trades if t.get("entry_zone") == entry_zone]
    if not zone_rows:
        return {"alerts": 0, "filled": 0, "fill_rate_pct": 0.0}
    filled = [t for t in zone_rows if _is_real_filled(t)]
    return {
        "alerts":        len(zone_rows),
        "filled":        len(filled),
        "fill_rate_pct": round(len(filled) / len(zone_rows) * 100, 1) if zone_rows else 0.0,
    }


def _score_buckets(trades: List[Dict[str, Any]], r_col: str) -> List[Dict[str, Any]]:
    filled = [t for t in trades if _is_real_filled(t)]
    if not filled:
        return []
    df = pd.DataFrame(filled)
    if "score" not in df.columns or r_col not in df.columns:
        return []
    edges  = [0, 2, 3, 4, 5, 6, 7, 99]
    labels = ["0-2", "2-3", "3-4", "4-5", "5-6", "6-7", "7+"]
    df["bucket"] = pd.cut(df["score"], bins=edges, labels=labels,
                          right=False, include_lowest=True)
    out = []
    for label in labels:
        sub = df[df["bucket"] == label]
        if sub.empty:
            continue
        out.append({
            "score_bucket":   label,
            "trades":         int(len(sub)),
            "win_rate_pct":   _win_rate(sub, r_col),
            "expectancy_r":   round(float(sub[r_col].mean()), 3),
        })
    return out


def _exit_reason_counts(trades: List[Dict[str, Any]]) -> Dict[str, int]:
    counts: Dict[str, int] = {}
    for t in trades:
        r = t.get("exit_reason", "unknown")
        counts[r] = counts.get(r, 0) + 1
    return counts


def _flag_vet_review(t: Dict[str, Any]) -> Tuple[bool, str]:
    r    = t.get("r_realised", 0)
    mfe  = t.get("mfe_r", 0)
    score = t.get("score", 0)
    if r > 0 and mfe > r * 2 and mfe > 1.5:
        return True, f"Left money on table — price reached {mfe:.1f}R, you booked {r:.1f}R"
    if r < 0 and mfe > 0.5:
        return True, f"Nearly worked — price reached +{mfe:.1f}R before reversing to SL"
    if score >= 4 and r < 0:
        return True, f"High-score setup ({score:.1f}) that lost — worth understanding why"
    return False, ""


# ---------------------------------------------------------------------------
# Sample-size discipline + exit-policy math
# ---------------------------------------------------------------------------

# Below this trade count, comparative verdicts are demoted to
# "insufficient sample" rather than asserted. Reader still sees the numbers.
MIN_N_FOR_VERDICT = 5

# Driver-engine guard (the §4 anti-phantom-pattern gate). A surfaced bucket is
# promoted to a "driver" only if it clears ALL THREE: enough trades, a real gap
# from the system base-rate, and a per-quarter sign that mostly holds. A bucket
# that misses any one is NOT hidden -- it is flagged "directional, thin" so the
# reader sees it but is told not to trust it as a rule. (Trader: flag, never hide.)
DRIVER_MIN_N          = 8      # below this: thin, flagged not promoted
DRIVER_MIN_EXR_GAP    = 0.20   # |bucket expR - base expR| must clear this (R)
DRIVER_PQ_SIGN_FRAC   = 0.60   # share of quarters whose sign matches the bucket


# ---------------------------------------------------------------------------
# Shared statistics engine (one implementation, reused by the headline, the
# driver engine, and the exit-recipe table). Every "is this real?" call in the
# email goes through here so the method is identical everywhere:
#   bootstrap 95% CI  +  per-quarter sign  +  a plain one-line verdict.
# RECOMMENDATIONS.md mandates CI + per-quarter on every "is it good" call.
# ---------------------------------------------------------------------------

def _quarter_of(ts_str: Any) -> Optional[str]:
    """Calendar quarter label (e.g. '2025Q3') from an alert timestamp. None on
    unparseable input. Matches exit_lab._quarter so backtest views agree."""
    if not ts_str:
        return None
    try:
        ts = pd.to_datetime(ts_str, utc=True)
    except Exception:
        return None
    return f"{ts.year}Q{(ts.month - 1) // 3 + 1}"




def _pq_sign_consistency(pq: Dict[str, float], expr: float) -> Tuple[int, int]:
    """(quarters agreeing in sign with `expr`, total quarters). A quarter at
    exactly 0 counts as neither for/against -- excluded from the numerator but
    kept in the denominator so it dilutes confidence honestly."""
    if not pq:
        return 0, 0
    want_pos = expr >= 0
    agree = sum(1 for v in pq.values() if (v > 0) == want_pos and v != 0)
    return agree, len(pq)


def _stat_block(values: List[float], ts_values: List[Any] = None) -> Dict[str, Any]:
    """The shared 'is this real?' computation over a list of R outcomes.

    Returns n, expR, CI (lo, hi), per-quarter dict, sign-consistency, and a
    plain verdict in {edge, unproven, loser, thin}. `ts_values` (alert
    timestamps aligned to `values`) drives the per-quarter sign; omit it to
    skip the per-quarter read (CI-only).
    """
    n = len(values)
    if n == 0:
        return {"n": 0, "expR": None, "ci": (None, None), "pq": {},
                "pq_agree": 0, "pq_total": 0, "verdict": "thin"}
    expr = round(sum(values) / n, 3)
    lo, hi = _bootstrap_ci(values)  # (None, None) when n < 5
    pq: Dict[str, float] = {}
    if ts_values is not None:
        by_q: Dict[str, List[float]] = {}
        for v, ts in zip(values, ts_values):
            q = _quarter_of(ts)
            if q is not None:
                by_q.setdefault(q, []).append(v)
        pq = {q: round(sum(x) / len(x), 3) for q, x in by_q.items() if x}
    agree, total = _pq_sign_consistency(pq, expr)

    # Verdict ladder. CI is the primary gate (RECOMMENDATIONS.md): an edge is
    # real only if its 95% CI clears 0. Thin sample (no CI) => never asserted.
    if lo is None:
        verdict = "thin"
    elif lo > 0:
        verdict = "edge"        # CI entirely above 0
    elif hi < 0:
        verdict = "loser"       # CI entirely below 0
    else:
        verdict = "unproven"    # CI straddles 0
    return {"n": n, "expR": expr, "ci": (lo, hi), "pq": pq,
            "pq_agree": agree, "pq_total": total, "verdict": verdict}




# ===========================================================================
# SHARED INSIGHT ENGINE
# ---------------------------------------------------------------------------
# One place that turns a slice of trades into a TRUSTWORTHY, ranked, plain-
# English read. Every "key insight" block in the email (by-pair column, MFE/MAE,
# pair×session narrative, second-look analysis) routes through here so nothing
# is ad-hoc. The discipline is always the same four steps:
#   1. RANK   — effect size = |avg-R gap vs base| × sqrt(N) (support-weighted).
#   2. GUARD  — a slice is only asserted as a rule if it clears the driver guard
#               (N ≥ DRIVER_MIN_N, |gap| ≥ DRIVER_MIN_EXR_GAP, per-quarter sign
#               holds). Below that it is "directional, thin" — surfaced, not
#               trusted. Same bar the driver engine uses, so verdicts agree.
#   3. CONNECT— the caller can pull the top guarded slice on several dimensions
#               at once and the narrative helpers stitch them ("worst pair is
#               GOLD, and its weakness concentrates in NY / weak breaks").
#   4. ACT    — every asserted insight ends in one recommendation, or an explicit
#               "no action — thin" when nothing cleared the guard.
# This is deterministic Python (no LLM): the numbers must be exact for SL/exit
# decisions, and it must reproduce byte-for-byte every run.
# ===========================================================================

def _slice_read(rows: List[Dict[str, Any]], base_exr: float,
                r_col: str = "r_realised") -> Optional[Dict[str, Any]]:
    """Core evaluator: score one slice of trades vs the book base-rate.

    Returns None for an empty slice, else a dict with the full guarded read:
      n, expR, gap (vs base), wr, ci, per-quarter sign, promoted (bool), and the
      component guard flags so a caller can explain WHY a slice was not trusted.
    `promoted` is the single source of "is this a rule?" — identical criteria to
    the driver engine.
    """
    rows = [r for r in rows if _is_real_filled(r)]
    n = len(rows)
    if n == 0:
        return None
    vals = [float(r.get(r_col) or 0.0) for r in rows]
    ts = [r.get("alert_ts") for r in rows]
    stat = _stat_block(vals, ts)
    gap = stat["expR"] - base_exr
    wins = sum(1 for v in vals if v > 0)
    losses = sum(1 for v in vals if v < 0)
    wr = (wins / (wins + losses) * 100) if (wins + losses) else None
    pq_ok = (stat["pq_total"] == 0
             or stat["pq_agree"] / stat["pq_total"] >= DRIVER_PQ_SIGN_FRAC)
    promoted = (n >= DRIVER_MIN_N and abs(gap) >= DRIVER_MIN_EXR_GAP and pq_ok)
    return {
        "n": n, "expR": stat["expR"], "gap": gap, "wr": wr, "ci": stat["ci"],
        "pq_agree": stat["pq_agree"], "pq_total": stat["pq_total"],
        "promoted": promoted,
        "n_ok": n >= DRIVER_MIN_N,
        "gap_ok": abs(gap) >= DRIVER_MIN_EXR_GAP,
        "pq_ok": pq_ok,
        "rank": abs(gap) * (n ** 0.5),
    }












def _runner_r(t: Dict[str, Any]) -> float:
    """R under TP1+runner reference policy: 50% closes at TP1 when TP1 hit,
    50% rides to whatever the TP2-ride reference did (r_if_exit_tp2 — original
    SL, BE after TP1). When TP1 never hits, the full position is the reference
    outcome.

    NOTE (2026-06-18): rides against `r_if_exit_tp2`, the TP2-ride REFERENCE —
    NOT `r_realised`, which is now the live TP1+BE@1R policy. Using r_realised
    here would collapse the runner to plain TP1.
    """
    if t.get("exit_reason") == "never_filled":
        return 0.0
    ref_tp2 = float(t.get("r_if_exit_tp2") or 0.0)
    bars_to_tp1 = t.get("bars_to_tp1")
    tp1_rr = float(t.get("tp1_rr") or 0.0)
    if bars_to_tp1 is not None and bars_to_tp1 >= 0:
        return 0.5 * tp1_rr + 0.5 * ref_tp2
    return ref_tp2


def _attach_runner_r(trades: List[Dict[str, Any]]) -> None:
    """Mutate trades in place to add `r_if_runner` (TP1+runner policy)."""
    for t in trades:
        if "r_if_runner" not in t:
            t["r_if_runner"] = round(_runner_r(t), 3)



# ---------------------------------------------------------------------------
# Killzone Alignment — OB session vs Fill session alignment buckets.
# Tests the SMC hypothesis: both-in-killzone trades > one-side > neither.
# ---------------------------------------------------------------------------

_ALIGNMENT_ORDER = ["Both", "OB only", "Fill only", "Neither"]

# Plain-English display labels for the four alignment buckets (2026-07). The raw
# keys stay in the data (and drive grouping); only the SHOWN label changes, so
# every alignment table across the email reads the same, unambiguous way.
_ALIGNMENT_LABELS = {
    "Both":      "OB formed + filled in killzone",
    "OB only":   "OB formed in killzone only",
    "Fill only": "Filled in killzone only",
    "Neither":   "Neither in killzone",
}


def _align_label(bucket: str) -> str:
    return _ALIGNMENT_LABELS.get(bucket, bucket)


def _killzone_alignment_table(trades: List[Dict[str, Any]], r_col: str
                              ) -> List[Dict[str, Any]]:
    """One row per alignment bucket: trades, win rate, avg R, total R."""
    filled = [t for t in trades if _is_real_filled(t)]
    if not filled:
        return []
    df = pd.DataFrame(filled)
    if "killzone_alignment" not in df.columns or r_col not in df.columns:
        return []
    out = []
    for bucket in _ALIGNMENT_ORDER:
        sub = df[df["killzone_alignment"] == bucket]
        if sub.empty:
            continue
        out.append({
            "bucket":       bucket,
            "trades":       int(len(sub)),
            "win_rate_pct": _win_rate(sub, r_col),
            "expectancy_r": round(float(sub[r_col].mean()), 3),
            "total_r":      round(float(sub[r_col].sum()), 3),
        })
    return out






# ---------------------------------------------------------------------------
# Trend alignment — did the zone trade WITH or AGAINST the H1 trend?
# The backtest fires counter-trend by design, so this answers the user's
# question directly: are against-trend setups losing consistently enough to
# filter out? trend_alignment is stamped per trade (h1_only_simulator) from the
# alert's h1_trend vs zone direction.
#
# One vocabulary now — the live path's. The backtest replay used to emit a
# separate dialect (with_trend / against_trend / no_trend); as of the Task 1
# parity fix (2026-07-05) both paths call smc_detector.derive_trend_alignment
# and emit with_trend / counter_trend / ambiguous. The old backtest keys are
# gone; this map only turns the shared values into display text.
# ---------------------------------------------------------------------------

_TREND_ORDER = ["With trend", "Against trend", "No trend"]

_TREND_LABEL_MAP = {
    "with_trend":    "With trend",
    "counter_trend": "Against trend",
    "ambiguous":     "No trend",
}







# ---------------------------------------------------------------------------
# Counterfactual "what if" filter analysis.
# For each filter dimension, compute aggregates under the filter and report
# delta vs baseline. Direct answer to: "if I had only taken trades that
# matched X, would I have made more money?"
# ---------------------------------------------------------------------------

_LOW_N_THRESHOLD = 10


def _cf_aggregate(sub: "pd.DataFrame", risk_usd: float) -> Dict[str, Any]:
    if sub.empty:
        return {"n": 0, "win_rate": None, "avg_r": 0.0, "total_pnl": 0.0}
    return {
        "n":         int(len(sub)),
        "win_rate":  _win_rate(sub, "r_realised"),  # wins/(wins+losses); None if all BE
        "avg_r":     float(sub["r_realised"].mean()),
        "total_pnl": float(sub["r_realised"].sum()) * risk_usd,
    }






def _counterfactual_dataframe(trades: List[Dict[str, Any]],
                              risk_usd: float) -> "pd.DataFrame":
    """Flat tabular form of the counterfactual analysis for Excel.

    Mirrors the row set built in _counterfactual_html so the Excel tab
    matches the email. Returns an empty frame if no filled trades or
    r_realised missing -- caller treats empty as 'skip the tab'.
    """
    filled = [t for t in trades if _is_real_filled(t)]
    if not filled:
        return pd.DataFrame()
    df = pd.DataFrame(filled)
    if "r_realised" not in df.columns:
        return pd.DataFrame()

    baseline = _cf_aggregate(df, risk_usd)
    if baseline["n"] == 0:
        return pd.DataFrame()

    rows: List[Dict[str, Any]] = [{
        "Section":      "Baseline",
        "Filter":       "All filled trades (no filter)",
        "Trades":       baseline["n"],
        "Win Rate %":   round(baseline["win_rate"], 1) if baseline["win_rate"] is not None else None,
        "Win Rate Delta (pp)": 0.0,
        "Avg R":        round(baseline["avg_r"], 3),
        "Total P&L":    round(baseline["total_pnl"], 2),
        "P&L Delta":    0.0,
        "Low N?":       "No",
    }]

    sections: List[Tuple[str, List[Tuple[str, "pd.Series"]]]] = []

    if "tp1_rr" in df.columns:
        tp1 = df["tp1_rr"].astype(float)
        sections.append(("TP1 R-multiple", [
            ("Restore live's 1.5R floor (TP1 R >= 1.5)", tp1 >= 1.5),
            ("Only TP1 R >= 2.0",        tp1 >= 2.0),
            ("Only TP1 R >= 2.5",        tp1 >= 2.5),
        ]))
    if "tp2_rr" in df.columns:
        tp2 = pd.to_numeric(df["tp2_rr"], errors="coerce")
        sections.append(("TP2 R-multiple", [
            ("Only TP2 R >= 2.0",        (tp2 >= 2.0) & tp2.notna()),
            ("Only TP2 R >= 2.5",        (tp2 >= 2.5) & tp2.notna()),
            ("Only TP2 R >= 3.0",        (tp2 >= 3.0) & tp2.notna()),
        ]))
    if "score" in df.columns:
        sc = pd.to_numeric(df["score"], errors="coerce")
        sections.append(("Score threshold", [
            ("Only score >= 3", sc >= 3),
            ("Only score >= 4", sc >= 4),
            ("Only score >= 5", sc >= 5),
        ]))
    if "confluences_present" in df.columns:
        conf_count = df["confluences_present"].astype(str).apply(
            lambda s: 0 if s in ("", "none", "nan") else len([c for c in s.split(",") if c])
        )
        sections.append(("Confluences", [
            ("Only >= 2 confluences", conf_count >= 2),
            ("Only >= 3 confluences", conf_count >= 3),
            ("Only >= 4 confluences", conf_count >= 4),
        ]))
    if "killzone_alignment" in df.columns:
        ka = df["killzone_alignment"]
        sections.append(("Killzone timing (OB-formation + fill candle)", [
            ("Only both candles in killzone",          ka == "Both"),
            ("Drop trades with neither candle in killzone", ka != "Neither"),
            ("Only trades filled in a killzone (both or fill-only)",
             ka.isin(["Both", "Fill only"])),
        ]))
    if "fill_session" in df.columns:
        fs = df["fill_session"]
        sections.append(("Fill session (session the order filled in)", [
            ("Only fills in London", fs == "London"),
            ("Only fills in NY",     fs == "NY"),
            ("Only fills in Asia",   fs == "Asia"),
        ]))
    if "fill_ts" in df.columns:
        dow = pd.to_datetime(df["fill_ts"], errors="coerce", utc=True).dt.day_name()
        sections.append(("Day of week", [
            ("Skip Monday",  dow != "Monday"),
            ("Skip Friday",  dow != "Friday"),
            ("Only Tue-Thu", dow.isin(["Tuesday", "Wednesday", "Thursday"])),
        ]))
    if "pd_alignment" in df.columns:
        pa = df["pd_alignment"]
        sections.append(("PD-array alignment", [
            ("PD-aligned (long+discount / short+premium)", pa == "aligned"),
            ("PD-counter (long+premium / short+discount)", pa == "counter"),
        ]))
    elif "pd_zone" in df.columns:
        pdz = df["pd_zone"]
        sections.append(("PD zone (legacy)", [
            ("Only Discount",     pdz == "discount"),
            ("Only Premium",      pdz == "premium"),
        ]))

    for sect_label, filters in sections:
        for label, mask in filters:
            sub = df[mask.fillna(False)]
            agg = _cf_aggregate(sub, risk_usd)
            # Guard columns (§11.4): the counterfactual leaves the body but only
            # enters the email via Act 5, which recomputes with _stat_block. Ship
            # the same guard here so the Excel row is self-judging — a filter is
            # only worth acting on if its remainder CI clears zero and the sign
            # holds across quarters.
            sub_rows = sub.to_dict("records")
            svals = [float(r.get("r_realised") or 0.0) for r in sub_rows
                     if _is_real_filled(r)]
            sts = [r.get("alert_ts") for r in sub_rows if _is_real_filled(r)]
            gstat = _stat_block(svals, sts)
            glo, ghi = gstat["ci"]
            rows.append({
                "Section":      sect_label,
                "Filter":       label,
                "Trades":       agg["n"],
                "Win Rate %":   round(agg["win_rate"], 1) if agg["win_rate"] is not None else None,
                "Win Rate Delta (pp)": (round(agg["win_rate"] - baseline["win_rate"], 1)
                                        if agg["win_rate"] is not None
                                        and baseline["win_rate"] is not None else None),
                "Avg R":        round(agg["avg_r"], 3) if agg["n"] else None,
                "Total P&L":    round(agg["total_pnl"], 2) if agg["n"] else None,
                "P&L Delta":    round(agg["total_pnl"] - baseline["total_pnl"], 2) if agg["n"] else None,
                # Guard columns.
                "CI lo":        glo,
                "CI hi":        ghi,
                "Quarters (sign held)": (f"{gstat['pq_agree']}/{gstat['pq_total']}"
                                         if gstat["pq_total"] else None),
                "CI cleared":   bool(glo is not None and glo > 0),
                "Low N?":       "Yes" if 0 < agg["n"] < _LOW_N_THRESHOLD else "No",
            })

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Exit-policy comparison: three policies + per-pair + per-session.
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Exit-recipe display labels + the LIVE baseline key. Shared by the recipe table,
# the winner rule, and the leak-table exit-fix join.
_EXIT_RECIPE_LABELS = {
    "baseline_liqTP_be1.0":  "TP1 + BE@1R (LIVE)",
    "B_be0.5":               "TP1, BE@0.5R",
    "B_be0.7":               "TP1, BE@0.7R",
    "C_fullTP_0.5R":         "Fixed TP 0.5R",
    "C_fullTP_1.0R":         "Fixed TP 1.0R",
    "C_fullTP_1.5R":         "Fixed TP 1.5R",
    "C_fullTP_2.0R":         "Fixed TP 2.0R",
    "D_partial50_1R_runLiq": "Partial 50% @1R, runner to TP1",
    # 3-target ladder (2026-07-17): zone (TP1) / wick (pool A magnet) / next pool.
    "E_zoneTP_be1.0":              "Zone TP (TP1) + BE@1R",
    "E_wickTP_be1.0":              "Wick TP + BE@1R",
    "E_partial_zone_then_wick":    "50% zone, 50% wick",
    "E_partial_zone_wick_nextpool": "Thirds: zone / wick / next pool",
}
_EXIT_BASELINE_KEY = "baseline_liqTP_be1.0"

# Dimensions the driver/mined engine screens. Categorical => value as-is;
# continuous => tertile buckets (low/mid/high) from this run's distribution.
_DRIVER_CATEGORICAL = [
    ("pair",                "Pair"),
    ("session",             "Alert session"),
    ("ob_session",          "OB session"),
    ("fill_session",        "Fill session"),
    ("bos_tag",             "Break type (BOS/CHoCH)"),
    ("bos_tier",            "Structure tier"),
    ("break_tier",          "Break-quality tier"),
    ("fvg_state",           "FVG state"),
    ("killzone_alignment",  "Killzone alignment"),
    ("trend_alignment",     "Trend alignment"),
    ("pd_zone",             "PD zone"),
    ("pd_alignment",        "PD alignment"),
    ("reversed_from_extreme", "CHoCH from extreme"),
    ("dow",                 "Day of week"),
]
_DRIVER_CONTINUOUS = [
    ("break_close_atr",   "Break close (ATR)"),
    ("break_body_atr",    "Break body (ATR)"),
    ("ob_range_atr",      "OB range / stop (ATR)"),
    ("fvg_size_atr",      "FVG size (ATR)"),
    ("impulse_leg_to_extreme_atr", "Impulse leg to extreme (ATR)"),
    ("ob_age_h1_bars",    "OB age (H1 bars)"),
    ("tp1_rr",            "TP1 distance (R)"),
    ("score",             "Confidence score"),
    ("atr_at_ob",         "ATR at OB (volatility)"),
]
_DOW_NAMES = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]


def _driver_buckets(filled: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """Build every candidate bucket across all screened dimensions.

    One bucket = {dim_label, value_label, rows}. Continuous features are split
    into low/mid/high tertiles on the non-null values present this run. Buckets
    with zero rows are skipped (they can't be a candidate). Day-of-week is
    derived here from alert_ts so it needs no new logged column."""
    df = pd.DataFrame(filled)
    if df.empty:
        return []
    # Derive day-of-week once (Mon..Sun) from the alert timestamp.
    if "alert_ts" in df.columns:
        _dt = pd.to_datetime(df["alert_ts"], utc=True, errors="coerce")
        df["dow"] = _dt.dt.dayofweek.map(
            lambda i: _DOW_NAMES[int(i)] if pd.notna(i) else None)

    out: List[Dict[str, Any]] = []

    for col, label in _DRIVER_CATEGORICAL:
        if col not in df.columns:
            continue
        for val, sub in df.groupby(df[col].astype("object")):
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            rows = sub.to_dict("records")
            if rows:
                out.append({"dim": label, "value": str(val), "rows": rows})

    for col, label in _DRIVER_CONTINUOUS:
        if col not in df.columns:
            continue
        ser = pd.to_numeric(df[col], errors="coerce")
        valid = ser.dropna()
        if len(valid) < DRIVER_MIN_N * 2 or valid.nunique() < 3:
            continue  # not enough spread to bucket meaningfully
        try:
            lo_q, hi_q = valid.quantile([1 / 3, 2 / 3]).tolist()
        except Exception:
            continue
        if lo_q == hi_q:
            continue
        bands = [
            (f"low (≤{lo_q:.2f})",          ser <= lo_q),
            (f"mid ({lo_q:.2f}–{hi_q:.2f})", (ser > lo_q) & (ser <= hi_q)),
            (f"high (>{hi_q:.2f})",         ser > hi_q),
        ]
        for band_label, mask in bands:
            sub = df[mask.fillna(False)]
            rows = sub.to_dict("records")
            if rows:
                out.append({"dim": label, "value": band_label, "rows": rows})

    return out


def _driver_two_way(filled: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """The handful of 2-way interactions a univariate pass can't see. Only the
    pairings SMC actually cares about: pair×session and PD-zone×break-type."""
    df = pd.DataFrame(filled)
    if df.empty:
        return []
    out: List[Dict[str, Any]] = []
    combos = [
        (("pair", "session"),  "Pair × session"),
        (("pd_zone", "bos_tag"), "PD zone × break type"),
    ]
    for (a, b), label in combos:
        if a not in df.columns or b not in df.columns:
            continue
        for (va, vb), sub in df.groupby([df[a].astype("object"),
                                         df[b].astype("object")]):
            if va is None or vb is None:
                continue
            rows = sub.to_dict("records")
            if len(rows) >= DRIVER_MIN_N:   # 2-way only worth showing if not trivially thin
                out.append({"dim": label, "value": f"{va} / {vb}", "rows": rows})
    return out


# SMC methodology priors — what a veteran EXPECTS each slice to do to expectancy.
# Keyed by (dim_label, matcher). Value = +1 if the slice SHOULD help (raise avg R),
# -1 if it SHOULD hurt. A row "conflicts with SMC" when the data's direction
# (gap sign) opposes this prior. Conflicts are flagged, never hidden — per the
# project rule, data-vs-SMC disagreement is a discussion point, not a conclusion.
def _smc_expected_direction(dim: str, value: str):
    """Return +1 (SMC expects this slice to HELP), -1 (SMC expects it to HURT),
    or None (no strong SMC prior — do not flag). dim = human label, value = raw
    bucket label as shown in the table."""
    d = (dim or "").lower()
    v = (value or "").lower()

    # Trend: with-trend is the draw on liquidity (help); against-trend fights it.
    if "trend alignment" in d:
        if "with trend" in v:
            return +1
        if "against trend" in v:
            return -1
        return None
    # FVG size: a bigger imbalance is real displacement (help); a tiny gap is noise.
    if "fvg size" in d:
        if v.startswith("high"):
            return +1
        if v.startswith("low"):
            return -1
        return None
    # FVG state: fresh gap unmitigated (help); stale = already discharged (hurt).
    if "fvg state" in d:
        if "fresh" in v:
            return +1
        if "stale" in v:
            return -1
        return None
    # Break quality / displacement: a stronger break past the level is conviction.
    if "break-quality tier" in d or "break close" in d or "break body" in d:
        if v.startswith("high") or "strong" in v:
            return +1
        if v.startswith("low") or "weak" in v:
            return -1
        return None
    # Killzone alignment: in-killzone formation/fill is the institutional window.
    if "killzone alignment" in d:
        if v == "both":
            return +1
        if v == "neither":
            return -1
        return None
    # PD array: aligned side (long in discount / short in premium) is the premise.
    if "pd alignment" in d:
        if "aligned" in v:
            return +1
        if "counter" in v:
            return -1
        return None
    return None




# ---------------------------------------------------------------------------
# Where the edge leaked — replaces the old "losing trades" bullet block.
# Same data as before plus MFE-leak, OB-age distribution, time-in-trade.
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# MFE / MAE — the excursion book. Drives SL placement and exit tuning.
#   mfe_r : best unrealised R the trade ever showed (peak). >= 0.
#   mae_r : worst unrealised R the trade ever showed (deepest drawdown). <= 0.
# The actionable questions:
#   • Winners: how much of the peak do we GIVE BACK (mfe - booked)? Big give-back
#     = the exit leaves money on the table -> a target/trail worth revisiting.
#   • Losers: how deep is the average drawdown before SL (mae)? And how many went
#     meaningfully GREEN first ("nearly worked") before reversing -> the SL / BE
#     placement leak. If losers routinely show +1R before dying, a break-even or
#     partial rule would have rescued real R.
# ---------------------------------------------------------------------------



def _same_bar_resolution_html(trades: List[Dict[str, Any]]) -> str:
    filled = [t for t in trades if _is_real_filled(t)]
    n = len(filled)
    if n == 0:
        return "<p style='color:#888;'>No resolved trades this period.</p>"
    fill_bar = [t for t in filled if int(t.get("bars_to_exit") or 0) == 0]
    losers = [t for t in filled if (t.get("r_realised") or 0) < 0]
    # Loser whose MFE reached TP1's price -> TP1 was touched in-bar but booked SL.
    tp1_touched = [t for t in losers
                   if (t.get("mfe_r") or 0) >= (t.get("tp1_rr") or 1e9)]
    pct = len(fill_bar) / n * 100
    tone = "#e74c3c" if pct >= 25 else ("#f39c12" if pct >= 10 else "#27ae60")
    return (
        f"<p style='font-size:13px;line-height:1.6;'>"
        f"<b style='color:{tone};'>{len(fill_bar)} of {n}</b> resolved trades "
        f"({pct:.0f}%) exited on the <b>fill bar</b> — the same H1 candle that "
        f"filled the limit also hit SL (and sometimes TP1). On H1 we cannot tell "
        f"whether TP or SL printed first, so these are booked <b>SL</b> "
        f"(pessimistic, no look-ahead). Of the losers, "
        f"<b>{len(tp1_touched)}</b> had TP1's price touched in-bar yet booked a "
        f"full −1R. Read the headline knowing this slice is H1-resolution-limited, "
        f"not clean directional losses — a lower-timeframe fill model would "
        f"reclassify some.</p>"
    )


# ---------------------------------------------------------------------------
# By-pair and by-session flat tables (replaces the sparse 2D grid)
# ---------------------------------------------------------------------------

_SESSION_ORDER = ["Asia", "London", "NY", "Other"]


























# ---------------------------------------------------------------------------
# Confluence per pair
# ---------------------------------------------------------------------------

_CONFLUENCE_COLS = {
    "FVG":          ("fvg_present",   lambda s: s == True),
    "Sweep":        ("sweep_present", lambda s: s == True),
    "Kill zone":    ("killzone_pts",  lambda s: s > 0),
    "Freshness":    ("freshness_pts", lambda s: s > 0),
    "Structure":    ("structure_pts", lambda s: s > 0),
}





# ---------------------------------------------------------------------------
# News blackout report section
# ---------------------------------------------------------------------------



# ---------------------------------------------------------------------------
# Loss pattern analysis
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Structure event performance breakdown (Major vs Minor, BOS vs CHoCH)
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Trade validation
# ---------------------------------------------------------------------------

def _validate_trades(trades: List[Dict[str, Any]]) -> List[str]:
    """Check structural invariants. Returns a list of violation strings (empty = clean)."""
    violations = []
    filled = [t for t in trades if _is_real_filled(t)]
    for t in filled:
        pair      = t.get("pair", "?")
        ts        = t.get("alert_ts", "?")
        direction = t.get("direction", "")
        entry     = t.get("entry")
        sl        = t.get("sl_initial")
        tp1       = t.get("tp1")
        tp2       = t.get("tp2")
        r         = t.get("r_realised", 0)
        reason    = t.get("exit_reason", "")

        if entry is None or sl is None or tp1 is None:
            violations.append(f"{pair} @ {ts}: missing entry/SL/TP1")
            continue

        if direction == "bullish":
            if sl >= entry:
                violations.append(f"{pair} @ {ts}: SL ({sl}) ≥ entry ({entry}) for LONG")
            if tp1 <= entry:
                violations.append(f"{pair} @ {ts}: TP1 ({tp1}) ≤ entry ({entry}) for LONG")
            if tp2 is not None and tp2 <= tp1:
                violations.append(f"{pair} @ {ts}: TP2 ({tp2}) ≤ TP1 ({tp1}) for LONG")
        elif direction == "bearish":
            if sl <= entry:
                violations.append(f"{pair} @ {ts}: SL ({sl}) ≤ entry ({entry}) for SHORT")
            if tp1 >= entry:
                violations.append(f"{pair} @ {ts}: TP1 ({tp1}) ≥ entry ({entry}) for SHORT")
            if tp2 is not None and tp2 >= tp1:
                violations.append(f"{pair} @ {ts}: TP2 ({tp2}) ≥ TP1 ({tp1}) for SHORT")

        # TP/SL exit should have correct sign
        if reason == "sl" and r > 0:
            violations.append(f"{pair} @ {ts}: exit=SL but r_realised={r:.2f} (positive — unexpected)")
        if reason in ("tp1", "tp2") and r < 0:
            violations.append(f"{pair} @ {ts}: exit={reason} but r_realised={r:.2f} (negative — unexpected)")

    return violations


def _validation_html(trades: List[Dict[str, Any]]) -> str:
    violations = _validate_trades(trades)
    if not violations:
        n_prox = len([t for t in trades
                      if _is_real_filled(t)
                      and t.get("entry_zone") == "proximal"])
        return (f"<p style='color:#27ae60;'>✓ All {n_prox} filled proximal trade "
                f"rows passed validation — entry/SL/TP levels are correctly "
                f"ordered, and exit outcomes match exit reasons.</p>")
    items = "".join(f"<li style='color:#e74c3c;font-family:monospace;font-size:12px;'>{v}</li>"
                    for v in violations)
    return (f"<p style='color:#e74c3c;'><b>⚠ {len(violations)} validation issue(s) found:</b></p>"
            f"<ul>{items}</ul>"
            f"<p style='font-size:12px;color:#888;'>These indicate calculation errors in the simulator. "
            f"Do not draw trading conclusions until they are resolved.</p>")


# ---------------------------------------------------------------------------
# Zone register (for Excel tab)
# ---------------------------------------------------------------------------

def _build_zone_register_df(trades: List[Dict[str, Any]]) -> pd.DataFrame:
    """One row per OB-touch event, showing both entry zone outcomes side by side."""
    if not trades:
        return pd.DataFrame()
    df = pd.DataFrame(trades)
    # Key = unique OB touch event: pair + alert_ts (each alert produces 2 rows)
    if "alert_ts" not in df.columns or "pair" not in df.columns:
        return pd.DataFrame()

    _exit_labels = {
        "sl": "Stop Loss Hit", "tp1": "TP1 Hit", "tp2": "TP2 Hit",
        "timeout": "Time Limit (48h)", "window_end": "End of Window",
        "sl_collision": "SL+TP Same Bar", "never_filled": "Never Filled",
        "friday_flat": "Weekend Flat (Fri)",
    }

    rows = []
    for (pair, alert_ts), grp in df.groupby(["pair", "alert_ts"]):
        prox = grp[grp["entry_zone"] == "proximal"].iloc[0].to_dict() if not grp[grp["entry_zone"] == "proximal"].empty else {}

        def _v(d, k, default=""):
            return d.get(k, default)

        news_blocked = bool(_v(prox, "news_blocked"))
        news_event_title    = _v(prox, "news_event_title")
        news_event_currency = _v(prox, "news_event_currency")
        news_event_source   = _v(prox, "news_event_source")
        news_event_ts       = _v(prox, "news_event_ts")
        kz_blocked = bool(_v(prox, "killzone_blocked"))
        ist_blocked_zr = bool(_v(prox, "ist_blocked"))

        # Prefer fill_session (when the trade was actually live). Fall back
        # to alert-hour session for never-filled OBs. Same logic for DOW.
        prox_fill_ts = _v(prox, "fill_ts")
        zr_dow = _day_of_week(prox_fill_ts or alert_ts)
        rows.append({
            "Proximal ID":             _v(prox, "setup_id"),
            "Pair":                    pair,
            "OB Candle (IST)":         _to_ist_str(_v(prox, "ob_timestamp")),
            "Scan / Alert Time (IST)": _to_ist_str(alert_ts),
            "Direction":               "Long" if _v(prox, "direction") == "bullish" else "Short",
            "Structure Event":         _v(prox, "bos_tag"),
            "Structure Tier":          _v(prox, "bos_tier"),
            "OB Age (H1 bars)":        _v(prox, "ob_age_h1_bars"),
            "Setup Score":             _v(prox, "score"),
            "Fill Session":            _v(prox, "fill_session"),
            "OB Session":              _v(prox, "ob_session"),
            "Killzone Alignment":      _v(prox, "killzone_alignment"),
            "Day of Week":             zr_dow,
            "Entry Price (Proximal)":  _v(prox, "entry"),
            "Stop Loss (raw)":         _v(prox, "sl_raw"),
            "Stop Loss":               _v(prox, "sl_initial"),
            "Take Profit 1":           _v(prox, "tp1"),
            "Take Profit 2":           _v(prox, "tp2"),
            "TP1 Reward:Risk":         _v(prox, "tp1_rr"),
            "TP2 Reward:Risk":         _v(prox, "tp2_rr"),
            "Proximal Filled?":        "Yes" if _v(prox, "exit_reason") != "never_filled" and prox else "No",
            "Proximal Outcome":        _exit_labels.get(_v(prox, "exit_reason"), _v(prox, "exit_reason")),
            "Proximal R":              _v(prox, "r_realised"),
            "Proximal Dollar P&L":     round(float(_v(prox, "pnl_usd", 0) or 0), 0) if prox else "",
            "FVG Present":             "Yes" if _v(prox, "fvg_present") else "No",
            "Sweep Present":           "Yes" if _v(prox, "sweep_present") else "No",
            "Confluences Active":      _v(prox, "confluences_present"),
            # News blackout audit columns. If "Yes" in News Blocked, this
            # row was excluded from every aggregate metric in the email.
            "News Blocked":            "Yes" if news_blocked else "No",
            "News Event":              news_event_title,
            "News Currency":           news_event_currency,
            "News Source":             news_event_source,
            "News Event Time (UTC)":   news_event_ts,
            "Killzone Blocked":        "Yes" if kz_blocked else "No",
            "IST Blocked":             "Yes" if ist_blocked_zr else "No",
        })

    return pd.DataFrame(rows)


def _day_of_week(ts_str: str) -> str:
    try:
        return pd.Timestamp(ts_str).day_name()[:3]
    except Exception:
        return ""


# ---------------------------------------------------------------------------
# CSV (machine-readable, column names unchanged — used by aggregate_runs.py)
# ---------------------------------------------------------------------------

# Events calendar for the automated news enrichment (Part D, 2026-07-16). The
# ONE maintenance seam: re-scrape (backtest/data/ff_calendar_scraper.py) before
# each baseline so this file covers the run window (see CANONICAL.md checklist).
_NEWS_EVENTS_CSV = Path(__file__).resolve().parent / "data" / "ff_calendar_2007_2026.csv"


def _enrich_news_columns(csv_path: Path) -> None:
    """Stamp the 5 news columns onto trades.csv in place (Part D, 2026-07-16).

    Runs at the END of the report build so news_fill/news_open land in the same
    artifact as the rest — no separate command. Fail-LOUD offline: this is a
    report build, not the live alert path, so a missing/short events file must
    RAISE rather than silently write empty news columns (CLAUDE.md's "no raise
    in the live alert path" does not apply here). Determinism/idempotence are
    covered by test_news_enrichment.py; a fresh run pairs true-UTC candles
    (import_mt5 Part B) with the no-double-correction enrichment (Part C).
    """
    import backtest.news_enrichment as ne  # local import: offline tooling only

    if not _NEWS_EVENTS_CSV.exists():
        raise FileNotFoundError(
            f"news events file missing: {_NEWS_EVENTS_CSV} — re-scrape with "
            "backtest/data/ff_calendar_scraper.py before the run (Part E seam)")
    trades = pd.read_csv(csv_path, low_memory=False)
    if "fill_ts" not in trades.columns:  # e.g. the "no trades this run" stub
        return
    events = ne.load_events(str(_NEWS_EVENTS_CSV))
    # Coverage guard: the events file must span the filled-trade window, else the
    # enrichment would silently NULL the uncovered tail. Raise loud instead.
    fills = pd.to_datetime(trades["fill_ts"], utc=True,
                           format="mixed", errors="coerce").dropna()
    if len(fills):
        need_lo, need_hi = fills.min(), fills.max()
        have_lo, have_hi = events["utc"].min(), events["utc"].max()
        if need_lo < have_lo or need_hi > have_hi:
            raise ValueError(
                "news events file does not cover the run window "
                f"(trades {need_lo}..{need_hi}, events {have_lo}..{have_hi}); "
                "re-scrape ff_calendar_scraper.py (Part E seam)")
    res, _stats = ne.enrich(trades, events)
    # Health check (Part D): the 5 columns must be present and news_fill must be
    # populated (not all-null) — an all-null news_fill means the join silently
    # matched nothing (bad events file or a wholesale coverage miss). Only assert
    # when there are filled trades to flag; a run with zero fills legitimately
    # has an all-null column.
    missing = [c for c in ne.NEW_COLS if c not in res.columns]
    if missing:
        raise AssertionError(f"news enrichment dropped columns: {missing}")
    if len(fills) and not res["news_fill"].notna().any():
        raise AssertionError(
            "news_fill is entirely null despite filled trades — events join "
            "matched nothing (check ff_calendar coverage / clock fix)")
    res.to_csv(csv_path, index=False)


def _trades_csv(trades: List[Dict[str, Any]], path: Path) -> None:
    if not trades:
        pd.DataFrame([{"info": "no trades this run"}]).to_csv(path, index=False)
        return
    # setup_id is stamped upstream in write_h1_only_report (on trades_all)
    # BEFORE any consumer runs, so the same T#### appears in trades.csv, both
    # Excel Trade tabs, the Zone Register, and the email vet-review table.
    for _t in trades:
        # Self-describing headline membership. Stamped from the ONE eligibility
        # rule (_headline_exclusion) so any consumer of this file can reproduce
        # the email headline directly: sum(pnl_usd where eligible_for_headline)
        # == summary headline. Without these columns the rule lives only in
        # reporting and the file silently over-counts unresolved/audit rows.
        _excl = _headline_exclusion(_t)
        _t["eligible_for_headline"] = (_excl == "")
        _t["headline_exclusion"] = _excl
    front_cols = [
        "setup_id",
        "pair", "alert_ts", "fill_ts", "exit_ts", "session",
        "direction", "event", "entry_zone",
        # entry/tp1/tp2 = SPREAD-PLACED execution levels; *_raw = pre-placement OB/
        # zone geometry (2026-07-22 spread shift audit, mirrors sl_raw/sl_initial).
        "entry", "entry_raw", "sl_raw", "sl_initial",
        "tp1", "tp1_raw", "tp2", "tp2_raw", "tp1_rr", "tp2_rr",
        # TP-placement audit (2026-07-15): tp1/tp2 above are the ZONE-EDGE
        # (traded) levels; these expose the raw swing wick they replaced, its RR,
        # and the source ("zone" opposing-OB edge used | "wick" fallback).
        "tp1_wick", "tp1_wick_rr", "tp1_zone_source",
        "tp2_wick", "tp2_zone_source",
        # 3-TARGET LADDER (backtest triple mode, 2026-07-17). Unambiguous names —
        # `tp2` above still means "next pool". tp_wick = pool-A liquidity wick
        # (buffered); tp_nextpool = the runner (next different pool); tp2_collapsed_
        # to_tp1 = wick landed on TP1; tp_targets = "triple" marker. Consumed by the
        # exit recipes via walk_multileg "tp_wick"/"tp_nextpool" specs.
        "tp_wick", "tp_wick_rr",
        "tp_nextpool", "tp_nextpool_rr", "tp_nextpool_zone_source",
        "tp2_collapsed_to_tp1", "tp_targets",
        "exit_reason", "exit_price",
        "r_realised", "r_if_exit_tp1", "r_if_exit_tp2", "pnl_usd",
        # Headline membership: True rows sum to the email headline. When False,
        # headline_exclusion names why (unresolved:*, below_score_floor,
        # ist_blocked). Stamped from the single _headline_exclusion rule.
        "eligible_for_headline", "headline_exclusion",
        "mfe_r", "mae_r", "r_capture_ratio",
        "sl_bar_was_sweep", "sl_swept_then_tp1",
        "sl_wick_depth_atr",
        # Outcome-time exit-track columns (2026-07-08; NEVER entry features).
        "sl_max_adverse_after_sweep_atr", "bars_sl_to_tp1_touch",
        "sl_recovered_to_entry", "sl_distance_atr",
        "sl_dist_atr_at_alert", "tp_dist_atr_at_alert",
        "bars_to_exit", "bars_to_tp1", "bars_to_tp2",
        "ob_to_fill_hours", "bars_break_to_pullback",
        "ob_age_h1_bars", "pd_zone",
        # Reversal book: exact CHoCH-origin-in-extreme flag (raw) + derived bool.
        "reversal_pct", "reversed_from_extreme",
        "score", "structure_pts", "sweep_pts", "fvg_pts",
        "freshness_pts", "killzone_pts", "confluences_present",
        "sl_collision", "model", "ob_timestamp", "bos_tag", "bos_tier",
        # Bars the true break candle sits before the confirmation candle (event-
        # candle fix, 2026-07-09). 0 = clean single-candle break. Immutable event
        # fact; audits the candle shift the fix introduced.
        "event_candle_delta",
        "bos_verdict", "fvg_present", "fvg_mitigation", "sweep_present", "ob_touches",
        # Setup-geometry features (ATR-normalized) — edge-discovery engine inputs.
        "break_close_atr", "break_body_atr", "break_excess", "break_tier",
        # is_mss (2026-07-21): descriptive MSS label — CHoCH break body >= the
        # data-derived MSS_BODY_ATR_MULT. NOT a proven separator (see ledger), no
        # gate/score. Sits by the break_* group it is derived from.
        "is_mss",
        "ob_range_atr", "fvg_size_atr", "impulse_leg_to_extreme_atr", "atr_at_ob", "atr_at_fill",
        # Walk-back geometry (A3, DECISION_GUARDRAILS.md) — logging only, no
        # gate. None for legacy zones built before this change.
        "ob_body_ratio", "ob_walkback_depth",
        # Choppiness Index on the alert's server trading day — daily trend-vs-
        # range regime at the alert bar. Observe-only, gates nothing.
        "chop_at_alert",
        # News blackout audit columns. INFORMATIONAL ONLY — news never gates
        # (the one eligibility rule is _headline_exclusion above).
        "news_blocked", "news_event_title", "news_event_currency",
        "news_event_source", "news_event_ts",
        # IST blackout audit columns. ist_blocked=True means this alert
        # fell outside the user's IST trading window and was excluded
        # from aggregates (live system would have suppressed it).
        "ist_blocked", "alert_utc_hour",
        "h1_trend", "trend_alignment", "trend_pd_agree",
        # Structure signals (STRUCTURE_SIGNALS_SPEC) — edge-discovery inputs, no
        # gate/email change yet. S2 ranging/pending-flip state at alert; S3 leg
        # leg extreme (extreme + clipped are audit support); S4 broken-wall PD
        # flags at OB formation. None per each column's rule.
        "structure_ranging_at_alert", "flip_pending_at_alert",
        "flip_pending_dir_at_alert",
        "leg_extreme_at_alert", "leg_extreme_clipped",
        "dr_ceiling_broken_at_ob", "dr_floor_broken_at_ob",
        # PD/PW liquidity pools (DAILY_BIAS_V4_SPEC §1.3) — observation only,
        # stamped at alert from strictly-prior bars. One source:
        # pool_builder.POOL_FEATURE_COLUMNS via _pool_features_at_alert.
        "day_state_at_fill",
        "pdh_status_at_fill", "pdl_status_at_fill",
        "pwh_status_at_fill", "pwl_status_at_fill",
        "dist_next_pool_above_atr", "dist_next_pool_below_atr",
        "next_pool_above_tier", "next_pool_below_tier",
        "trade_toward_pool", "last_sweep_age_h1", "last_sweep_tier",
        # EQH/EQL equal-level clusters (2026-07-14) — observation only,
        # stamped at alert from strictly-prior bars. One source:
        # eq_pools.EQ_FEATURE_COLUMNS via _eq_features_at_alert.
        "eqh_above_dist_atr", "eqh_above_size",
        "eql_below_dist_atr", "eql_below_size",
        "eq_trade_toward",
        "eq_sl_gap_atr", "eq_sl_at_risk",
        "eq_last_sweep_age_h1", "eq_last_sweep_side",
        "eq_intact_above_count", "eq_intact_below_count",
        # Sweep v2 (rebuilt pool-anchored sweep, 2026-07-18) — observation
        # only, re-labelled from the FROZEN ob['sweep_v2'] snapshot stamped at
        # OB build. One source: liquidity_sweep.SWEEP2_FEATURE_COLUMNS via
        # _sweep2_features. Legacy sweep_pts / sweep_present stay unchanged
        # (score parity); superseded for ANALYSIS by these columns.
        "sweep2_present", "sweep2_tier", "sweep2_level",
        "sweep2_pierce_atr", "sweep2_rejection_ratio", "sweep2_follow_atr",
        "sweep2_pools_swept", "sweep2_rn_aligned", "sweep2_rn_dist_atr",
        "sweep2_eq_size", "sweep2_age_at_alert_h1", "sweep2_tiers_checked",
        # SETUP-LIQ (this trade's own stop/target vs swing liquidity, 2026-07-20)
        # — 6 columns from setup_liq._setup_liq_features. Reads 1 & 2 anchor on
        # the trade SL/TP1 (level-calc, not OB-build-frozen); Read 3.2 is the
        # leg-extreme-was-a-sweep payload scalar. See TRUTH_LEDGER setup_liq_* rows.
        "setup_liq_stop_present", "setup_liq_stop_offset_atr", "setup_liq_stop_tier",
        "setup_liq_tp_present", "setup_liq_tp_offset_atr",
        "setup_liq_legextreme_swept",
        # SESSION H/L sweep+break (SESSION_SWEEP_STUDY_SPEC, 2026-07-21) — 3 columns
        # from session_levels._session_level_features_at_alert. DST-honest session
        # windows resolved per candle; sweep/break via reused pool_builder.pool_status.
        # ALERT-time, observation only, pair-specific study. See TRUTH_LEDGER
        # session_level_* rows.
        "session_level_event", "session_level_which", "session_level_side",
        "session_level_pair_relevant",
    ]
    df = pd.DataFrame(trades)
    cols_present = [c for c in front_cols if c in df.columns]
    rest = [c for c in df.columns if c not in cols_present]
    df[cols_present + rest].to_csv(path, index=False)


# ---------------------------------------------------------------------------
# Excel (human-readable — plain English column names, formatted)
# ---------------------------------------------------------------------------

_EXCEL_COL_NAMES = {
    # A: identifiers
    "setup_id":          "Setup ID",
    "pair":              "Currency Pair",
    # B–F: all timestamps, grouped for easy chart cross-reference
    "ob_time_ist":       "OB Candle (IST)",
    "event_time_ist":    "Event Candle (IST)",
    "alert_time_ist":    "Scan / Alert Time (IST)",
    "fill_time_ist":     "Entry Fill (IST)",
    "sl_hit_time_ist":   "SL Hit (IST)",
    "tp_fill_time_ist":  "TP Fill (IST)",
    # trade context
    "direction":         "Direction",
    "fill_session":      "Fill Session",
    "ob_session":        "OB Session",
    "killzone_alignment": "Killzone Alignment",
    "h1_trend":          "H1 Trend",
    "trend_alignment":   "Trend Alignment",
    "entry_zone":        "Entry Type",
    # levels
    "entry":             "Entry Price",
    "entry_raw":         "Entry Price (raw)",
    "sl_initial":        "Stop Loss",
    "tp1":               "Take Profit 1",
    "tp1_raw":           "Take Profit 1 (raw)",
    "tp2":               "Take Profit 2",
    "tp2_raw":           "Take Profit 2 (raw)",
    "tp1_rr":            "TP1 Reward:Risk",
    "tp2_rr":            "TP2 Reward:Risk",
    # PD array
    "pd_zone":           "PD Zone",
    "pd_alignment":      "PD Alignment",
    "pd_pct":            "PD Array % (entry)",
    "reversal_pct":          "CHoCH Origin In Extreme (raw)",
    "reversed_from_extreme": "Reversed From Extreme (CHoCH)",
    # outcome
    "exit_reason":       "How Trade Closed",
    "exit_price":        "Exit Price",
    "r_realised":        "R Achieved (LIVE: TP1+BE@1R)",
    "pnl_usd":           "Dollar P&L (LIVE: TP1+BE@1R)",
    "r_if_exit_tp1":     "R if Closed at TP1",
    "pnl_usd_tp1":       "Dollar P&L (TP1-only)",
    "r_if_runner":       "R if TP1 + Runner",
    "pnl_usd_runner":    "Dollar P&L (TP1+Runner)",
    "mfe_r":             "Best Price Reached (R)",
    "mae_r":             "Worst Price Reached (R)",
    "bars_to_exit":      "Hours Held",
    "bars_to_tp1":       "Hours to TP1 (-1 if never)",
    # setup quality
    "score":             "Setup Score (0–8)",
    "confluences_present": "Confluences Active",
    "fvg_present":       "FVG Present",
    "sweep_present":     "Liquidity Sweep Present",
    "bos_tag":           "Structure Event (BOS / CHoCH)",
    "bos_tier":          "Structure Tier (Major / Minor)",
    "break_close_atr":   "Break ATR Multiple",
    "break_excess":      "Break × Over Floor",
    "break_body_atr":    "Break Body (ATR)",
    "break_tier":        "Break Quality",
    "ob_range_atr":      "OB Range / Stop (ATR)",
    "fvg_size_atr":      "FVG Size (ATR)",
    "impulse_leg_to_extreme_atr": "Impulse Leg to Extreme (ATR)",
    "atr_at_ob":         "ATR at OB (price)",
    "chop_at_alert":     "Market Regime (chop 0-100)",
    "vet_review":        "Worth Reviewing",
    "vet_review_reason": "Why Worth Reviewing",
    # news / session audit
    "news_blocked":         "News Blocked",
    "news_event_title":     "News Event",
    "news_event_currency":  "News Currency",
    "news_event_source":    "News Source",
    "news_event_ts":        "News Event Time (UTC)",
    "ist_blocked":          "IST Window Blocked",
    "killzone_blocked":     "Killzone Blocked",
    "killzone_windows":     "Killzone Window(s)",
    "alert_utc_hour":       "Alert Hour (UTC)",
}

_EXIT_LABELS = {
    "sl":           "Stop Loss Hit",
    "tp1":          "TP1 Hit",
    "tp2":          "TP2 Hit",
    "timeout":      "Time Limit Reached (48h)",
    "window_end":   "End of Test Window",
    "sl_collision": "SL and TP Same Bar — SL Taken",
    "never_filled": "Order Never Filled",
    "friday_flat":  "Closed Before Weekend (Fri)",
}

_ENTRY_LABELS = {
    "proximal": "Proximal (OB edge)",
}

_DIR_LABELS = {
    "bullish": "Long",
    "bearish": "Short",
}


def _to_ist_str(ts_val: Any) -> str:
    """Convert a UTC ISO timestamp string (or pd.Timestamp) to IST 'YYYY-MM-DD HH:MM'.

    All stored timestamps in the backtest pipeline are UTC (data_loader pins
    every H1 frame to UTC and the simulator localises naïve timestamps to UTC
    before emitting). So a bare conversion to Asia/Kolkata is correct — no
    silent assumption about the source zone.
    """
    if ts_val is None or ts_val == "" or (isinstance(ts_val, float) and pd.isna(ts_val)):
        return ""
    try:
        ts = pd.Timestamp(ts_val)
        if ts.tzinfo is None:
            ts = ts.tz_localize("UTC")
        return ts.tz_convert("Asia/Kolkata").strftime("%Y-%m-%d %H:%M")
    except Exception:
        return ""


# ---------------------------------------------------------------------------
# Relocated reference tabs (§10) — cut from the email body, live in Excel now.
# All sourced from backtest.insights (the canonical DataFrame analytics that
# aggregate_runs.py also uses) so the numbers match across the whole system.
# ---------------------------------------------------------------------------

def _confluences_tab_df(filled: List[Dict[str, Any]]) -> "pd.DataFrame":
    """Confluence uplift (overall + per pair). Settled non-predictive — kept as
    a reference tab per §A2."""
    if not filled:
        return pd.DataFrame()
    df = pd.DataFrame(filled)
    rows = []
    overall = _insights.confluence_attribution(df, "r_realised")
    for name, d in overall.items():
        rows.append({"Scope": "ALL", "Confluence": name, "N with": d["n_with"],
                     "N without": d["n_without"], "Exp with": d["exp_with"],
                     "Exp without": d["exp_without"], "Uplift R": d["uplift_r"],
                     "Verdict": d["verdict"]})
    if "pair" in df.columns:
        for pair, sub in df.groupby("pair"):
            per = _insights.confluence_attribution(sub, "r_realised")
            for name, d in per.items():
                rows.append({"Scope": pair, "Confluence": name, "N with": d["n_with"],
                             "N without": d["n_without"], "Exp with": d["exp_with"],
                             "Exp without": d["exp_without"], "Uplift R": d["uplift_r"],
                             "Verdict": d["verdict"]})
    return pd.DataFrame(rows)


def _setup_badges_tab_df(filled: List[Dict[str, Any]]) -> "pd.DataFrame":
    """Setup-badge validation buckets (§A3). Data from insights.setup_badge_validation."""
    if not filled:
        return pd.DataFrame()
    df = pd.DataFrame(filled)
    res = _insights.setup_badge_validation(df, "r_realised")
    buckets = res.get("buckets") or []
    if not buckets:
        return pd.DataFrame()
    rows = []
    for b in buckets:
        rows.append({"Badge": b["badge"], "N": b["n"], "Win Rate %": b["win_rate_pct"],
                     "Avg R": b["expectancy_r"], "CI lo": b.get("ci_lo_95"),
                     "CI hi": b.get("ci_hi_95")})
    df_out = pd.DataFrame(rows)
    df_out.attrs["verdict"] = res.get("verdict", "")
    return df_out


def _pair_session_tab_df(filled: List[Dict[str, Any]]) -> "pd.DataFrame":
    """Full pair×session matrix (§A5). Data from insights.pair_session_matrix."""
    if not filled:
        return pd.DataFrame()
    df = pd.DataFrame(filled)
    cells = _insights.pair_session_matrix(df, "r_realised")
    if not cells:
        return pd.DataFrame()
    rows = [{"Pair": c["pair"], "Alert session": c["session"], "N": c["n"],
             "Win Rate %": c["win_rate_pct"], "Avg R": c["expectancy_r"],
             "CI lo": c["ci_lo_95"], "CI hi": c["ci_hi_95"],
             "Live-eligible": c["live_eligible"], "Confidence": c["confidence"]}
            for c in cells]
    return pd.DataFrame(rows)


def _break_ladder_tab_df(filled: List[Dict[str, Any]]) -> "pd.DataFrame":
    """Break-quality ATR ladder (§6f full ladder → Excel). One row per
    (event, measure, ATR tercile) with N/WR/avg R. Deterministic terciles."""
    if not filled:
        return pd.DataFrame()
    df = pd.DataFrame(filled)
    rows = []
    for measure, mlabel in (("break_close_atr", "Break close (ATR)"),
                            ("break_body_atr", "Break body (ATR)")):
        if measure not in df.columns:
            continue
        ser = pd.to_numeric(df[measure], errors="coerce")
        valid = ser.dropna()
        if len(valid) < 6 or valid.nunique() < 3:
            continue
        lo_q, hi_q = valid.quantile([1 / 3, 2 / 3]).tolist()
        if lo_q == hi_q:
            continue
        bands = [(f"low (<={lo_q:.2f})", ser <= lo_q),
                 (f"mid ({lo_q:.2f}-{hi_q:.2f})", (ser > lo_q) & (ser <= hi_q)),
                 (f"high (>{hi_q:.2f})", ser > hi_q)]
        for evt in sorted(df.get("event", pd.Series(dtype=object)).dropna().unique()):
            for blabel, mask in bands:
                sub = df[mask.fillna(False) & (df["event"] == evt)]
                if sub.empty:
                    continue
                rows.append({"Event": evt, "Measure": mlabel, "Band": blabel,
                             "N": len(sub),
                             "Win Rate %": _win_rate(sub, "r_realised"),
                             "Avg R": round(float(sub["r_realised"].mean()), 3)})
    return pd.DataFrame(rows)


def _try_excel(trades: List[Dict[str, Any]], path: Path,
               risk_usd: float = 250.0) -> Optional[Path]:
    """Write human-readable Excel. Returns path or None on failure.

    `risk_usd` is used by the What If counterfactual tab to scale R-deltas
    into $-deltas. Defaults to 250 if caller doesn't pass it.
    """
    # Only filled trades in the Excel — never_filled are counted in fill rate
    # but are not trade outcomes and would confuse the spreadsheet.
    filled = [t for t in trades if _is_real_filled(t)]
    if not filled:
        return None
    try:
        # Make sure r_if_runner is populated for every row so the new Excel
        # column has data. _attach_runner_r is idempotent.
        _attach_runner_r(filled)
        # Dollar P&L under TP1-only and TP1+runner. Same formula as the
        # TP2-ride scoreboard uses internally (R × risk_usd), so per-row
        # sums reconcile to the email's headline totals.
        for t in filled:
            t["pnl_usd_tp1"]    = round(float(t.get("r_if_exit_tp1") or 0.0) * risk_usd, 0)
            t["pnl_usd_runner"] = round(float(t.get("r_if_runner")  or 0.0) * risk_usd, 0)
        df = pd.DataFrame(filled)

        # Day of week — sourced from fill_ts (when the trade was actually
        # live). never_filled rows fall back to alert_ts. Previous code keyed
        # off alert_ts which conflated setup-day with trading-day.
        if "fill_ts" in df.columns:
            df["day_of_week"] = df.apply(
                lambda r: _day_of_week(r.get("fill_ts") or r.get("alert_ts")),
                axis=1,
            )
        elif "alert_ts" in df.columns:
            df["day_of_week"] = df["alert_ts"].apply(_day_of_week)

        # IST timestamp columns (for TradingView verification). Source columns
        # are UTC ISO strings. SL/TP IST cells split exit_ts by exit_reason so
        # a row only carries the IST timestamp for the outcome that actually
        # happened — empty otherwise.
        df["ob_time_ist"]    = df["ob_timestamp"].apply(_to_ist_str)    if "ob_timestamp"  in df.columns else ""
        df["event_time_ist"] = df["bos_timestamp"].apply(_to_ist_str) if "bos_timestamp" in df.columns else ""
        df["alert_time_ist"] = df["alert_ts"].apply(_to_ist_str)      if "alert_ts"      in df.columns else ""
        df["fill_time_ist"]  = df["fill_ts"].apply(_to_ist_str)       if "fill_ts"       in df.columns else ""

        def _sl_ist(row):
            return _to_ist_str(row.get("exit_ts")) if row.get("exit_reason") in ("sl", "sl_collision") else ""

        def _tp_ist(row):
            return _to_ist_str(row.get("exit_ts")) if row.get("exit_reason") in ("tp1", "tp2") else ""

        df["sl_hit_time_ist"] = df.apply(_sl_ist, axis=1) if "exit_ts" in df.columns else ""
        df["tp_fill_time_ist"] = df.apply(_tp_ist, axis=1) if "exit_ts" in df.columns else ""

        # Compute vet_review for each trade.
        reviews = [_flag_vet_review(t) for t in filled]
        df["vet_review"]        = ["Yes" if r[0] else "No" for r in reviews]
        df["vet_review_reason"] = [r[1] for r in reviews]

        # Human-friendly value mapping.
        if "direction" in df.columns:
            df["direction"] = df["direction"].map(_DIR_LABELS).fillna(df["direction"])
        if "entry_zone" in df.columns:
            df["entry_zone"] = df["entry_zone"].map(_ENTRY_LABELS).fillna(df["entry_zone"])
        if "exit_reason" in df.columns:
            df["exit_reason"] = df["exit_reason"].map(_EXIT_LABELS).fillna(df["exit_reason"])
        if "fvg_present" in df.columns:
            df["fvg_present"] = df["fvg_present"].map({True: "Yes", False: "No"}).fillna("")
        if "sweep_present" in df.columns:
            df["sweep_present"] = df["sweep_present"].map({True: "Yes", False: "No"}).fillna("")
        if "reversed_from_extreme" in df.columns:
            # True/False/None -> Yes/No/blank. Blank = not a CHoCH (no reversal
            # origin) or the flag was never stamped. Matches the fvg/sweep style.
            df["reversed_from_extreme"] = (
                df["reversed_from_extreme"].map({True: "Yes", False: "No"}).fillna(""))

        # Select and rename columns — insert day_of_week right after the
        # session/alignment block (pair, 5× IST timestamps, direction,
        # fill_session, ob_session, killzone_alignment = first 10 keys).
        _col_names_with_dow = dict(_EXCEL_COL_NAMES)
        _col_names_with_dow["day_of_week"] = "Day of Week"
        # +1 vs the pre-setup_id layout: setup_id is now the first key, so the
        # session/alignment block ends at index 12 (was 11, +1 for Event Candle).
        # Keeps Day of Week in the same logical slot after killzone_alignment.
        _split_at = 12
        desired = [c for c in list(_EXCEL_COL_NAMES.keys())[:_split_at]
                   + ["day_of_week"]
                   + list(_EXCEL_COL_NAMES.keys())[_split_at:]
                   if c in df.columns]
        out_df = df[desired].rename(columns=_col_names_with_dow)

        # Zone register (one row per OB — the proximal trade for that OB).
        zone_df = _build_zone_register_df(trades)

        # Proximal is the only entry zone (50% mean entry removed 2026-07).
        prox_df = out_df[out_df["Entry Type"] == _ENTRY_LABELS["proximal"]] if "Entry Type" in out_df.columns else out_df

        col_widths = {
            "Setup ID": 10,
            "Currency Pair": 14, "Direction": 10,
            "Fill Session": 13, "OB Session": 13, "Killzone Alignment": 18,
            "H1 Trend": 12, "Trend Alignment": 16,
            "Day of Week": 11,
            "Entry Type": 18, "Entry Price": 12, "Stop Loss": 12,
            "Take Profit 1": 13, "Take Profit 2": 13,
            "TP1 Reward:Risk": 14, "TP2 Reward:Risk": 14,
            "How Trade Closed": 24, "Exit Price": 12,
            "R Achieved (LIVE: TP1+BE@1R)": 22, "Dollar P&L (LIVE: TP1+BE@1R)": 24,
            "R if Closed at TP1": 18, "Dollar P&L (TP1-only)": 20,
            "R if TP1 + Runner": 18, "Dollar P&L (TP1+Runner)": 22,
            "Proximal R": 12, "Proximal Dollar P&L": 18,
            "Stop Loss (raw)": 14,
            "Best Price Reached (R)": 20, "Worst Price Reached (R)": 20,
            "Hours Held": 10, "Hours to TP1 (-1 if never)": 22,
            "Setup Score (0–8)": 14,
            "Confluences Active": 22,
            "FVG Present": 12, "Liquidity Sweep Present": 22,
            "Structure Event (BOS / CHoCH)": 24,
            "Structure Tier (Major / Minor)": 24,
            "Break ATR Multiple": 16, "Break × Over Floor": 16,
            "Break Body (ATR)": 14, "Break Quality": 14,
            "Worth Reviewing": 15, "Why Worth Reviewing": 40,
            "OB Candle (IST)": 18, "Event Candle (IST)": 18, "Scan / Alert Time (IST)": 22,
            "Entry Fill (IST)": 18,
            "SL Hit (IST)": 18, "TP Fill (IST)": 18,
            "PD Zone": 12, "PD Alignment": 14, "PD Array % (entry)": 18,
        }

        def _style_trades_sheet(ws, from_openpyxl):
            PatternFill, Font, Alignment = (
                from_openpyxl["PatternFill"],
                from_openpyxl["Font"],
                from_openpyxl["Alignment"],
            )
            green_fill = PatternFill("solid", fgColor="C6EFCE")
            red_fill   = PatternFill("solid", fgColor="FFC7CE")
            grey_fill  = PatternFill("solid", fgColor="F2F2F2")

            for cell in ws[1]:
                cell.font      = Font(bold=True, color="FFFFFF")
                cell.fill      = PatternFill("solid", fgColor="2C3E50")
                cell.alignment = Alignment(wrap_text=True)

            headers = [cell.value for cell in ws[1]]
            pnl_cols = [headers.index(h) + 1 for h in
                        ("Dollar P&L (LIVE: TP1+BE@1R)", "Dollar P&L",
                         "Proximal Dollar P&L")
                        if h in headers]
            rev_col = headers.index("Worth Reviewing") + 1 if "Worth Reviewing" in headers else None
            nb_col  = headers.index("News Blocked") + 1   if "News Blocked" in headers else None

            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                base_fill = grey_fill if row_idx % 2 == 0 else None
                for cell in row:
                    if base_fill:
                        cell.fill = base_fill
                if pnl_cols:
                    net = sum(
                        ws.cell(row=row_idx, column=c).value or 0
                        for c in pnl_cols
                        if isinstance(ws.cell(row=row_idx, column=c).value, (int, float))
                    )
                    fill = green_fill if net > 0 else (red_fill if net < 0 else None)
                    if fill:
                        for cell in row:
                            cell.fill = fill
                if rev_col and ws.cell(row=row_idx, column=rev_col).value == "Yes":
                    for cell in row:
                        cell.fill = PatternFill("solid", fgColor="FFEB9C")

            if nb_col:
                news_fill = PatternFill("solid", fgColor="E8DAEF")
                for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                    if ws.cell(row=row_idx, column=nb_col).value in (True, "True", "Yes", 1, "1"):
                        for cell in row:
                            cell.fill = news_fill

            for i, col in enumerate(ws.columns, start=1):
                header = ws.cell(row=1, column=i).value
                ws.column_dimensions[col[0].column_letter].width = col_widths.get(header, 14)
            ws.freeze_panes = "A2"

        with pd.ExcelWriter(path, engine="openpyxl") as xw:
            prox_df.to_excel(xw, sheet_name="Proximal", index=False)
            if not zone_df.empty:
                zone_df.to_excel(xw, sheet_name="Zone Register", index=False)

            # Killzone alignment tab — one row per bucket (Both / OB only / Fill only / Neither)
            kz_align_rows = _killzone_alignment_table(filled, "r_realised")
            if kz_align_rows:
                kz_df = pd.DataFrame(kz_align_rows).rename(columns={
                    "bucket":       "Killzone Alignment",
                    "trades":       "Trades",
                    "win_rate_pct": "Win Rate %",
                    "expectancy_r": "Avg R per Trade",
                    "total_r":      "Total R",
                })
                # Plain-English bucket labels, consistent with the email tables.
                kz_df["Killzone Alignment"] = kz_df["Killzone Alignment"].map(
                    _align_label).fillna(kz_df["Killzone Alignment"])
                kz_df.to_excel(xw, sheet_name="Killzone Alignment", index=False)

            # What-If counterfactual tab — now with guard columns (§11.4): CI
            # lo/hi, quarters-sign, ci_cleared. Cut from the body; its guarded
            # survivors re-enter only through Act 5.
            cf_df = _counterfactual_dataframe(filled, risk_usd)
            if not cf_df.empty:
                cf_df.to_excel(xw, sheet_name="What If", index=False)

            # Second Look tab — the full flagged-trade list moved OUT of the
            # email (2026-07). The email now carries only the analysis; every
            # individual flagged row lives here so nothing is lost.
            sl_df = _second_look_df(filled)
            if not sl_df.empty:
                sl_df.to_excel(xw, sheet_name="Second Look", index=False)

            # Relocated reference tabs (§10) — cut from the email body.
            conf_df = _confluences_tab_df(filled)
            if not conf_df.empty:
                conf_df.to_excel(xw, sheet_name="Confluences", index=False)
            badge_df = _setup_badges_tab_df(filled)
            if not badge_df.empty:
                badge_df.to_excel(xw, sheet_name="Setup Badges", index=False)
            ladder_df = _break_ladder_tab_df(filled)
            if not ladder_df.empty:
                ladder_df.to_excel(xw, sheet_name="Break Ladder", index=False)
            psm_df = _pair_session_tab_df(filled)
            if not psm_df.empty:
                psm_df.to_excel(xw, sheet_name="Pair x Session", index=False)

            try:
                from openpyxl.styles import PatternFill, Font, Alignment
                _opx = {"PatternFill": PatternFill, "Font": Font, "Alignment": Alignment}

                _style_trades_sheet(xw.sheets["Proximal"], _opx)

                if "Zone Register" in xw.sheets:
                    zws = xw.sheets["Zone Register"]
                    for cell in zws[1]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill("solid", fgColor="1A5490")
                    for col in zws.columns:
                        zws.column_dimensions[col[0].column_letter].width = 18
                    zws.freeze_panes = "A2"

                for extra_sheet in ("Killzone Alignment", "What If", "Second Look",
                                    "Confluences", "Setup Badges", "Break Ladder",
                                    "Pair x Session"):
                    if extra_sheet in xw.sheets:
                        ews = xw.sheets[extra_sheet]
                        for cell in ews[1]:
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = PatternFill("solid", fgColor="1A5490")
                            cell.alignment = Alignment(wrap_text=True)
                        for col in ews.columns:
                            ews.column_dimensions[col[0].column_letter].width = 22
                        ews.freeze_panes = "A2"

            except ImportError:
                pass  # openpyxl styles not available — plain Excel still written

        return path
    except Exception as e:
        # Stash the real cause so the email footer can name it instead of
        # blaming openpyxl for every failure (locked file on OneDrive, bad
        # column, disk error all landed here as "openpyxl not installed").
        _try_excel.last_error = f"{type(e).__name__}: {e}"
        print(f"  [excel write skipped: {e}]")
        return None


# ---------------------------------------------------------------------------
# HTML helpers
# ---------------------------------------------------------------------------

def _m(amount: float) -> str:
    sign = "+" if amount >= 0 else "-"
    return f"{sign}${abs(amount):,.0f}"


def _r(v: float) -> str:
    sign = "+" if v >= 0 else ""
    return f"{sign}{v:.2f}R"


def _wr_str(v: Optional[float]) -> str:
    """Render a win-rate value. None (no resolved trade — all breakevens) shows
    as an em-dash, never 0%."""
    return "—" if v is None else f"{v:.0f}%"






def _vet_review_category(t: Dict[str, Any]) -> Optional[str]:
    """Category of a flagged trade: 'left_money' | 'nearly_worked' |
    'high_score_loss' | None. Derived from the SAME thresholds _flag_vet_review
    uses, so the email analysis and the Excel tab agree row-for-row."""
    r = t.get("r_realised", 0)
    mfe = t.get("mfe_r", 0)
    score = t.get("score", 0)
    if r > 0 and mfe > r * 2 and mfe > 1.5:
        return "left_money"
    if r < 0 and mfe > 0.5:
        return "nearly_worked"
    if score >= 4 and r < 0:
        return "high_score_loss"
    return None


def _second_look_df(trades: List[Dict[str, Any]]) -> "pd.DataFrame":
    """Full flagged-trade list for the Excel 'Second Look' tab (moved out of the
    email). One row per flagged trade with the columns needed to investigate it."""
    rows = []
    for t in trades:
        if not _is_real_filled(t):
            continue
        cat = _vet_review_category(t)
        if cat is None:
            continue
        _, reason = _flag_vet_review(t)
        rows.append({
            "Setup ID":       t.get("setup_id", ""),
            "Pair":           t.get("pair", ""),
            "Direction":      "Long" if t.get("direction") == "bullish" else "Short",
            "Alert session":  t.get("session", ""),
            "Category":       {"left_money": "Left money on table",
                               "nearly_worked": "Nearly worked",
                               "high_score_loss": "High-score loss"}[cat],
            "R Achieved":     t.get("r_realised", 0),
            "Peak Reached (MFE R)": t.get("mfe_r", 0),
            "Setup Score":    t.get("score", 0),
            "Break Body (ATR)": t.get("break_body_atr"),
            "Alert Time (IST)": _to_ist_str(t.get("alert_ts")),
            "What to look at": reason,
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Pair groups for split per-email reports
# ---------------------------------------------------------------------------

# Email bucket 1 ("original" / Book A): the original 4 FX majors + GOLD.
# Email bucket 2 ("new" / Book B): the new evaluation FX + BTC.
# NAS100 is excluded ENTIRELY (run-level, 2026-06-30 trader decision): it
# generates no trade rows and appears in no email, Excel, or audit.
# (Var names kept for minimal blast radius; the SETS are what the split reads.)
FOREX_PAIRS = {"EURUSD", "NZDUSD", "USDJPY", "USDCHF", "GOLD", "XAUUSD"}
INDEX_COMMODITY_PAIRS = {"GBPUSD", "AUDUSD", "USDCAD", "EURJPY", "BTCUSD"}


def _filter_meta_by_pairs(meta: Dict[str, Any], allowed: set) -> Dict[str, Any]:
    """Return a copy of meta with per-pair fields scoped to `allowed`.

    The audit sections in the per-group emails should only show counters
    for pairs in that group. We touch the per-pair dicts (killzone drops,
    killzone windows) and the pair list; other meta fields are unchanged.
    """
    out = dict(meta)
    out["pairs"] = [p for p in meta.get("pairs", []) if p in allowed]
    drops = meta.get("killzone_drops_by_pair") or {}
    out["killzone_drops_by_pair"] = {k: v for k, v in drops.items() if k in allowed}
    windows = meta.get("killzone_windows_by_pair") or {}
    out["killzone_windows_by_pair"] = {k: v for k, v in windows.items() if k in allowed}
    out["killzone_dropped_alerts"] = int(sum(out["killzone_drops_by_pair"].values()))
    return out


# ===========================================================================
# THE SIX-ACT EMAIL (EMAIL_REBUILD_SPEC, 2026-07-02)
# ---------------------------------------------------------------------------
# One document, six acts, fixed order. Every act opens with one plain-English
# sentence and closes with an action line. All numbers route through the shared
# spine (_stat_block / _slice_read) so "is it real?" is judged identically
# everywhere. BANKABLE = from r_realised or a replayed order; PATH = a peak/touch
# (mfe/mae/"reached"), NEVER phrased as money. Row-level detail lives in Excel.
# ===========================================================================

# --- Design-system tokens (§8.1, locked) -----------------------------------
_C = {
    "surface": "#fcfcfb", "page": "#f9f9f7",
    "ink": "#0b0b0b", "ink2": "#52514e", "muted": "#898781",
    "hair": "#e1e0d9", "baseline": "#c3c2b7",
    "blue": "#2a78d6", "red": "#e34948", "neutral": "#f0efec",
    "good": "#0ca30c", "warn": "#fab219", "serious": "#ec835a", "critical": "#d03b3b",
}

# Verdict → (icon, word, tint bg, ink). Icon+word+color, never color alone (§8.1).
_VERDICT_BANNER = {
    "edge":     ("&#10004;", "EDGE REAL",    "#e3f4e3", _C["good"]),
    "unproven": ("?",        "UNPROVEN",     "#fef3d9", "#9a7400"),
    "loser":    ("&#10008;", "LOSING EDGE",  "#f8e0e0", _C["critical"]),
    "thin":     ("&#8230;",  "TOO THIN",     "#eeeeec", _C["muted"]),
}


def _assert_body_size(html: str, filename: str) -> None:
    """Hard build-time budget (§8.6): Gmail clips bodies over ~102 KB and hides
    the trust section exactly when it matters. Fail the build, never silently
    ship an over-budget email. Row-level detail belongs in Excel."""
    size = len(html.encode("utf-8"))
    if size >= 90_000:
        raise ValueError(
            f"Email body {filename} is {size:,} bytes (>= 90 KB budget). Gmail "
            f"clips ~102 KB — move row-level content to Excel and re-render.")


def _bankable_badge() -> str:
    return (f"<span style='display:inline-block;background:{_C['ink']};color:#fff;"
            f"font-size:10px;font-weight:700;padding:1px 6px;border-radius:3px;"
            f"letter-spacing:0.04em;'>BANKABLE</span>")


def _path_badge() -> str:
    return (f"<span style='display:inline-block;border:1px solid {_C['hair']};"
            f"color:{_C['muted']};font-size:10px;font-weight:700;padding:0 5px;"
            f"border-radius:3px;letter-spacing:0.04em;'>PATH</span>")




def _months_between(meta: Dict[str, Any]) -> int:
    try:
        s = pd.to_datetime(meta.get("start"), utc=True)
        e = pd.to_datetime(meta.get("end"), utc=True)
        return max(1, round((e - s).days / 30.44))
    except Exception:
        return 12


def _quarter_totals(filled: List[Dict[str, Any]], r_col: str = "r_realised"
                    ) -> List[Tuple[str, float]]:
    """[(quarter, total R), ...] chronological, over resolved rows."""
    by_q: Dict[str, float] = {}
    for t in filled:
        if not _is_real_filled(t):
            continue
        q = _quarter_of(t.get("alert_ts"))
        if q is None:
            continue
        by_q[q] = by_q.get(q, 0.0) + float(t.get(r_col) or 0.0)
    return sorted(by_q.items(), key=lambda kv: kv[0])


# ---------------------------------------------------------------------------
# ACT 1 — PULSE
# ---------------------------------------------------------------------------

def _act1_html(prox_trades: List[Dict[str, Any]], sb_prox: Dict[str, Any],
               fill_prox: Dict[str, Any], group_meta: Dict[str, Any],
               risk_usd: float) -> str:
    filled = [t for t in prox_trades if _is_real_filled(t)]
    vals = [float(t.get("r_realised") or 0.0) for t in filled]
    ts = [t.get("alert_ts") for t in filled]
    stat = _stat_block(vals, ts)
    lo, hi = stat["ci"]
    months = _months_between(group_meta)

    icon, word, tint, vink = _VERDICT_BANNER.get(stat["verdict"],
                                                 _VERDICT_BANNER["thin"])
    banner_tail = {"edge": "a real edge", "unproven": "could still be luck",
                   "loser": "a real loser", "thin": "too thin to call"}[stat["verdict"]]

    banner = (
        f"<div style='background:{tint};padding:14px 18px;border-radius:6px;"
        f"margin-bottom:14px;'>"
        f"<span style='font-size:20px;font-weight:800;color:{vink};'>"
        f"{icon}&nbsp;{word}</span>"
        f"<span style='font-size:13px;color:{_C['ink2']};margin-left:10px;'>"
        f"&mdash; {banner_tail}</span></div>"
    )

    # Deterministic sentence.
    pnl = sb_prox.get("total_pnl_usd", 0)
    total_r = sb_prox.get("total_r", 0)
    n = sb_prox.get("trades", 0)
    is_phrase = {"edge": "is a real edge", "unproven": "could still be luck",
                 "loser": "is a real loser",
                 "thin": "is too thin to call"}[stat["verdict"]]
    if lo is not None:
        range_phrase = (f"the honest range is {lo:+.2f}R to {hi:+.2f}R per trade, "
                        f"so this {is_phrase}")
    else:
        range_phrase = (f"there are too few trades for an honest range, so this "
                        f"{is_phrase}")
    exp_r = stat["expR"] if stat["expR"] is not None else 0.0
    sentence = (
        f"<p style='font-size:14px;color:{_C['ink']};margin-bottom:14px;'>"
        f"<b>{n}</b> trades over <b>{months}</b> months made "
        f"<b>{_m(pnl)}</b>. The average trade is <b>{_r(exp_r)}</b> &mdash; "
        f"{range_phrase}. <b>{stat['pq_agree']}</b> of <b>{stat['pq_total']}</b> "
        f"quarters pointed the same way.</p>"
    )

    # 6 KPI tiles (3×2). All BANKABLE.
    dd_r = _insights.max_drawdown_r(vals) if vals else 0.0
    streak = _insights.longest_losing_streak(vals) if vals else 0
    wr = sb_prox.get("win_rate_pct")
    ci_sub = (f"CI [{lo:+.2f}, {hi:+.2f}]" if lo is not None else "thin sample")
    _w, _l, _be = sb_prox.get("wins", 0), sb_prox.get("losses", 0), sb_prox.get("breakevens", 0)
    _be_pct = (_be / n * 100) if n else 0.0
    wr_sub = f"{_w}W &middot; {_l}L &middot; {_be}BE ({_be_pct:.0f}%)"
    tiles = [
        ("P&amp;L", f"{_m(pnl)}", f"{_r(total_r)} total"),
        ("Expectancy", f"{_r(exp_r)}", ci_sub),
        ("Win rate", _wr_str(wr), wr_sub),
        ("N filled / fill rate", f"{n}", f"{fill_prox.get('fill_rate_pct', 0):.0f}% of alerts"),
        ("Max drawdown", f"{dd_r:.1f}R", f"{_m(-dd_r * risk_usd)}"),
        ("Longest losing streak", f"{streak}", "consecutive losses"),
    ]
    tile_cells = ""
    for i, (label, value, sub) in enumerate(tiles):
        if i % 3 == 0:
            tile_cells += "<tr>" if i == 0 else "</tr><tr>"
        tile_cells += (
            f"<td width='33%' style='padding:4px;'>"
            f"<div style='border:1px solid {_C['hair']};border-radius:8px;"
            f"background:{_C['surface']};padding:10px 12px;'>"
            f"<div style='font-size:11px;text-transform:uppercase;letter-spacing:0.05em;"
            f"color:{_C['muted']};'>{label}</div>"
            f"<div style='font-size:24px;font-weight:700;color:{_C['ink']};"
            f"font-variant-numeric:tabular-nums;'>{value}</div>"
            f"<div style='font-size:12px;color:{_C['ink2']};'>{sub}</div>"
            f"</div></td>"
        )
    tile_cells += "</tr>"
    tiles_html = (f"<div style='margin-bottom:6px;'>{_bankable_badge()} "
                  f"<span style='font-size:11px;color:{_C['muted']};'>all tiles from "
                  f"realised P&amp;L</span></div>"
                  f"<table width='100%' style='border-collapse:separate;"
                  f"border-spacing:0;'>{tile_cells}</table>")

    # Quarter chips.
    chips = ""
    for q, qr in _quarter_totals(filled):
        # chip shows AVG R sign (matches _stat_block.pq); use total sign for color.
        avg = stat["pq"].get(q)
        if avg is None:
            continue
        pos = avg >= 0
        bg = "#e3f0fb" if pos else "#fbe3e3"
        fg = _C["blue"] if pos else _C["red"]
        chips += (f"<span style='display:inline-block;background:{bg};color:{fg};"
                  f"font-size:12px;font-weight:600;padding:3px 9px;border-radius:12px;"
                  f"margin:0 4px 4px 0;font-variant-numeric:tabular-nums;'>"
                  f"{q} {avg:+.2f}R</span>")
    chips_html = (f"<div style='margin-top:12px;'>{chips}"
                  f"<span style='font-size:12px;color:{_C['muted']};margin-left:4px;'>"
                  f"sign held {stat['pq_agree']}/{stat['pq_total']}</span></div>")

    action = _act_action(
        {"edge": "Edge is real this period — Act 5 says where to lean in.",
         "unproven": "No action on the headline — the edge is unproven; Act 5 shows "
                     "the only changes that cleared the guard.",
         "loser": "The book loses over this window — Act 4 shows where the red "
                  "concentrates; Act 5 the filter tests.",
         "thin": "Too few trades to act — keep collecting; see Act 5."}[stat["verdict"]])

    return (f'<div class="act">{banner}{sentence}{tiles_html}{chips_html}{action}</div>')


def _act_action(text: str) -> str:
    """The closing action line every act ends with (§6)."""
    return (f"<p style='margin-top:12px;font-size:13px;color:{_C['ink']};"
            f"border-left:3px solid {_C['ink']};padding-left:10px;'>"
            f"<b>Action:</b> {text}</p>")


# ---------------------------------------------------------------------------
# ACT 2 — WHAT THE YEAR LOOKED LIKE
# ---------------------------------------------------------------------------

# Exit reasons that are real resolutions (not audit-only). friday_flat is a real
# weekend-flatten exit; timeout/window_end/never_filled are audit-only (§_EXCLUDE).
def _act2_html(prox_trades: List[Dict[str, Any]], out_dir: Path,
               chart_prefix: str, risk_usd: float) -> str:
    filled = [t for t in prox_trades if _is_real_filled(t)]
    # Order by fill date for the equity curve.
    def _fill_key(t):
        try:
            return pd.to_datetime(t.get("fill_ts") or t.get("alert_ts"), utc=True)
        except Exception:
            return pd.Timestamp.min.tz_localize("UTC")
    ordered = sorted(filled, key=_fill_key)
    r_seq = [float(t.get("r_realised") or 0.0) for t in ordered]

    # Quarter boundary indices for the equity chart vlines.
    q_bounds, seen_q, prev_q = [], set(), None
    for i, t in enumerate(ordered):
        q = _quarter_of(t.get("alert_ts"))
        if q != prev_q and q is not None:
            if prev_q is not None:
                q_bounds.append(i)
            prev_q = q

    qtotals = _quarter_totals(filled)

    # Charts (best-effort; alt text carries the numbers if images blocked).
    equity_img = quarter_img = ""
    try:
        from backtest import report_charts as _charts
        eq_path = out_dir / f"{chart_prefix}_equity.png"
        qb_path = out_dir / f"{chart_prefix}_quarters.png"
        total_r = sum(r_seq)
        if _charts.equity_curve_png(r_seq, eq_path, q_bounds):
            equity_img = (
                f"<img src='{eq_path.name}' width='640' "
                f"alt='Equity curve, cumulative {total_r:+.1f}R over {len(r_seq)} "
                f"trades' style='width:640px;max-width:100%;height:auto;'>")
        if _charts.quarter_bars_png(qtotals, qb_path):
            qtxt = ", ".join(f"{q} {v:+.1f}R" for q, v in qtotals)
            quarter_img = (
                f"<img src='{qb_path.name}' width='640' "
                f"alt='Quarterly totals: {qtxt}' "
                f"style='width:640px;max-width:100%;height:auto;margin-top:8px;'>")
    except Exception as e:
        equity_img = (f"<p style='font-size:12px;color:{_C['muted']};'>"
                      f"[chart unavailable: {type(e).__name__}]</p>")

    # Exit mix — horizontal stacked HTML bar (real resolutions only).
    counts = _exit_reason_counts(filled)
    real = {k: v for k, v in counts.items() if k not in _EXCLUDE_REASONS}
    total_real = sum(real.values()) or 1
    # Group into tp1 / sl / breakeven-ish / other.
    seg_order = [("tp1", _C["good"], "TP1"), ("sl", _C["red"], "SL"),
                 ("friday_flat", _C["blue"], "Weekend flat")]
    seg_html = ""
    legend = ""
    shown = set()
    for key, color, name in seg_order:
        c = real.get(key, 0)
        if c == 0:
            continue
        shown.add(key)
        pct = c / total_real * 100
        seg_html += (f"<td width='{pct:.2f}%' style='background:{color};height:22px;'"
                     f"></td>")
        legend += (f"<span style='font-size:12px;color:{_C['ink2']};margin-right:12px;'>"
                   f"<span style='display:inline-block;width:10px;height:10px;"
                   f"background:{color};border-radius:2px;'></span> {name}: {c}</span>")
    other = sum(v for k, v in real.items() if k not in shown)
    if other:
        pct = other / total_real * 100
        seg_html += (f"<td width='{pct:.2f}%' style='background:{_C['neutral']};"
                     f"height:22px;'></td>")
        legend += (f"<span style='font-size:12px;color:{_C['ink2']};margin-right:12px;'>"
                   f"<span style='display:inline-block;width:10px;height:10px;"
                   f"background:{_C['neutral']};border-radius:2px;'></span> other: {other}</span>")
    exit_mix = (
        f"<h4 style='margin:16px 0 6px;'>How trades ended</h4>"
        f"<table width='100%' style='border-collapse:collapse;border-radius:4px;"
        f"overflow:hidden;'><tr>{seg_html}</tr></table>"
        f"<div style='margin-top:6px;'>{legend}</div>"
        f"<p style='font-size:11px;color:{_C['muted']};margin-top:4px;'>"
        f"Audit-only exits (never filled / timeout / data-window end) are "
        f"excluded here and never feed P&amp;L.</p>")

    # Tempo line.
    def _median(xs):
        xs = sorted(x for x in xs if x is not None)
        if not xs:
            return None
        m = len(xs) // 2
        return xs[m] if len(xs) % 2 else (xs[m - 1] + xs[m]) / 2
    win_bars = _median([int(t.get("bars_to_exit") or 0) for t in filled
                        if (t.get("r_realised") or 0) > 0])
    loss_bars = _median([int(t.get("bars_to_exit") or 0) for t in filled
                         if (t.get("r_realised") or 0) < 0])
    weeks = max(1, len(qtotals) * 13)
    per_week = len(filled) / weeks
    tempo = (
        f"<p style='font-size:13px;color:{_C['ink2']};margin-top:14px;'>"
        f"<b>Tempo:</b> winners resolve in ~{win_bars if win_bars is not None else '—'} "
        f"H1 bars, losers ~{loss_bars if loss_bars is not None else '—'}; "
        f"~{per_week:.1f} trades/week.</p>")

    # Capture tile (PATH-labeled).
    wins = [t for t in filled if (t.get("r_realised") or 0) > 0]
    avg_booked = (sum(float(t.get("r_realised") or 0) for t in wins) / len(wins)
                  if wins else 0.0)
    avg_mfe = (sum(float(t.get("mfe_r") or 0) for t in wins) / len(wins)
               if wins else 0.0)
    capture = (avg_booked / avg_mfe * 100) if avg_mfe > 0 else 0.0
    capture_tile = (
        f"<div style='border:1px solid {_C['hair']};border-radius:8px;"
        f"background:{_C['surface']};padding:10px 12px;margin-top:12px;'>"
        f"{_path_badge()} "
        f"<span style='font-size:13px;color:{_C['ink']};'>Winners banked "
        f"<b>{capture:.0f}%</b> of their best in-trade price on average.</span>"
        f"<div style='font-size:11px;color:{_C['muted']};margin-top:2px;'>"
        f"PATH = a peak on the path, not money booked; the bankable exit answer is "
        f"the recipe table in Act 5.</div></div>")

    intro = (f"<p style='font-size:14px;color:{_C['ink']};margin-bottom:12px;'>"
             f"How the year actually felt to trade &mdash; the equity path, the "
             f"quarter-by-quarter swing, and how positions closed.</p>")
    action = _act_action("Read the drawdown depth and streak against your risk "
                         "tolerance; the cause of the swings is Acts 3 and 4.")
    return (f'<div class="act">{intro}{equity_img}{quarter_img}{exit_mix}'
            f'{tempo}{capture_tile}{action}</div>')


# ---------------------------------------------------------------------------
# ACT 3 / ACT 4 — the shared cause-panel engine
# ---------------------------------------------------------------------------

# Fixed SMC cause dimensions, ALWAYS shown (thin or not) (§Act 3 Panel A).
# (column, human label, bucketer). Bucketer None => categorical value as-is.
def _bucket_seq_depth(v) -> Optional[str]:
    try:
        n = int(float(v))
    except (TypeError, ValueError):
        return None
    return "4+" if n >= 4 else str(n)


def _clean_event(v) -> Optional[str]:
    """Display-only: collapse the detector's doubled event labels ('BOS BOS',
    'CHoCH CHoCH') to a single token. Render-layer fix — the raw `event` value is
    upstream trade data this task must not touch (§0)."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip()
    parts = s.split()
    if len(parts) == 2 and parts[0] == parts[1]:
        return parts[0]
    return s


_CAUSE_DIMS = [
    ("event",              "Event type",         _clean_event),
    ("bos_sequence_count", "Continuation depth", _bucket_seq_depth),
    ("bos_tier",           "Break tier",         None),
    ("killzone_alignment", "Killzone alignment", None),
    ("pd_alignment",       "PD alignment",       None),
    ("trend_alignment",    "Trend alignment",    None),
    ("fill_session",       "Fill session",       None),
    ("pair",               "Pair",               None),
]


def _cause_rows(filled: List[Dict[str, Any]], base_exr: float,
                side: str) -> List[Dict[str, Any]]:
    """Score every value of every fixed cause dimension via the shared engine.
    `side` = 'win' keeps positive-gap rows, 'loss' keeps negative-gap rows.
    Returns rows sorted by rank desc, each with dim/value/read."""
    out = []
    for col, label, bucketer in _CAUSE_DIMS:
        groups: Dict[str, List[Dict[str, Any]]] = {}
        for t in filled:
            raw = t.get(col)
            val = bucketer(raw) if bucketer else raw
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            groups.setdefault(str(val), []).append(t)
        for val, rows in groups.items():
            read = _slice_read(rows, base_exr, "r_realised")
            if read is None:
                continue
            gap = read["gap"]
            if side == "win" and gap <= 0:
                continue
            if side == "loss" and gap >= 0:
                continue
            out.append({"dim": label, "value": val, **read})
    out.sort(key=lambda r: r["rank"], reverse=True)
    return out


def _cause_panel_html(rows: List[Dict[str, Any]], side: str,
                      limit: int = 8) -> str:
    """Render the fixed-dimension cause panel (Act 3/4 Panel A)."""
    if not rows:
        return (f"<p style='font-size:13px;color:{_C['muted']};'>No "
                f"{'winning' if side == 'win' else 'leaking'} cause dimension "
                f"diverged from the book this period.</p>")
    body = ""
    for r in rows[:limit]:
        lo, hi = r["ci"]
        ci_s = f"[{lo:+.2f},{hi:+.2f}]" if lo is not None else "thin"
        pq_s = (f"{r['pq_agree']}/{r['pq_total']}q" if r["pq_total"] else "—")
        if r["promoted"]:
            verdict = (f"<span style='color:{_C['good'] if side=='win' else _C['critical']};"
                       f"font-weight:700;'>driver</span>")
        else:
            verdict = f"<span style='color:{_C['muted']};'>directional, thin</span>"
        # SMC-conflict marker.
        exp = _smc_expected_direction(r["dim"], r["value"])
        if exp is not None and (exp * r["gap"]) < 0:
            verdict += (f" <span style='color:{_C['serious']};font-weight:700;'>"
                        f"&#9888; conflicts SMC</span>")
        gap_color = _C["good"] if r["gap"] >= 0 else _C["critical"]
        wr_s = "—" if r["wr"] is None else f"{r['wr']:.0f}%"
        body += (
            f"<tr>"
            f"<td style='text-align:left;'><b>{r['dim']}:</b> {r['value']}</td>"
            f"<td style='text-align:right;font-variant-numeric:tabular-nums;'>{r['n']}</td>"
            f"<td style='text-align:right;'>{wr_s}</td>"
            f"<td style='text-align:right;font-variant-numeric:tabular-nums;'>{_r(r['expR'])}</td>"
            f"<td style='text-align:right;color:{gap_color};font-variant-numeric:tabular-nums;'>{r['gap']:+.2f}R</td>"
            f"<td style='text-align:right;font-size:11px;color:{_C['muted']};'>{ci_s}</td>"
            f"<td style='text-align:right;font-size:11px;color:{_C['muted']};'>{pq_s}</td>"
            f"<td style='text-align:left;'>{verdict}</td>"
            f"</tr>")
    return (f"<table width='100%'><thead><tr>"
            f"<th style='text-align:left;'>Cause</th><th style='text-align:right;'>N</th>"
            f"<th style='text-align:right;'>WR</th><th style='text-align:right;'>Avg R</th>"
            f"<th style='text-align:right;'>vs base</th><th style='text-align:right;'>CI</th>"
            f"<th style='text-align:right;'>Qtrs</th><th style='text-align:left;'>Verdict</th>"
            f"</tr></thead><tbody>{body}</tbody></table>")


def _mined_slices(filled: List[Dict[str, Any]], base_exr: float,
                  side: str) -> List[Dict[str, Any]]:
    """Panel B — mined patterns incl. 2-ways, reusing the driver bucket builder.
    Returns ONLY promoted slices on the requested side, top by rank."""
    cands = _driver_buckets(filled) + _driver_two_way(filled)
    scored = []
    for c in cands:
        rows = c["rows"]
        if len(rows) < 2 or len(rows) == len(filled):
            continue
        read = _slice_read(rows, base_exr, "r_realised")
        if read is None or not read["promoted"]:
            continue
        if side == "win" and read["gap"] <= 0:
            continue
        if side == "loss" and read["gap"] >= 0:
            continue
        scored.append({"dim": c["dim"], "value": c["value"], **read})
    scored.sort(key=lambda r: r["rank"], reverse=True)
    return scored


def _mined_panel_html(slices: List[Dict[str, Any]], side: str) -> str:
    if not slices:
        return (f"<p style='font-size:13px;color:{_C['muted']};'>No new "
                f"{'winner' if side == 'win' else 'leak'} pattern cleared the guard "
                f"this period.</p>")
    items = ""
    for s in slices[:3]:
        lo, hi = s["ci"]
        ci_s = f"[{lo:+.2f},{hi:+.2f}]" if lo is not None else "thin"
        items += (f"<li style='margin-bottom:4px;'><b>{s['dim']}: {s['value']}</b> "
                  f"&mdash; {_r(s['expR'])} ({s['gap']:+.2f}R vs base), N={s['n']}, "
                  f"CI {ci_s}, {s['pq_agree']}/{s['pq_total']}q</li>")
    return f"<ul style='padding-left:18px;font-size:13px;'>{items}</ul>"


def _act3_html(prox_trades: List[Dict[str, Any]]) -> str:
    filled = [t for t in prox_trades if _is_real_filled(t)]
    if not filled:
        return '<div class="act"><p>No filled trades this period.</p></div>'
    base_exr = sum(float(t.get("r_realised") or 0.0) for t in filled) / len(filled)

    cause = _cause_rows(filled, base_exr, "win")
    mined = _mined_slices(filled, base_exr, "win")

    # The stitch (deterministic prose): top promoted win slice × best sub-dim.
    stitch = ""
    top = next((r for r in cause if r["promoted"]), None) or (mined[0] if mined else None)
    if top:
        sub_txt = ""
        subrows = [t for t in filled
                   if str(top.get("value")) == str(_val_for(t, top["dim"]))]
        if subrows:
            sub = _best_subslice(subrows, base_exr)
            if sub:
                sub_txt = f" &mdash; mostly {sub['dim']}: {sub['value']}"
        stitch = (
            f"<p style='font-size:14px;color:{_C['ink']};margin-top:12px;'>"
            f"<b>The stitch:</b> {top['dim']} = {top['value']} carries the edge: "
            f"{_r(top['expR'])} over {top['n']} trades, sign held "
            f"{top['pq_agree']}/{top['pq_total']} quarters{sub_txt}.</p>")
    elif cause:
        stitch = (f"<p style='font-size:13px;color:{_C['muted']};margin-top:12px;'>"
                  f"No win cause cleared the guard &mdash; the strongest tilt is "
                  f"directional only, not a rule yet.</p>")

    intro = (f"<p style='font-size:14px;color:{_C['ink']};margin-bottom:10px;'>"
             f"Where the winners come from. Every SMC dimension scored against the "
             f"book's own average; a <b>driver</b> cleared the guard (enough trades, "
             f"a real gap, steady across quarters), everything else is a hint.</p>")
    action = _act_action(
        "Lean into promoted win drivers via sizing/priority (not a code change); "
        "see Act 5." if top and top.get("promoted")
        else "No win driver cleared the guard &mdash; no action, thin.")
    return (f'<div class="act">{intro}'
            f'<h4>Cause panel &mdash; fixed SMC dimensions</h4>'
            f'{_cause_panel_html(cause, "win")}'
            f'<h4 style="margin-top:16px;">Discovered patterns (mined, promoted only)</h4>'
            f'{_mined_panel_html(mined, "win")}{stitch}{action}</div>')


def _verdict_pill_html(verdict: str) -> str:
    """Small pill using the same icon/word/colour as the Act 1 banner, so a
    pair's verdict and the book's verdict read as the same claim."""
    icon, word, tint, vink = _VERDICT_BANNER.get(verdict, _VERDICT_BANNER["thin"])
    return (f"<span style='background:{tint};color:{vink};font-size:11px;"
            f"font-weight:700;padding:2px 8px;border-radius:10px;"
            f"white-space:nowrap;'>{icon}&nbsp;{word}</span>")


def _pair_table_html(prox_trades: List[Dict[str, Any]], risk_usd: float) -> str:
    """Per-pair scoreboard: N · win rate · expectancy · total P&L · verdict.
    Same filled-only, r_realised basis as the headline — sums reconcile.
    Sorted by P&L (best pair first). Verdict is each pair's OWN bootstrap 95%
    CI test (_stat_block), not inferred from win rate or P&L sign — a pair
    with 3 trades and a big P&L still reads THIN, not EDGE. No drawdown per
    pair (a whole-book measure)."""
    rows = _per_pair_breakdown(prox_trades, "r_realised", risk_usd)
    if not rows:
        return ('<div class="act"><p style="color:%s;">No filled trades this '
                'period.</p></div>' % _C["muted"])
    body = ""
    for p in rows:
        pnl = p["total_pnl_usd"]
        pnl_color = _C["good"] if pnl >= 0 else _C["critical"]
        body += (
            f"<tr>"
            f"<td style='text-align:left;'><b>{p['pair']}</b></td>"
            f"<td style='text-align:right;font-variant-numeric:tabular-nums;'>{p['trades']}</td>"
            f"<td style='text-align:right;font-variant-numeric:tabular-nums;'>{_wr_str(p['win_rate_pct'])}</td>"
            f"<td style='text-align:right;font-variant-numeric:tabular-nums;'>{_r(p['expectancy_r'])}</td>"
            f"<td style='text-align:right;font-variant-numeric:tabular-nums;color:{pnl_color};'>{_m(pnl)}</td>"
            f"<td style='text-align:center;'>{_verdict_pill_html(p['verdict'])}</td>"
            f"</tr>")
    intro = (f"<p style='font-size:14px;color:{_C['ink']};margin-bottom:10px;'>"
             f"Every pair on the same basis as the headline (filled trades, "
             f"realised R). Which instruments carry the book, which drag it. "
             f"Verdict is each pair's own CI test — thin per-pair samples stay "
             f"THIN even with a big headline P&amp;L.</p>")
    table = (
        f"<table width='100%'><thead><tr>"
        f"<th style='text-align:left;'>Pair</th>"
        f"<th style='text-align:right;'>N</th>"
        f"<th style='text-align:right;'>Win rate</th>"
        f"<th style='text-align:right;'>Expectancy</th>"
        f"<th style='text-align:right;'>Total P&amp;L</th>"
        f"<th style='text-align:center;'>Verdict</th>"
        f"</tr></thead><tbody>{body}</tbody></table>")
    return f'<div class="act">{intro}{table}</div>'


def _val_for(t: Dict[str, Any], dim_label: str):
    """Map a cause dim human-label back to the trade's bucketed value."""
    for col, label, bucketer in _CAUSE_DIMS:
        if label == dim_label:
            raw = t.get(col)
            return bucketer(raw) if bucketer else raw
    return None


def _best_subslice(rows: List[Dict[str, Any]], base_exr: float
                   ) -> Optional[Dict[str, Any]]:
    """Best sub-dimension within a slice (for the stitch). Scans the cause dims
    and returns the strongest promoted-or-directional sub-value."""
    best = None
    for col, label, bucketer in _CAUSE_DIMS:
        groups: Dict[str, List[Dict[str, Any]]] = {}
        for t in rows:
            raw = t.get(col)
            val = bucketer(raw) if bucketer else raw
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            groups.setdefault(str(val), []).append(t)
        if len(groups) < 2:  # a sub-dimension needs contrast to be interesting
            continue
        for val, sub in groups.items():
            read = _slice_read(sub, base_exr, "r_realised")
            if read is None:
                continue
            if best is None or read["rank"] > best["rank"]:
                best = {"dim": label, "value": val, **read}
    return best


# ---------------------------------------------------------------------------
# ACT 4 — WHERE IT LEAKS + the stitched leak table
# ---------------------------------------------------------------------------

def _sink_by_join(group_sink: List[Dict[str, Any]]) -> Dict[tuple, Dict[str, float]]:
    """Index the exit-lab sink by (pair, ob_timestamp, direction) → {config: r}.
    Same unique join key the pipeline uses (pair+alert_ts is NOT unique)."""
    idx: Dict[tuple, Dict[str, float]] = {}
    for r in group_sink or []:
        if r.get("entry_zone") != "proximal":
            continue
        if r.get("exit_reason") in _EXCLUDE_REASONS:
            continue
        key = (r.get("pair"), str(r.get("ob_timestamp")), r.get("direction"))
        idx.setdefault(key, {})[r.get("config")] = float(r.get("r") or 0.0)
    return idx


def _leak_buckets(filled: List[Dict[str, Any]], base_exr: float
                  ) -> List[Dict[str, Any]]:
    """Rank leak buckets by SHARE OF LEAK (sum of negative R in bucket ÷ total
    negative R of the book). Returns promoted-or-thin-flagged buckets."""
    total_neg = sum(float(t.get("r_realised") or 0.0) for t in filled
                    if (t.get("r_realised") or 0) < 0)
    if total_neg == 0:
        return []
    cands = _cause_rows(filled, base_exr, "loss")
    # attach share-of-leak + the bucket's rows for the join.
    for c in cands:
        rows = [t for t in filled if str(_val_for(t, c["dim"])) == str(c["value"])]
        c["rows"] = rows
        neg = sum(float(t.get("r_realised") or 0.0) for t in rows
                  if (t.get("r_realised") or 0) < 0)
        c["leak_share"] = (neg / total_neg) if total_neg else 0.0
    cands.sort(key=lambda c: c["leak_share"], reverse=True)
    return cands


def _best_exit_fix(bucket_rows: List[Dict[str, Any]],
                   sink_idx: Dict[tuple, Dict[str, float]]
                   ) -> Optional[Dict[str, Any]]:
    """Within a leak bucket, replay each recipe (from the sink) and find the best
    avg-R improvement over LIVE. Paired per-trade delta → _stat_block verdict.
    BANKABLE (replayed orders). Returns None if the bucket doesn't join."""
    # Build per-trade paired R: (live_r, {config: r}) for each joinable trade.
    live_by_cfg: Dict[str, List[float]] = {}
    deltas_by_cfg: Dict[str, List[float]] = {}
    ts_by_cfg: Dict[str, List[Any]] = {}
    for t in bucket_rows:
        key = (t.get("pair"), str(t.get("ob_timestamp")), t.get("direction"))
        cfgs = sink_idx.get(key)
        if not cfgs or _EXIT_BASELINE_KEY not in cfgs:
            continue
        live_r = cfgs[_EXIT_BASELINE_KEY]
        for cfg, rv in cfgs.items():
            if cfg == _EXIT_BASELINE_KEY:
                continue
            live_by_cfg.setdefault(cfg, []).append(live_r)
            deltas_by_cfg.setdefault(cfg, []).append(rv - live_r)
            ts_by_cfg.setdefault(cfg, []).append(t.get("alert_ts"))
    best = None
    for cfg, deltas in deltas_by_cfg.items():
        if not deltas:
            continue
        stat = _stat_block(deltas, ts_by_cfg[cfg])
        avg_delta = stat["expR"]
        if avg_delta is None or avg_delta <= 0:
            continue
        lo, hi = stat["ci"]
        pq_ok = (stat["pq_total"] == 0
                 or stat["pq_agree"] / stat["pq_total"] >= DRIVER_PQ_SIGN_FRAC)
        promoted = (lo is not None and lo > 0 and pq_ok)
        if best is None or avg_delta > best["avg_delta"]:
            best = {"cfg": cfg, "label": _EXIT_RECIPE_LABELS.get(cfg, cfg),
                    "avg_delta": avg_delta, "ci": (lo, hi), "n": stat["n"],
                    "pq_agree": stat["pq_agree"], "pq_total": stat["pq_total"],
                    "promoted": promoted}
    return best


def _act4_html(prox_trades: List[Dict[str, Any]], group_sink: List[Dict[str, Any]],
               risk_usd: float) -> Dict[str, Any]:
    """Returns {'html': str, 'candidates': [...]} — the candidates feed Act 5."""
    filled = [t for t in prox_trades if _is_real_filled(t)]
    if not filled:
        return {"html": '<div class="act"><p>No filled trades.</p></div>',
                "candidates": []}
    base_exr = sum(float(t.get("r_realised") or 0.0) for t in filled) / len(filled)
    sink_idx = _sink_by_join(group_sink)

    cause = _cause_rows(filled, base_exr, "loss")
    buckets = _leak_buckets(filled, base_exr)

    # The stitched leak table (the centerpiece).
    def _median(xs):
        xs = sorted(x for x in xs if x is not None)
        if not xs:
            return None
        m = len(xs) // 2
        return xs[m] if len(xs) % 2 else (xs[m - 1] + xs[m]) / 2

    rows_html = ""
    stitch_prose = ""
    candidates = []
    for c in buckets[:5]:
        rows = c["rows"]
        med_bars = _median([int(t.get("bars_to_exit") or 0) for t in rows])
        med_mae = _median([float(t.get("mae_r") or 0.0) for t in rows
                           if t.get("mae_r") is not None])
        fix = _best_exit_fix(rows, sink_idx)
        if fix and fix["avg_delta"] is not None:
            flo, fhi = fix["ci"]
            fix_verdict = ("driver" if fix["promoted"] else "directional, thin")
            fix_txt = (f"{fix['label']} +{fix['avg_delta']:.2f}R "
                       f"<span style='font-size:11px;color:{_C['muted']};'>"
                       f"({fix_verdict})</span>")
        else:
            fix_txt = f"<span style='color:{_C['muted']};'>none</span>"
        promoted = c["promoted"]
        verdict = ("driver" if promoted else "directional, thin")
        vcolor = _C["critical"] if promoted else _C["muted"]
        rows_html += (
            f"<tr>"
            f"<td style='text-align:left;'><b>{c['dim']}:</b> {c['value']}</td>"
            f"<td style='text-align:right;font-variant-numeric:tabular-nums;'>{c['n']}</td>"
            f"<td style='text-align:right;font-variant-numeric:tabular-nums;'>{_r(c['expR'])}</td>"
            f"<td style='text-align:right;font-variant-numeric:tabular-nums;'>{c['leak_share']*100:.0f}%</td>"
            f"<td style='text-align:right;'>{med_bars if med_bars is not None else '—'}</td>"
            f"<td style='text-align:right;font-variant-numeric:tabular-nums;'>"
            f"{med_mae:.2f}R" if med_mae is not None else "—"
            f"</td>"
            f"<td style='text-align:left;'>{fix_txt}</td>"
            f"<td style='text-align:left;color:{vcolor};'>{verdict}</td>"
            f"</tr>")
        # Connect-the-dots prose for the single worst bucket.
        if not stitch_prose:
            if fix and fix["promoted"]:
                fix_sentence = (f"{fix['label']} rescues it: +{fix['avg_delta']:.2f}R "
                                f"per trade in this bucket (replayed order, BANKABLE).")
            else:
                fix_sentence = ("No exit recipe rescues it &mdash; candidate for a "
                                "filter test, see Act 5.")
            stitch_prose = (
                f"<p style='font-size:14px;color:{_C['ink']};margin-top:12px;'>"
                f"<b>{c['dim']}: {c['value']}</b> is {c['leak_share']*100:.0f}% of the "
                f"total leak ({_r(c['expR'])} avg over {c['n']} trades, "
                f"{c['pq_agree']}/{c['pq_total']} quarters). {fix_sentence}</p>")
        # Feed Act 5.
        candidates.append({"bucket": c, "fix": fix})

    leak_table = (
        f"<h4 style='margin-top:16px;'>The stitched leak table</h4>"
        f"<div style='margin-bottom:4px;'>{_bankable_badge()} exit fix &nbsp; "
        f"{_path_badge()} MAE</div>"
        f"<table width='100%'><thead><tr>"
        f"<th style='text-align:left;'>Leak bucket</th><th style='text-align:right;'>N</th>"
        f"<th style='text-align:right;'>Avg R</th><th style='text-align:right;'>Leak share</th>"
        f"<th style='text-align:right;'>Bars held (med)</th>"
        f"<th style='text-align:right;'>Worst drawdown before close (med)</th>"
        f"<th style='text-align:left;'>Best exit fix</th><th style='text-align:left;'>Verdict</th>"
        f"</tr></thead><tbody>{rows_html}</tbody></table>"
        if rows_html else
        f"<p style='font-size:13px;color:{_C['muted']};'>No losing bucket to stitch "
        f"this period.</p>")

    # Losers-green touch hint (PATH, one line).
    green_losers = [t for t in filled
                    if (t.get("r_realised") or 0) < 0 and (t.get("mfe_r") or 0) >= 0.5]
    touch_hint = (
        f"<p style='font-size:12px;color:{_C['ink2']};margin-top:10px;'>"
        f"{_path_badge()} <b>{len(green_losers)}</b> losers touched &ge;+0.5R "
        f"in-trade before reversing (PATH &mdash; a touch is not an exit). The "
        f"bankable answer is the break-even rows of the recipe table in Act 5.</p>")

    intro = (f"<p style='font-size:14px;color:{_C['ink']};margin-bottom:10px;'>"
             f"Where the red concentrates &mdash; and, next to each leak, the best "
             f"exit recipe that would have rescued it (a replayed order, not a "
             f"peak).</p>")
    action = _act_action(
        "The stitched table pairs each leak with its bankable fix or flags it for a "
        "filter test &mdash; Act 5 ranks which to try first.")
    html = (f'<div class="act">{intro}'
            f'<h4>Cause panel &mdash; leak side</h4>{_cause_panel_html(cause, "loss")}'
            f'{leak_table}{stitch_prose}{touch_hint}{action}</div>')
    return {"html": html, "candidates": candidates}


# ---------------------------------------------------------------------------
# ACT 5 — WHAT TO CHANGE NEXT (max 3 cards)
# ---------------------------------------------------------------------------

def _act5_html(prox_trades: List[Dict[str, Any]], group_sink: List[Dict[str, Any]],
               leak_candidates: List[Dict[str, Any]], head_stat: Dict[str, Any],
               risk_usd: float) -> str:
    filled = [t for t in prox_trades if _is_real_filled(t)]
    n_book = len(filled)
    base_exr = (sum(float(t.get("r_realised") or 0.0) for t in filled) / n_book
                if n_book else 0.0)

    cards = []          # each: {claim, guard, delta, watch, impact, seen}
    seen_buckets = set()

    # (1) Any exit recipe that beats LIVE under the strict rule (§11.1).
    win = _recipe_winner(group_sink)
    if win and win["promoted"]:
        lo, hi = win["ci"]
        dollar = win["avg_delta"] * risk_usd * win["n_book"]
        cards.append({
            "claim": f"Switch to {win['label']} &mdash; beats LIVE by "
                     f"{win['avg_delta']:+.2f}R/trade.",
            "guard": f"CI [{lo:+.2f}, {hi:+.2f}], sign {win['pq_agree']}/{win['pq_total']}q, "
                     f"N={win['n']}.",
            "delta": f"{_m(dollar)} over {win['n_book']} trades if it held "
                     f"(replayed orders, BANKABLE).",
            "watch": "Re-check on the 18-yr run; keep only if the paired CI stays "
                     "clear of zero.",
            "impact": abs(win["avg_delta"]) * (win["n"] ** 0.5)})

    # (2) & (3) leak-bucket candidates.
    for cand in leak_candidates:
        c = cand["bucket"]
        fix = cand["fix"]
        bkey = (c["dim"], c["value"])
        if bkey in seen_buckets or not c["promoted"]:
            continue
        seen_buckets.add(bkey)
        if fix and fix["promoted"]:
            lo, hi = fix["ci"]
            dollar = fix["avg_delta"] * risk_usd * fix["n"]
            cards.append({
                "claim": f"For {c['dim']}={c['value']}, use {fix['label']} &mdash; "
                         f"+{fix['avg_delta']:.2f}R/trade in this cluster.",
                "guard": f"CI [{lo:+.2f}, {hi:+.2f}], sign {fix['pq_agree']}/{fix['pq_total']}q, "
                         f"N={fix['n']}. Scoped to a cluster &mdash; needs long-history "
                         f"confirmation before scoping.",
                "delta": f"{_m(dollar)} over {fix['n']} in-cluster trades (replayed, "
                         f"BANKABLE).",
                "watch": "Confirm the cluster holds on the 18-yr run before scoping the "
                         "recipe to it.",
                "impact": abs(fix["avg_delta"]) * (fix["n"] ** 0.5)})
        else:
            # filter-test candidate: re-aggregate book WITHOUT the bucket.
            rows = c["rows"]
            keep = [t for t in filled if t not in rows]
            kvals = [float(t.get("r_realised") or 0.0) for t in keep]
            kts = [t.get("alert_ts") for t in keep]
            kstat = _stat_block(kvals, kts)
            lo, hi = kstat["ci"]
            base_dollar = base_exr * risk_usd * n_book
            new_dollar = (kstat["expR"] or 0.0) * risk_usd * len(keep)
            cards.append({
                "claim": f"Filter-test: drop {c['dim']}={c['value']} &mdash; it is "
                         f"{c['leak_share']*100:.0f}% of the leak with no exit rescue.",
                "guard": (f"Remainder CI [{lo:+.2f}, {hi:+.2f}], "
                          f"sign {kstat['pq_agree']}/{kstat['pq_total']}q, N={kstat['n']}."
                          + ("" if (lo is not None and lo > 0)
                             else " <b>not CI-cleared</b>")),
                "delta": f"Book expectancy moves {_r(base_exr)} &rarr; {_r(kstat['expR'] or 0)} "
                         f"(from {_m(base_dollar)} to {_m(new_dollar)}, fewer trades).",
                "watch": "A filter is a real trade-count cut &mdash; confirm the remainder "
                         "CI clears zero on more data before gating.",
                "impact": abs(c["expR"]) * (c["n"] ** 0.5)})

    # (4) Top promoted win bucket → lean-in (sizing/priority, not code).
    win_causes = _cause_rows(filled, base_exr, "win")
    top_win = next((r for r in win_causes if r["promoted"]), None)
    if top_win and (top_win["dim"], top_win["value"]) not in seen_buckets:
        lo, hi = top_win["ci"]
        cards.append({
            "claim": f"Lean into {top_win['dim']}={top_win['value']} &mdash; "
                     f"{top_win['gap']:+.2f}R above the book.",
            "guard": f"CI [{lo:+.2f}, {hi:+.2f}], sign {top_win['pq_agree']}/{top_win['pq_total']}q, "
                     f"N={top_win['n']}. Sizing/priority only &mdash; not a code change.",
            "delta": "No replayed delta (this is a weighting call, not an exit change).",
            "watch": "Confirm the tilt persists next run before changing sizing.",
            "impact": abs(top_win["gap"]) * (top_win["n"] ** 0.5)})

    cards.sort(key=lambda c: c["impact"], reverse=True)
    cards = cards[:3]

    intro = (f"<p style='font-size:14px;color:{_C['ink']};margin-bottom:10px;'>"
             f"The only changes that earned a place this period &mdash; ranked by "
             f"impact, capped at three. Each shows its guard honestly.</p>")

    if not cards:
        # Honest no-change card with n_needed estimate.
        lo, hi = head_stat["ci"]
        n_needed = ""
        if lo is not None and head_stat["expR"] is not None:
            width = hi - lo
            # CI width ∝ 1/√n; to roughly halve width need ~4× N.
            target = max(0, int(n_book * 3))
            n_needed = f" Next checkpoint: ~{target} more trades to tighten this CI."
        body = (
            f"<div style='border:1px solid {_C['hair']};border-radius:8px;"
            f"background:{_C['surface']};padding:14px;'>"
            f"<b>No change earned a recommendation this period.</b> The edge is "
            f"<b>{head_stat['verdict']}</b>; keep collecting.{n_needed}</div>")
        return f'<div class="act">{intro}{body}{_act_action("No action — keep collecting; re-check next run.")}</div>'

    card_html = ""
    for c in cards:
        card_html += (
            f"<div style='border:1px solid {_C['hair']};border-radius:8px;"
            f"background:{_C['surface']};padding:12px 14px;margin-bottom:10px;'>"
            f"<div style='font-size:14px;font-weight:700;color:{_C['ink']};"
            f"margin-bottom:6px;'>{c['claim']}</div>"
            f"<div style='font-size:12px;color:{_C['ink2']};margin-bottom:3px;'>"
            f"<b>Guard:</b> {c['guard']}</div>"
            f"<div style='font-size:12px;color:{_C['ink2']};margin-bottom:3px;'>"
            f"<b>Replayed delta:</b> {c['delta']}</div>"
            f"<div style='font-size:12px;color:{_C['ink2']};'>"
            f"<b>Watch:</b> {c['watch']}</div></div>")
    action = _act_action("Try the top card first; confirm every guard on the next "
                         "(longer-history) run before making it live.")
    return f'<div class="act">{intro}{card_html}{action}</div>'


# ---------------------------------------------------------------------------
# Recipe ranking — the §11.1 fix (paired CI vs LIVE, trophy only when it clears)
# ---------------------------------------------------------------------------

def _recipe_ranked(group_sink: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """One ranked read per recipe over the headline population, PLUS the paired
    delta-vs-LIVE _stat_block for each non-LIVE recipe. This is the single source
    for both the recipe table (Act 5 feed / Act 6) and the winner rule."""
    prox = [r for r in (group_sink or [])
            if r.get("entry_zone") == "proximal"
            and not (r.get("ist_blocked") or r.get("weekend_blocked"))
            and r.get("exit_reason") not in _EXCLUDE_REASONS]
    if not prox:
        return []
    # Index per-trade R by join key so we can pair against LIVE.
    by_key: Dict[tuple, Dict[str, float]] = {}
    ts_by_key: Dict[tuple, Any] = {}
    for r in prox:
        key = (r.get("pair"), str(r.get("ob_timestamp")), r.get("direction"))
        by_key.setdefault(key, {})[r.get("config")] = float(r.get("r") or 0.0)
        ts_by_key[key] = r.get("alert_ts")

    by_cfg: Dict[str, List[Dict[str, Any]]] = {}
    for r in prox:
        by_cfg.setdefault(r.get("config"), []).append(r)

    out = []
    for cfg, rows in by_cfg.items():
        vals = [float(r.get("r") or 0.0) for r in rows]
        ts = [r.get("alert_ts") for r in rows]
        stat = _stat_block(vals, ts)
        w = sum(1 for v in vals if v > 0)
        l = sum(1 for v in vals if v < 0)
        wr = (100.0 * w / (w + l)) if (w + l) else None
        # Paired delta vs LIVE (same trades).
        paired = None
        if cfg != _EXIT_BASELINE_KEY:
            deltas, dts = [], []
            for key, cfgs in by_key.items():
                if cfg in cfgs and _EXIT_BASELINE_KEY in cfgs:
                    deltas.append(cfgs[cfg] - cfgs[_EXIT_BASELINE_KEY])
                    dts.append(ts_by_key[key])
            if deltas:
                pstat = _stat_block(deltas, dts)
                plo, phi = pstat["ci"]
                pq_ok = (pstat["pq_total"] == 0
                         or pstat["pq_agree"] / pstat["pq_total"] >= DRIVER_PQ_SIGN_FRAC)
                paired = {"avg_delta": pstat["expR"], "ci": pstat["ci"],
                          "n": pstat["n"], "pq_agree": pstat["pq_agree"],
                          "pq_total": pstat["pq_total"],
                          "promoted": (plo is not None and plo > 0 and pq_ok)}
        out.append({"cfg": cfg, "stat": stat, "wr": wr,
                    "total_r": round(sum(vals), 1), "paired": paired,
                    "n_book": len(rows)})
    out.sort(key=lambda x: (x["stat"]["expR"] if x["stat"]["expR"] is not None
                            else -1e9), reverse=True)
    return out


def _recipe_winner(group_sink: List[Dict[str, Any]]) -> Optional[Dict[str, Any]]:
    """The recipe that beats LIVE under the STRICT §11.1 rule: paired delta CI
    clears 0 AND per-quarter sign holds. None if none clears (trophy-less)."""
    ranked = _recipe_ranked(group_sink)
    winners = [x for x in ranked
               if x["paired"] and x["paired"]["promoted"]
               and x["paired"]["avg_delta"] and x["paired"]["avg_delta"] > 0]
    if not winners:
        return None
    winners.sort(key=lambda x: x["paired"]["avg_delta"], reverse=True)
    best = winners[0]
    return {"cfg": best["cfg"], "label": _EXIT_RECIPE_LABELS.get(best["cfg"], best["cfg"]),
            "avg_delta": best["paired"]["avg_delta"], "ci": best["paired"]["ci"],
            "n": best["paired"]["n"], "pq_agree": best["paired"]["pq_agree"],
            "pq_total": best["paired"]["pq_total"], "promoted": True,
            "n_book": best["n_book"]}


def _recipe_table_html(group_sink: List[Dict[str, Any]], risk_usd: float,
                       sink_exact: bool, headline_filled_n: Optional[int]) -> str:
    """The exit-recipe ranking table (§2 KEEP+FIX). Trophy ONLY on a recipe whose
    PAIRED delta-vs-LIVE CI clears 0 (§11.1) — not the pooled-average top row."""
    if not group_sink:
        return (f"<p style='color:{_C['muted']};'>No exit-lab rows this run.</p>")
    ranked = _recipe_ranked(group_sink)
    if not ranked:
        return (f"<p style='color:{_C['muted']};'>No proximal fills to study exits.</p>")

    # Count invariant (pipeline only; skipped for reconstructed preview sinks).
    if sink_exact and headline_filled_n is not None:
        recipe_n = next((x["n_book"] for x in ranked
                         if x["cfg"] == _EXIT_BASELINE_KEY), None)
        if recipe_n is None:
            recipe_n = max((x["n_book"] for x in ranked), default=0)
        if recipe_n != headline_filled_n:
            raise ValueError(
                f"Exit-table population ({recipe_n} fills per recipe) != headline "
                f"filled-proximal count ({headline_filled_n}). Sink and headline "
                f"scoring different books — refuse to emit a contaminated ranking.")

    winner = _recipe_winner(group_sink)
    win_cfg = winner["cfg"] if winner else None
    live_exr = next((x["stat"]["expR"] for x in ranked
                     if x["cfg"] == _EXIT_BASELINE_KEY), None)

    # Self-check line (moved to Act 6 caller; here we render just the table).
    body = ""
    for x in ranked:
        cfg, stat = x["cfg"], x["stat"]
        label = _EXIT_RECIPE_LABELS.get(cfg, cfg)
        lo, hi = stat["ci"]
        ci_s = f"[{lo:+.2f}, {hi:+.2f}]" if lo is not None else "thin"
        pq_s = f"{stat['pq_agree']}/{stat['pq_total']}q" if stat["pq_total"] else "—"
        if cfg == _EXIT_BASELINE_KEY:
            vs = "<i>incumbent</i>"
        elif x["paired"] and x["paired"]["avg_delta"] is not None:
            d = x["paired"]["avg_delta"]
            plo, phi = x["paired"]["ci"]
            cleared = x["paired"]["promoted"]
            color = _C["good"] if cleared else _C["muted"]
            vs = (f"<span style='color:{color};'>{d:+.2f}R</span>"
                  f"<span style='font-size:11px;color:{_C['muted']};'> "
                  f"CI[{plo:+.2f},{phi:+.2f}]</span>" if plo is not None
                  else f"<span style='color:{color};'>{d:+.2f}R</span>")
        else:
            vs = "—"
        is_win = cfg == win_cfg
        is_live = cfg == _EXIT_BASELINE_KEY
        bg = "#e3f4e3" if is_win else ("#fef3d9" if is_live else "")
        flag = " &#127942;" if is_win else ""
        total_pnl = x["total_r"] * risk_usd
        body += (
            f"<tr style='background:{bg};'>"
            f"<td style='text-align:left;'><b>{label}</b>{flag}</td>"
            f"<td style='text-align:right;'>{stat['n']}</td>"
            f"<td style='text-align:right;'>{_wr_str(x['wr'])}</td>"
            f"<td style='text-align:right;font-variant-numeric:tabular-nums;'>{_r(stat['expR'])}</td>"
            f"<td style='text-align:right;font-variant-numeric:tabular-nums;'>{x['total_r']:+.1f}R</td>"
            f"<td style='text-align:right;font-variant-numeric:tabular-nums;'>{_m(total_pnl)}</td>"
            f"<td style='text-align:right;font-size:11px;color:{_C['muted']};'>{ci_s}</td>"
            f"<td style='text-align:right;font-size:11px;color:{_C['muted']};'>{pq_s}</td>"
            f"<td style='text-align:left;'>{vs}</td></tr>")

    if winner:
        header_note = (f"<b>{winner['label']}</b> beats LIVE with confidence "
                       f"(paired CI clears zero).")
    else:
        header_note = "No recipe beats LIVE with confidence this period."
    note = (f"<p style='font-size:12px;color:{_C['muted']};margin-top:6px;'>"
            f"{header_note} <b>vs LIVE</b> is the paired per-trade delta (same "
            f"trades) with its own CI &mdash; a higher pooled average alone is not "
            f"a winner. Same-bar spikes are uncapturable on H1, so tight-TP recipes "
            f"flatter here vs reality.</p>")
    return (f"<div style='margin-bottom:4px;'>{_bankable_badge()} "
            f"<span style='font-size:11px;color:{_C['muted']};'>replayed orders "
            f"over the same post-fill bars</span></div>"
            f"<table width='100%'><thead><tr>"
            f"<th style='text-align:left;'>Exit recipe</th><th style='text-align:right;'>N</th>"
            f"<th style='text-align:right;'>WR</th><th style='text-align:right;'>Avg R</th>"
            f"<th style='text-align:right;'>Total R</th><th style='text-align:right;'>P&amp;L</th>"
            f"<th style='text-align:right;'>CI</th><th style='text-align:right;'>Qtrs</th>"
            f"<th style='text-align:left;'>vs LIVE</th>"
            f"</tr></thead><tbody>{body}</tbody></table>{note}")


def _recipe_selfcheck(group_sink: List[Dict[str, Any]]) -> str:
    """Baseline replay vs committed r_realised — the engine self-check (moved to
    Act 6 per §Act 6)."""
    prox = [r for r in (group_sink or [])
            if r.get("entry_zone") == "proximal"
            and not (r.get("ist_blocked") or r.get("weekend_blocked"))
            and r.get("exit_reason") not in _EXCLUDE_REASONS]
    base_rows = [r for r in prox if r.get("config") == _EXIT_BASELINE_KEY]
    if not base_rows:
        return ""
    repl = sum(float(r.get("r") or 0) for r in base_rows) / len(base_rows)
    comm = sum(float(r.get("committed_r") or 0) for r in base_rows) / len(base_rows)
    diff = abs(repl - comm)
    ok = diff <= 0.01
    color = _C["good"] if ok else _C["critical"]
    return (f"<li><b>Engine self-check:</b> baseline replay avg {_r(round(repl,3))} "
            f"vs committed {_r(round(comm,3))} (|diff| {diff:.3f}) &mdash; "
            f"<span style='color:{color};font-weight:700;'>"
            f"{'PASS' if ok else 'FAIL — side-channel not faithful'}</span>.</li>")


# ---------------------------------------------------------------------------
# ACT 6 — CAN I TRUST THIS RUN
# ---------------------------------------------------------------------------

# Placeholder token replaced post-gate-eval by run_backtest.py (§Act 6). Gates
# read the report's own headline (circular otherwise), so they run AFTER render.
_GATES_TOKEN = "<!--GATES-->"


def _act6_html(prox_trades: List[Dict[str, Any]], group_sink: List[Dict[str, Any]],
               group_meta: Dict[str, Any], raw_alert_n: int, filter_line: str,
               any_asserted: bool, preview: bool) -> str:
    filled = [t for t in prox_trades if _is_real_filled(t)]

    gates = (f"<p style='font-size:12px;color:{_C['muted']};'>"
             f"Gates not evaluated (preview).</p>" if preview else _GATES_TOKEN)

    # Funnel.
    ist_blocked = sum(1 for t in prox_trades if t.get("ist_blocked"))
    never_filled = sum(1 for t in prox_trades
                       if t.get("exit_reason") == "never_filled")
    resolved = len(filled)
    n_sim = len(prox_trades)
    funnel = (
        f"<p style='font-size:13px;color:{_C['ink2']};'>"
        f"<b>Funnel:</b> {raw_alert_n} raw alerts &rarr; {n_sim} simulated "
        f"&rarr; {n_sim - never_filled} filled &rarr; {resolved} resolved.</p>")

    # Exclusions.
    timeout = sum(1 for t in prox_trades if t.get("exit_reason") == "timeout")
    window_end = sum(1 for t in prox_trades if t.get("exit_reason") == "window_end")
    excl = (
        f"<p style='font-size:12px;color:{_C['muted']};'>"
        f"Audit-only (never in P&amp;L): never-filled {never_filled}, timeout "
        f"{timeout}, data-window-end {window_end}, IST-blocked {ist_blocked}. "
        f"Standing policy: unresolved positions never feed expectancy.</p>")

    same_bar = _same_bar_resolution_html(filled)
    selfcheck = _recipe_selfcheck(group_sink)
    validation = _validation_html(filled)

    settled = (
        f"<ul style='padding-left:18px;font-size:12px;color:{_C['ink2']};'>"
        f"<li>Confidence score: settled noise (Spearman ~0.05) &mdash; logged, not gated.</li>"
        f"<li>Confluences: settled non-predictive &mdash; see Excel.</li>"
        f"{selfcheck}</ul>")

    thin_warn = ""
    if not any_asserted:
        thin_warn = (f"<p style='font-size:13px;color:{_C['serious']};font-weight:600;"
                     f"margin-top:8px;'>&#9888; Nothing cleared the guard this period "
                     f"&mdash; the book is thin, not clean. Treat every number as "
                     f"directional.</p>")

    intro = (f"<p style='font-size:14px;color:{_C['ink']};margin-bottom:10px;'>"
             f"The honesty checks: did the gates pass, where did trades drop out, and "
             f"what the H1 resolution can and can't tell you.</p>")
    action = _act_action("If a gate failed above, the run is not trusted &mdash; "
                         "fix and re-run before acting on any number in this email.")
    return (f'<div class="act">{intro}'
            f'<h4>Gates</h4>{gates}{funnel}{excl}'
            f'<h4 style="margin-top:12px;">Backtest fidelity</h4>{same_bar}'
            f'<p style="font-size:13px;color:{_C["ink2"]};margin-top:8px;">{filter_line}</p>'
            f'{validation}{settled}{thin_warn}{action}</div>')


def _build_group_html(
    group_label: str,
    group_trades_all: List[Dict[str, Any]],
    group_meta: Dict[str, Any],
    risk_usd: float,
    out_dir: Path,
    html_filename: str,
    excel_filename: str,
    exit_lab_sink: List[Dict[str, Any]] = None,
    preview: bool = False,
    write_excel: bool = True,
    sink_exact: bool = True,
) -> Dict[str, Any]:
    """Build one HTML email + one Excel for a single pair group.

    Proximal-only (the live model — 50% mean entry is dead and shown nowhere in
    the body). Section order is decision-first (2026-06-30 overhaul):
      1. Headline (P&L + CI + per-quarter + verdict)
      2. Exit recipe ranking (the point — full per-recipe data for edge engine)
      3. Where losses / wins concentrate (the dynamic driver engine)
      4. Backtest fidelity (H1 same-bar honesty)
      5. By pair
      6. Structure events · Trend alignment · Killzone alignment · Break quality
      7. Filters + validation (one-line counters + the validation guard)
      8. Appendix (confluences, by-session, pair×session, counterfactual, vet)
    Every aggregate is recomputed against this group's pairs so the email is
    internally consistent. Returns the group summary for summary.json.
    """
    # `group_trades_all` is ALREADY the core set (below-floor + out-of-IST rows
    # removed run-wide upstream). The hard-block filter below is a defensive
    # no-op kept for symmetry.
    trades = [t for t in group_trades_all if not _is_hard_blocked(t)]
    blocked_trades = [t for t in group_trades_all if t.get("news_blocked")]
    kz_blocked_trades = [t for t in group_trades_all if t.get("killzone_blocked")]

    # Proximal is the only live model. 50% rows still exist in the simulator
    # output (and the Excel Zone Register, for reference) but never enter the
    # email body or this group's headline aggregates.
    prox_trades = [t for t in trades if t.get("entry_zone") == "proximal"]

    sb_prox     = _aggregate_for_exit(prox_trades, "r_realised",    risk_usd)
    sb_prox_tp1 = _aggregate_for_exit(prox_trades, "r_if_exit_tp1", risk_usd)
    sb_prox_tp2 = _aggregate_for_exit(prox_trades, "r_if_exit_tp2", risk_usd)
    fill_prox   = _fill_rate(trades, "proximal")

    # Reconciliation invariant: headline P&L must equal the per-trade sum.
    from_scoreboard = round(float(sb_prox.get("total_pnl_usd", 0)), 2)
    _filled = [t for t in prox_trades if _is_real_filled(t)]
    from_trades = round(sum(float(t.get("pnl_usd") or 0) for t in _filled), 2)
    if abs(from_scoreboard - from_trades) > 0.01:
        raise AssertionError(
            f"P&L reconciliation failed for {group_label}/proximal: "
            f"scoreboard={from_scoreboard} vs per-trade-sum={from_trades}.")

    # Excel: this group's filled+blocked rows (audit rows preserved). The Excel
    # still carries BOTH entry zones side by side as reference data. Skipped when
    # write_excel is False (fast preview iteration); the "attached" footer then
    # reflects that the workbook was not rebuilt for this render.
    _try_excel.last_error = None
    excel_ok = (write_excel and _try_excel(
        group_trades_all, out_dir / excel_filename, risk_usd=risk_usd) is not None)

    # Exit-lab sink scoped to this group's pairs (the sink rows carry `pair`).
    allowed_pairs = set(group_meta.get("pairs", []))
    group_sink = [r for r in (exit_lab_sink or [])
                  if r.get("pair") in allowed_pairs] if allowed_pairs else []

    # Filter line (2026-07): news blackout and the killzone filter are BOTH off
    # live — neither gates an alert — so their old "removed 0 / dropped X" lines
    # were dead information and are gone. `killzone_dropped_alerts` in meta only
    # COUNTS alerts formed outside a killzone window; those trades are still
    # simulated and STILL feed the headline (verified: killzone_blocked never
    # appears in _headline_exclusion). So a non-zero meta count is normal and NOT
    # a problem. The real red flag is a trade being EXCLUDED from P&L for a
    # killzone reason — that would mean killzone got re-wired into the gate. We
    # check the actual exclusion tags, not the label counter.
    kz_excluded = sum(
        1 for t in prox_trades
        if "killzone" in _headline_exclusion(t).lower())
    if kz_excluded > 0:
        filter_line = (
            f"<b style='color:#e74c3c;'>&#9888; PROBLEM: {kz_excluded} trade(s) "
            f"were EXCLUDED from P&amp;L for a killzone reason.</b> Killzone is "
            f"supposed to be a display signal only — it must NOT gate P&amp;L. "
            f"This means killzone got re-wired into the headline exclusion path — "
            f"investigate before trusting this run.")
    else:
        filter_line = ("Validation guard: no filter (news / killzone) excluded "
                       "any trade from P&amp;L this period — both are display-only "
                       "as intended.")

    pairs_str  = ", ".join(group_meta.get("pairs", []))
    regime_str = group_meta.get("regime", "")

    # Chart file prefix per book (report_forex.html -> chart_forex).
    chart_prefix = "chart_" + html_filename.replace("report_", "").replace(".html", "")
    raw_alert_n = int(group_meta.get("raw_alert_n") or 0)
    # Real run-id for the run-log footer (out_dir is .../<run_id>/preview in
    # preview mode, .../<run_id> in pipeline mode).
    run_id_str = out_dir.parent.name if preview else out_dir.name

    # --- Assemble the six acts (deterministic Python) ----------------------
    # Headline stat drives the Act 1 banner + Act 5's no-change fallback.
    _hfilled = [t for t in prox_trades if _is_real_filled(t)]
    _hvals = [float(t.get("r_realised") or 0.0) for t in _hfilled]
    _hts = [t.get("alert_ts") for t in _hfilled]
    head_stat = _stat_block(_hvals, _hts)

    act1 = _act1_html(prox_trades, sb_prox, fill_prox, group_meta, risk_usd)
    act2 = _act2_html(prox_trades, out_dir, chart_prefix, risk_usd)
    act3 = _act3_html(prox_trades)
    act_pair = _pair_table_html(prox_trades, risk_usd)
    act4 = _act4_html(prox_trades, group_sink, risk_usd)
    act5 = _act5_html(prox_trades, group_sink, act4["candidates"], head_stat, risk_usd)
    recipe_tbl = _recipe_table_html(
        group_sink, risk_usd, sink_exact,
        headline_filled_n=len([t for t in prox_trades if _is_real_filled(t)
                               and not (t.get("ist_blocked") or t.get("weekend_blocked"))]))

    # "Anything asserted?" = did any act promote a driver / recipe winner? Drives
    # the Act 6 thin-warning.
    _base_exr = (sum(_hvals) / len(_hvals)) if _hvals else 0.0
    any_asserted = bool(
        [r for r in _cause_rows(_hfilled, _base_exr, "win") if r["promoted"]]
        or [r for r in _cause_rows(_hfilled, _base_exr, "loss") if r["promoted"]]
        or _recipe_winner(group_sink)
        or head_stat["verdict"] in ("edge", "loser"))
    act6 = _act6_html(prox_trades, group_sink, group_meta, raw_alert_n,
                      filter_line, any_asserted, preview)

    if not write_excel:
        excel_note = "workbook not rebuilt for this preview render"
    elif excel_ok:
        excel_note = "every filled trade for this group, plain-English headers."
    else:
        _err = getattr(_try_excel, "last_error", None) or "no filled trades to write"
        excel_note = (
            f"<span style='color:#d03b3b;'>FAILED — {_err}</span>")

    html = f"""<!DOCTYPE html>
<html lang="en"><head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, Arial, sans-serif;
          background: {_C['page']}; color: {_C['ink']}; font-size: 14px; line-height: 1.55; }}
  .wrap {{ max-width: 680px; margin: 0 auto; background: {_C['surface']}; }}
  .top-band {{ background: {_C['ink']}; color: #fff; padding: 18px 24px; }}
  .top-band h1 {{ font-size: 17px; font-weight: 700; margin-bottom: 4px; }}
  .top-band .meta {{ font-size: 12px; color: #c9c8c2; }}
  .act {{ padding: 22px 24px; border-bottom: 1px solid {_C['hair']}; }}
  .act-head {{ font-size: 12px; font-weight: 800; text-transform: uppercase;
               letter-spacing: 0.08em; color: {_C['muted']}; margin-bottom: 12px; }}
  h4 {{ font-size: 12px; font-weight: 700; color: {_C['ink2']}; margin: 14px 0 6px;
        text-transform: uppercase; letter-spacing: 0.04em; }}
  table {{ width: 100%; border-collapse: collapse; font-size: 12px; margin-top: 4px; }}
  th {{ background: {_C['neutral']}; color: {_C['ink2']}; padding: 6px 8px;
        font-weight: 700; font-size: 11px; border-bottom: 1px solid {_C['baseline']}; }}
  td {{ padding: 6px 8px; border-bottom: 1px solid {_C['hair']}; color: {_C['ink']}; }}
  tr:last-child td {{ border-bottom: none; }}
  a {{ color: {_C['blue']}; }}
  .footer {{ padding: 16px 24px; background: {_C['page']}; font-size: 11px; color: {_C['muted']}; }}
</style>
</head>
<body>
<div class="wrap">

<div class="top-band">
  <h1>{group_label} &mdash; {group_meta.get('start')} to {group_meta.get('end')}</h1>
  <div class="meta">
    {pairs_str} &nbsp;&middot;&nbsp; 1R = ${risk_usd:.0f} &nbsp;&middot;&nbsp;
    Regime: {regime_str} &nbsp;&middot;&nbsp; H1 bars only, proximal entry, spread modelled on the stop
  </div>
</div>

<div class="act"><div class="act-head">Act 1 &middot; Pulse</div>{act1}</div>
<div class="act"><div class="act-head">Act 2 &middot; What the year looked like</div>{act2}</div>
<div class="act"><div class="act-head">Act 3 &middot; Where the edge comes from</div>{act3}</div>
<div class="act"><div class="act-head">By pair</div>{act_pair}</div>
<div class="act"><div class="act-head">Act 4 &middot; Where it leaks</div>{act4['html']}</div>
<div class="act"><div class="act-head">Act 5 &middot; What to change next</div>
  <div style="margin-bottom:12px;">{act5}</div>
  <h4>Exit recipe ranking &mdash; which exit pays most</h4>
  <p style="font-size:12px;color:{_C['ink2']};margin-bottom:4px;">Every recipe replayed
  over the same post-fill bars as the live trade &mdash; the bankable menu the cards draw from.</p>
  {recipe_tbl}
</div>
<div class="act"><div class="act-head">Act 6 &middot; Can I trust this run</div>{act6}</div>

<div class="footer">
  <b>Attached ({excel_filename}):</b>
  Trades tab &mdash; {excel_note}
  Zone Register (one row per OB), Second Look (flagged trades), and the relocated
  reference tabs: Counterfactual (with CI/quarters/ci_cleared guard columns),
  Confluences, Setup badges, Break-quality ATR ladder, Pair&times;session matrix.
  <br><br>
  <b>Limitations:</b> spread modelled on the stop (worst-case fill); slippage and
  swap not modelled; exits at H1 bar boundaries; same-bar SL+TP collision resolves
  SL-first (pessimistic).
  <br><br>
  <b>Run log:</b> <code>backtest/results/{run_id_str}/</code> &middot;
  verify with <code>git log --grep="Backtest logs: {run_id_str}"</code>
</div>

</div></body></html>"""

    if not preview:
        _assert_body_size(html, html_filename)
    (out_dir / html_filename).write_text(html, encoding="utf-8")

    return {
        "label": group_label,
        "pairs": group_meta.get("pairs", []),
        # Proximal only — 50% is dead and no longer reported. Keys kept stable
        # for the registry / aggregate_runs readers (they read *proximal*).
        "scoreboards": {
            "proximal_realised":  sb_prox,
            "proximal_exit_tp1":  sb_prox_tp1,
            "proximal_exit_tp2":  sb_prox_tp2,
        },
        "fill_rate_proximal": fill_prox,
        "news_blocked_trade_rows": len(blocked_trades),
        "ist_blocked_trade_rows":  0,
        "killzone_dropped_alerts": int(group_meta.get("killzone_dropped_alerts") or 0),
        "score_floor":                  SCORE_FLOOR,
        "below_score_floor_trade_rows": 0,
        "html_file":  html_filename,
        "excel_file": excel_filename,
    }


# ---------------------------------------------------------------------------
# Main report writer
# ---------------------------------------------------------------------------

def write_h1_only_report(
    run_id: str,
    trades: List[Dict[str, Any]],
    raw_alerts: List[Dict[str, Any]],
    meta: Dict[str, Any],
    risk_usd: float = 250.0,
    out_root: Path = None,
    exit_lab_sink: List[Dict[str, Any]] = None,
    preview: bool = False,
    write_excel: bool = True,
    sink_exact: bool = True,
) -> Path:
    """Build the per-group HTML reports (+ summary/CSV/Excel in pipeline mode).

    `preview=True` (used by render_report.py): rows already carry `setup_id`, so
    trade IDs are NOT re-claimed (never burn IDs on a preview); HTML is written to
    a `preview/` subdir so the committed reports are never overwritten; the on-disk
    summary.json / trades.csv are NOT rewritten. Nothing in this path commits,
    pushes, or emails. `write_excel=False` additionally skips the (slow) Excel
    rebuild during fast design iteration.
    """
    base = out_root if out_root is not None else (Path(__file__).parent / "results")
    run_dir = Path(base) / run_id
    out_dir = (run_dir / "preview") if preview else run_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    # HARD EXCLUSION (one only, as of 2026-06-30): a trade is pulled out of the
    # run ENTIRELY -- no table, metric, score bucket, Excel/CSV row, or per-group
    # report -- if it is outside the IST trading window (live never scans before
    # 09:00 IST, so such an alert could never have happened live). It shows up in
    # EXACTLY ONE place: a small audit list (pair / time / score / would-be R).
    # SCORE NO LONGER EXCLUDES (trader decision 2026-06-30): we want the true,
    # unfiltered P&L of every OB-touch, so audit_below_floor is always empty now
    # (the score column is still recorded per row for analysis).
    raw_trades = list(trades)
    audit_below_floor = []  # score floor retired — no row is excluded on score
    audit_ist = [t for t in raw_trades if t.get("ist_blocked")]
    # CORE set: survives the IST gate. THE set every downstream path sees.
    trades_all = [t for t in raw_trades if not _is_hard_blocked(t)]
    trades = list(trades_all)
    # Global non-resetting trade ID. Claimed atomically from the counter file
    # so IDs are unique across every run forever (A0001…Z9999, 260 000 slots).
    # Stamped before any consumer reads the dicts; every output (CSV, Excel,
    # HTML, email vet-review) sees the same ID for the same trade.
    # Preview mode: rows already carry setup_id (read from trades.csv); claiming
    # here would permanently burn IDs for a throwaway render. Only the pipeline
    # path stamps IDs.
    if trades_all and not preview:
        _first_id = _claim_trade_ids(len(trades_all))
        for _i, _t in enumerate(trades_all):
            _t["setup_id"] = _int_to_trade_id(_first_id + _i)
    blocked_trades = [t for t in trades_all if t.get("news_blocked")]
    kz_blocked_trades = [t for t in trades_all if t.get("killzone_blocked")]

    # Proximal is the only entry zone (50% mean entry removed 2026-07). This
    # filter is now a defensive no-op — every row is proximal — kept so any
    # future re-introduction of a second zone can't silently leak into metrics.
    prox_trades = [t for t in trades if t.get("entry_zone") == "proximal"]

    # r_realised = the LIVE policy (TP1 + break-even at +1R). Every per-pair /
    # per-session / score breakdown uses it so headline + breakdowns reconcile.
    sb_prox    = _aggregate_for_exit(prox_trades, "r_realised",     risk_usd)
    sb_prox_tp2 = _aggregate_for_exit(prox_trades, "r_if_exit_tp2", risk_usd)
    sb_prox_tp1 = _aggregate_for_exit(prox_trades, "r_if_exit_tp1", risk_usd)

    pp_prox   = _per_pair_breakdown(prox_trades,  "r_realised", risk_usd)
    ss_prox   = _per_session_breakdown(prox_trades, "r_realised", risk_usd)
    fill_prox = _fill_rate(trades, "proximal")

    # Score buckets over the CORE set only (no sub-floor buckets by design).
    score_buckets    = _score_buckets(prox_trades, "r_realised")
    exit_counts_prox = _exit_reason_counts(prox_trades)

    # News-blackout audit. Every blocked row is listed (without the trade
    # outcome — that would imply we counted it; we did not). The list is
    # intentionally redundant with the Excel sheet so the user can grep
    # summary.json for a specific event without opening the workbook.
    blocked_audit = [
        {
            "pair":     t.get("pair"),
            "alert_ts": t.get("alert_ts"),
            "entry_zone": t.get("entry_zone"),
            "event_title":    t.get("news_event_title"),
            "event_currency": t.get("news_event_currency"),
            "event_source":   t.get("news_event_source"),
            "event_ts":       t.get("news_event_ts"),
        }
        for t in blocked_trades
    ]

    # Small audit lists for the two HARD exclusions. These rows appear nowhere
    # else in the run; this is their only home in summary.json. We carry the
    # would-have R so the user can see the cost/benefit of each gate.
    def _excl_audit(rows):
        out = []
        for t in rows:
            if t.get("entry_zone") != "proximal":
                continue  # one row per alert (proximal is canonical)
            real = t.get("exit_reason") not in _EXCLUDE_REASONS
            out.append({
                "pair":         t.get("pair"),
                "alert_ts":     t.get("alert_ts"),
                "score":        t.get("score"),
                "exit_reason":  t.get("exit_reason"),
                "would_have_r": (round(float(t.get("r_realised", 0.0)), 3)
                                 if real else None),
            })
        return out
    below_floor_audit = _excl_audit(audit_below_floor)
    ist_audit         = _excl_audit(audit_ist)

    summary = {
        "run_id":              run_id,
        "meta":                meta,
        "risk_per_trade_usd":  risk_usd,
        "total_trade_rows":    len(trades),
        "fill_rate_proximal":  fill_prox,
        "exit_reason_counts_proximal": exit_counts_prox,
        # Scoreboards under the exit policies. r_realised is the LIVE policy
        # (TP1 + break-even at +1R); the tp1/tp2 columns are pure hypotheticals
        # and named accordingly. Proximal only — 50% mean entry is dead.
        "scoreboards": {
            "proximal_realised":  sb_prox,
            "proximal_exit_tp1":  sb_prox_tp1,
            "proximal_exit_tp2":  sb_prox_tp2,
        },
        # Per-pair / per-session breakdowns use r_realised so they reconcile
        # to the headline total. Keys carry the column name explicitly.
        "per_pair_proximal_realised": pp_prox,
        "per_session_proximal_realised": ss_prox,
        "score_buckets_proximal_realised": score_buckets,
        # News blackout is INFORMATIONAL ONLY (live never suppresses on news);
        # these rows ARE counted in the metrics above. Listed for reference.
        "news_blocked_trade_rows": len(blocked_trades),
        "news_blocked_audit":      blocked_audit,
        # Killzone is a scoring signal, not a gate. Counts come from run_backtest
        # meta -- the dropped alerts never produced trade rows.
        "killzone_dropped_alerts":  int(meta.get("killzone_dropped_alerts") or 0),
        "killzone_drops_by_pair":   dict(meta.get("killzone_drops_by_pair") or {}),
        "killzone_windows_by_pair": dict(meta.get("killzone_windows_by_pair") or {}),
        # TWO HARD EXCLUSIONS — these rows appear NOWHERE else in the run, only
        # in their own audit list here (with would-have R). Below score floor,
        # and outside the IST trading window.
        "score_floor":                  SCORE_FLOOR,
        # *_trade_rows = all proximal rows; *_audit = one row per alert with
        # would-have R, for human reading.
        "below_score_floor_trade_rows": len(audit_below_floor),
        "below_score_floor_audit":      below_floor_audit,
        "ist_blocked_trade_rows":       len(audit_ist),
        "ist_blocked_audit":            ist_audit,
    }

    # Reconciliation invariant. The headline P&L (sb_prox) MUST equal the sum
    # of per-trade pnl_usd for the same population. If this fails, the email is
    # publishing inconsistent numbers -- fail loud.
    def _reconcile(zone_label, scoreboard, zone_trades):
        from_scoreboard = round(float(scoreboard.get("total_pnl_usd", 0)), 2)
        filled = [t for t in zone_trades if _is_real_filled(t)]
        from_trades = round(sum(float(t.get("pnl_usd") or 0) for t in filled), 2)
        if abs(from_scoreboard - from_trades) > 0.01:
            raise AssertionError(
                f"P&L reconciliation failed for {zone_label}: "
                f"scoreboard={from_scoreboard} vs per-trade-sum={from_trades}. "
                f"This means the email headline contradicts the trade rows. "
                f"Fix the aggregator before shipping the report."
            )
    _reconcile("proximal", sb_prox, prox_trades)

    # Files. Use trades_all for CSV and Excel so blocked rows appear in
    # the audit outputs (column news_blocked + event metadata). Metrics
    # were computed above on the filtered `trades`, so summary stats are
    # unaffected by this. Preview mode NEVER rewrites the committed run
    # artifacts (trades.csv / raw_alerts.jsonl / summary.json) — the harness
    # reads them, it does not regenerate them.
    if not preview:
        _trades_csv(trades_all, out_dir / "trades.csv")
        # News columns are RUN-PRODUCED (Part D, 2026-07-16): enrich the CSV in
        # place right after writing it, so the 5 news_* columns ship in the same
        # artifact. Fail-loud offline if the events file is missing/short.
        _enrich_news_columns(out_dir / "trades.csv")
        _try_excel(trades_all, out_dir / "trades.xlsx", risk_usd=risk_usd)
        with open(out_dir / "raw_alerts.jsonl", "w") as f:
            for a in raw_alerts:
                f.write(json.dumps(a, default=str) + "\n")
        with open(out_dir / "summary.json", "w") as f:
            json.dump(summary, f, indent=2, default=str)

    # NOTE (2026-06-30 overhaul): the combined `report.html` is GONE. It was
    # never emailed (reporting_email.py sends only the two per-group HTMLs) and
    # rendered the 13-section zone block twice. The two group emails below are
    # now the only HTML output. summary.json / trades.csv / trades.xlsx are
    # unchanged for update_registry.py + aggregate_runs.py.

    # ---- Per-group emails (Book A: original FX+Gold | Book B: new FX+BTC) ----
    # The ONLY HTML output. Each is proximal-only, decision-first, with every
    # aggregate recomputed against the group's pair subset. by_group KEYS and
    # filenames are kept stable (registry / aggregate_runs depend on them).
    forex_trades_all = [t for t in trades_all if t.get("pair") in FOREX_PAIRS]
    indcom_trades_all = [t for t in trades_all if t.get("pair") in INDEX_COMMODITY_PAIRS]

    # Raw-alert counts per book for the Act 6 funnel (raw alerts -> simulated).
    def _raw_n(allowed):
        return sum(1 for a in (raw_alerts or []) if a.get("pair") in allowed)

    forex_meta  = _filter_meta_by_pairs(meta, FOREX_PAIRS)
    indcom_meta = _filter_meta_by_pairs(meta, INDEX_COMMODITY_PAIRS)
    forex_meta["raw_alert_n"]  = _raw_n(FOREX_PAIRS)
    indcom_meta["raw_alert_n"] = _raw_n(INDEX_COMMODITY_PAIRS)

    by_group = {}
    if forex_trades_all:
        by_group["forex"] = _build_group_html(
            "Original (FX majors + Gold)", forex_trades_all, forex_meta, risk_usd,
            out_dir, "report_forex.html", "forex_trades.xlsx",
            exit_lab_sink=exit_lab_sink, preview=preview, write_excel=write_excel,
            sink_exact=sink_exact,
        )
    if indcom_trades_all:
        by_group["gold_nas"] = _build_group_html(
            "New (new FX + BTC)", indcom_trades_all, indcom_meta, risk_usd,
            out_dir, "report_gold_nas.html", "nas_xau_trades.xlsx",
            exit_lab_sink=exit_lab_sink, preview=preview, write_excel=write_excel,
            sink_exact=sink_exact,
        )

    if by_group and not preview:
        # Fold per-group summaries into the combined summary.json so future
        # tooling can read the partitioned numbers without re-running anything.
        # Preview never touches the committed summary.json.
        summary["by_group"] = by_group
        with open(out_dir / "summary.json", "w") as f:
            json.dump(summary, f, indent=2, default=str)

    return out_dir
