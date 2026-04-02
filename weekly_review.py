import yfinance as yf
import pandas as pd
import json, os, smtplib, requests, time
from datetime import datetime, timedelta
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open("config.json") as f:
    config = json.load(f)

GEMINI_KEY    = os.environ["GEMINI_API_KEY"]
GMAIL_ADDRESS = os.environ["GMAIL_ADDRESS"]
GMAIL_PASS    = os.environ["GMAIL_APP_PASSWORD"]
ALERT_EMAIL   = os.environ["ALERT_EMAIL"]

RISK_PER_TRADE = config["account"]["balance"] * config["account"]["risk_percent"] / 100
MIN_SCORE      = config.get("scoring", {}).get("min_score_to_send", 7)
FATIGUE_THRESH = config.get("scoring", {}).get("zone_fatigue_threshold", 3)

ALERT_LOG_FILE = "alert_log.json"
try:
    with open(ALERT_LOG_FILE) as f:
        alert_log = json.load(f)
    print(f"  Loaded {len(alert_log)} total log entries")
except Exception:
    alert_log = []
    print("  No alert log found")

def ist_now():
    return datetime.utcnow() + timedelta(hours=5, minutes=30)

# ── Helpers ────────────────────────────────────────────────────────────────────
def clean_df(df):
    if df is None or df.empty:
        return None
    if isinstance(df.columns, pd.MultiIndex):
        df.columns = [col[0] for col in df.columns]
    return df

def get_session(utc_hour):
    if   2  <= utc_hour < 8:  return "Asian"
    elif 8  <= utc_hour < 13: return "London"
    elif 13 <= utc_hour < 21: return "New York"
    else:                     return "Off-Hours"

def get_pair_type(pair_name):
    """Get asset class from config for a pair."""
    for p in config["pairs"]:
        if p["name"] == pair_name:
            return p.get("pair_type", "forex")
    return "forex"

# ── Gemini call ────────────────────────────────────────────────────────────────
def call_gemini(prompt, retries=2):
    url  = (f"https://generativelanguage.googleapis.com/v1beta/"
            f"models/gemini-2.5-flash:generateContent?key={GEMINI_KEY}")
    body = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {
            "thinkingConfig": {
                "thinkingBudget": 0
            },
            "maxOutputTokens": 2500
        }
    }
    for attempt in range(retries + 1):
        try:
            r      = requests.post(url, json=body, timeout=120)
            result = r.json()
            if "candidates" not in result:
                err_code = result.get("error", {}).get("code", 0)
                if err_code == 429 and attempt < retries:
                    wait = 10 * (attempt + 1)
                    print(f"    Gemini rate limit — retrying in {wait}s...")
                    time.sleep(wait)
                    continue
                raise ValueError(f"No candidates (code {err_code})")
            raw = result["candidates"][0]["content"]["parts"][0]["text"].strip()
            if raw.startswith("```"):
                raw = raw.split("\n", 1)[1].rsplit("```", 1)[0]
            return json.loads(raw), None
        except Exception as e:
            if attempt < retries:
                print(f"    Gemini attempt {attempt+1} failed ({e}), retrying...")
                time.sleep(5)
            else:
                return None, f"Gemini error: {str(e)}"

# ── Entry-gated outcome check ──────────────────────────────────────────────────
def check_entry_gated_outcome(alert):
    pair   = alert['pair']
    symbol = next((p['symbol'] for p in config['pairs'] if p['name'] == pair), None)
    if not symbol:
        return "pending", None
    try:
        bias = str(alert.get('bias', 'LONG')).upper()
        if bias not in ('LONG', 'SHORT'):
            return "pending", None
        sl  = float(alert.get('sl',  0) or 0)
        tp1 = float(alert.get('tp1', 0) or 0)
        if sl <= 0 or tp1 <= 0:
            return "pending", None
        try:
            entry = float(str(alert.get('entry', '0')).split('-')[0].strip() or 0)
        except Exception:
            entry = 0
        if entry <= 0:
            return "pending", None

        alert_time = datetime.strptime(alert['timestamp_utc'], "%Y-%m-%d %H:%M")
        df = clean_df(yf.download(symbol,
            start=(alert_time - timedelta(hours=1)).strftime('%Y-%m-%d'),
            interval="15m", progress=False))
        if df is None:
            return "pending", None

        entry_reached = False
        for ts, row in df.iterrows():
            try:
                ts_n = ts.replace(tzinfo=None) if hasattr(ts, 'tzinfo') and ts.tzinfo else ts
                if ts_n < alert_time:
                    continue
                h = float(row['High'])
                l = float(row['Low'])
            except Exception:
                continue

            if not entry_reached:
                if bias == "LONG"  and l <= sl:  return "not_triggered", None
                if bias == "SHORT" and h >= sl:  return "not_triggered", None
                if bias == "LONG"  and l <= entry <= h: entry_reached = True
                if bias == "SHORT" and l <= entry <= h: entry_reached = True
            else:
                if bias == "LONG":
                    if l <= sl:  return "loss",    sl
                    if h >= tp1: return "win_tp1", tp1
                elif bias == "SHORT":
                    if h >= sl:  return "loss",    sl
                    if l <= tp1: return "win_tp1", tp1

        if not entry_reached:
            age_h = (datetime.utcnow() - alert_time).total_seconds() / 3600
            if age_h >= 48:
                return "not_triggered", None
        return "pending", None
    except Exception as e:
        print(f"  Outcome check error {pair}: {e}")
        return "pending", None

# ── Outcome update pipeline ───────────────────────────────────────────────────
def update_outcomes():
    updated = 0
    for alert in alert_log:
        if alert.get('outcome') in ('win_tp1', 'loss', 'not_triggered', 'invalidated'):
            continue
        try:
            alert_time = datetime.strptime(alert['timestamp_utc'], "%Y-%m-%d %H:%M")
            age_hours  = (datetime.utcnow() - alert_time).total_seconds() / 3600
            if not (4 <= age_hours <= 336):
                continue
            sl  = float(alert.get('sl',  0) or 0)
            tp1 = float(alert.get('tp1', 0) or 0)
            if sl <= 0 or tp1 <= 0:
                print(f"    Skipping {alert['pair']} ({alert['timestamp_utc']}) — no SL/TP.")
                continue
            print(f"    Scanning {alert['pair']} ({alert['timestamp_utc']})...")
            outcome, outcome_price = check_entry_gated_outcome(alert)
            if outcome != 'pending':
                alert['outcome']            = outcome
                alert['outcome_price']      = outcome_price
                alert['outcome_checked_at'] = datetime.utcnow().strftime("%Y-%m-%d %H:%M")
                print(f"      → {outcome}" + (f" at {outcome_price}" if outcome_price else ""))
                updated += 1
            else:
                print(f"      → Still pending.")
        except Exception as e:
            print(f"  Update error {alert.get('pair','?')}: {e}")

    with open(ALERT_LOG_FILE, "w") as f:
        json.dump(alert_log, f, indent=2)
    print(f"  Updated {updated} outcomes (entry-gated — 0 Gemini calls)")

# ── Weekly alerts window (FIXED date label) ───────────────────────────────────
def get_weekly_alerts():
    """
    Returns alerts from Monday 00:00 IST through Friday 23:59 IST.
    Saturday/Sunday runs always show the JUST-ENDED Mon-Fri.
    Mid-week runs show Mon-now.
    The review_period label is derived from the SAME dates as the data.
    """
    ist = ist_now()
    days_to_monday = ist.weekday()
    this_monday_ist = ist.replace(hour=0, minute=0, second=0, microsecond=0) \
                      - timedelta(days=days_to_monday)
    this_friday_ist = this_monday_ist + timedelta(days=4, hours=23, minutes=59, seconds=59)

    week_start_utc = this_monday_ist - timedelta(hours=5, minutes=30)
    week_end_ist   = min(ist, this_friday_ist)
    week_end_utc   = week_end_ist - timedelta(hours=5, minutes=30)

    weekly = [a for a in alert_log
              if a.get('alert_type') == 'zone'
              and week_start_utc
              <= datetime.strptime(a['timestamp_utc'], "%Y-%m-%d %H:%M")
              <= week_end_utc]

    end_label = "Fri" if ist >= this_friday_ist else ist.strftime('%a')
    # Store review period for email label — derived from SAME dates
    global review_period_label
    review_period_label = f"{this_monday_ist.strftime('%d %b')} – {week_end_ist.strftime('%d %b %Y')}"

    print(f"  {len(weekly)} zone alerts | {this_monday_ist.strftime('%d %b')} (Mon) → "
          f"{week_end_ist.strftime('%d %b')} ({end_label})")
    return weekly

review_period_label = "—"

# ── Estimated P&L ─────────────────────────────────────────────────────────────
def estimate_pnl(alert):
    outcome = alert.get('outcome', 'pending')
    if outcome not in ('win_tp1', 'loss'):
        return 0.0, "—"
    try:
        bias  = alert.get('bias', '')
        entry = float(str(alert.get('entry', '0')).split('-')[0].strip() or 0)
        sl    = float(alert.get('sl',  0) or 0)
        tp1   = float(alert.get('tp1', 0) or 0)
        if entry <= 0 or sl <= 0 or tp1 <= 0:
            return 0.0, "—"
        risk_pts   = (entry - sl)    if bias == "LONG" else (sl    - entry)
        reward_pts = (tp1   - entry) if bias == "LONG" else (entry - tp1)
        if risk_pts <= 0:
            return 0.0, "—"
        rr = reward_pts / risk_pts
        if outcome == 'win_tp1':
            return round(RISK_PER_TRADE * rr, 2), f"+{round(reward_pts, 5)}"
        else:
            return round(-RISK_PER_TRADE, 2), f"-{round(risk_pts, 5)}"
    except Exception:
        return 0.0, "—"


# ── Gemini weekly analysis ────────────────────────────────────────────────────
def build_weekly_analysis(weekly_alerts, wins, losses, invalidated_count,
                          not_triggered_count, pending, win_rate, pair_stats,
                          asset_class_stats, alert2_triggered, alert2_invalidated):
    triggered = [a for a in weekly_alerts if a.get('outcome') in ('win_tp1', 'loss')]

    session_stats = {}
    for a in triggered:
        try:
            utc_hour = datetime.strptime(a['timestamp_utc'], "%Y-%m-%d %H:%M").hour
            session  = get_session(utc_hour)
            if session not in session_stats:
                session_stats[session] = {'wins': 0, 'losses': 0}
            if a.get('outcome') == 'win_tp1': session_stats[session]['wins']   += 1
            else:                             session_stats[session]['losses'] += 1
        except Exception:
            pass
    for s in session_stats.values():
        w, l = s['wins'], s['losses']
        s['win_rate'] = round(w/(w+l)*100, 1) if (w+l) > 0 else None

    hour_buckets = {}
    for a in triggered:
        try:
            utc_dt   = datetime.strptime(a['timestamp_utc'], "%Y-%m-%d %H:%M")
            ist_hour = (utc_dt + timedelta(hours=5, minutes=30)).hour
            bucket   = f"{ist_hour:02d}:00"
            if bucket not in hour_buckets:
                hour_buckets[bucket] = {'wins': 0, 'losses': 0}
            if a.get('outcome') == 'win_tp1': hour_buckets[bucket]['wins']   += 1
            else:                             hour_buckets[bucket]['losses'] += 1
        except Exception:
            pass
    loss_clusters = [f"{h} IST ({v['losses']} losses)"
                     for h, v in sorted(hour_buckets.items()) if v['losses'] >= 2]

    # Zone fatigue — computed in Python
    zone_alert_counts = {}
    for a in weekly_alerts:
        p = a.get('pair', '')
        zone_alert_counts[p] = zone_alert_counts.get(p, 0) + 1
    python_zone_flags = [
        f"{pair}: {count} zone alerts this week"
        for pair, count in zone_alert_counts.items()
        if count >= FATIGUE_THRESH + 2  # Only flag if significantly above threshold
    ]

    # Geo split
    geo_alerts   = [a for a in weekly_alerts if a.get('geo_flag', False)]
    clean_alerts = [a for a in weekly_alerts if not a.get('geo_flag', False)]
    geo_wins     = sum(1 for a in geo_alerts   if a.get('outcome') == 'win_tp1')
    geo_losses   = sum(1 for a in geo_alerts   if a.get('outcome') == 'loss')
    cl_wins      = sum(1 for a in clean_alerts if a.get('outcome') == 'win_tp1')
    cl_losses    = sum(1 for a in clean_alerts if a.get('outcome') == 'loss')
    geo_wr = round(geo_wins/(geo_wins+geo_losses)*100, 1) if (geo_wins+geo_losses) > 0 else 0
    cl_wr  = round(cl_wins/(cl_wins+cl_losses)*100, 1)   if (cl_wins+cl_losses)   > 0 else 0

    # News-flagged trades (separate from geo_flag)
    news_alerts  = [a for a in weekly_alerts if a.get('news_flag', 'none').lower() != 'none']
    news_wins    = sum(1 for a in news_alerts if a.get('outcome') == 'win_tp1')
    news_losses  = sum(1 for a in news_alerts if a.get('outcome') == 'loss')
    news_wr      = round(news_wins/(news_wins+news_losses)*100, 1) if (news_wins+news_losses) > 0 else 0

    alert_summary = []
    for a in triggered:
        alert_summary.append({
            "pair":       a.get('pair', ''),
            "pair_type":  a.get('pair_type', get_pair_type(a.get('pair', ''))),
            "ist_time":   a.get('ist_time', ''),
            "bias":       a.get('bias', ''),
            "confidence": a.get('confidence_score', 0),
            "score_breakdown": a.get('score_breakdown', {}),
            "confluences": a.get('confluences', []),
            "outcome":    a.get('outcome', 'pending'),
            "geo_flag":   a.get('geo_flag', False),
            "news_flag":  a.get('news_flag', 'none'),
        })

    prompt = f"""You are an elite SMC trading analyst reviewing one week of automated zone alerts.
Win rate = wins / (wins + losses). Pending, not_triggered, and invalidated are EXCLUDED from win rate.
Outcomes determined by entry-gated SL/TP candle scan (3-step: SL before entry → entry reached → SL vs TP1).

REVIEW PERIOD: {review_period_label}

WEEKLY NUMBERS:
- Total zone alerts: {len(weekly_alerts)}
- Triggered (entry reached, resolved): {wins}W / {losses}L = {win_rate:.1f}% win rate
- Not triggered (SL hit before entry or entry never reached): {not_triggered_count}
- Invalidated (setup cancelled before entry): {invalidated_count}
- Still pending: {pending}
- Entry Alerts sent (trigger confirmed): {alert2_triggered}
- Invalidation emails sent: {alert2_invalidated}

ASSET CLASS BREAKDOWN:
{json.dumps(asset_class_stats, indent=2)}

GEOPOLITICAL SPLIT:
- Clean technical: {len(clean_alerts)} alerts → {cl_wins}W/{cl_losses}L → {cl_wr}% WR
- Geo-flagged: {len(geo_alerts)} alerts → {geo_wins}W/{geo_losses}L → {geo_wr}% WR
- News-active: {len(news_alerts)} alerts → {news_wins}W/{news_losses}L → {news_wr}% WR

SESSION BREAKDOWN (triggered only):
{json.dumps(session_stats, indent=2)}

LOSS CLUSTERS (IST hours with 2+ losses):
{str(loss_clusters) if loss_clusters else "None"}

PAIR PERFORMANCE:
{json.dumps(pair_stats, indent=2)}

TRIGGERED TRADE DATA:
{json.dumps(alert_summary, indent=2)}

Return ONLY raw JSON. No markdown. No code fences.
{{
  "overall_grade": "A/B/C/D",
  "grade_comment": "one sentence on triggered trades only",
  "best_pair": "pair or none",
  "worst_pair": "pair or none",
  "pattern_insight": "which weighted confluence combination appeared most in winners",
  "confidence_calibration": "do higher-scored alerts win more — cite numbers from score_breakdown",
  "session_summary": "best and worst session with win rates",
  "session_recommendation": "one actionable sentence",
  "timing_observation": "IST windows where losses cluster",
  "timing_recommendation": "one actionable sentence",
  "asset_class_insight": "compare forex vs indices vs crypto vs metals win rates",
  "geo_insight": "compare geo-flagged vs clean technical win rates — always provide even if no geo trades",
  "not_triggered_insight": "what does the not_triggered rate tell us about entry aggressiveness",
  "zone_fatigue_flags": {json.dumps(python_zone_flags)},
  "streak_summary": "notable win/loss streaks",
  "drawdown_flag": "none or describe if 3+ consecutive losses",
  "improvement_suggestion": "one specific actionable change based on this week",
  "entry_alert_observation": "how many progressed to entry alert vs invalidated"
}}"""

    result, err = call_gemini(prompt)
    if err:
        print(f"  Gemini analysis error: {err}")
        return None
    return result


# ── Excel journal builder ─────────────────────────────────────────────────────
def build_excel_journal(weekly_alerts, analysis):
    wb = openpyxl.Workbook()

    C_HDR_BG  = "1A1A2E"; C_HDR_FG = "FFFFFF"; C_SEC_BG = "2C3E6B"
    C_WIN     = "D5F5E3"; C_LOSS   = "FADBD8"; C_INVALID = "FDEBD0"
    C_PENDING = "FDFEFE"; C_NOT_TRIG = "D6EAF8"; C_ALT = "F2F3F4"
    C_ID = "D6EAF8"; C_SETUP = "D5F5E3"; C_RISK = "FDEDEC"
    C_TRIG = "FEF9E7"; C_OUT = "EBF5FB"; C_LEARN = "F5EEF8"

    thin = Side(style='thin', color='CCCCCC')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    def hfont(sz=10): return Font(name='Arial', bold=True, color=C_HDR_FG, size=sz)
    def cfont(bold=False, color="000000", size=9):
        return Font(name='Arial', bold=bold, color=color, size=size)
    def bg(c): return PatternFill("solid", fgColor=c)
    def al(h='left', wrap=True): return Alignment(horizontal=h, vertical='center', wrap_text=wrap)
    def bc(cell): cell.border = bdr

    # ── TAB 1: WEEKLY SUMMARY ─────────────────────────────────────────────────
    ws_s = wb.active
    ws_s.title = "Weekly Summary"
    ws_s.sheet_view.showGridLines = False
    ws_s.column_dimensions['A'].width = 38
    ws_s.column_dimensions['B'].width = 28

    ist_date = ist_now().strftime("%d %b %Y")

    wins_list   = [a for a in weekly_alerts if a.get('outcome') == 'win_tp1']
    losses_list = [a for a in weekly_alerts if a.get('outcome') == 'loss']
    inv_list    = [a for a in weekly_alerts if a.get('outcome') == 'invalidated']
    nt_list     = [a for a in weekly_alerts if a.get('outcome') == 'not_triggered']
    pend_list   = [a for a in weekly_alerts if a.get('outcome') == 'pending']
    triggered   = wins_list + losses_list

    w_n = len(wins_list); l_n = len(losses_list); inv_n = len(inv_list)
    nt_n = len(nt_list); pend_n = len(pend_list)
    wr = round(w_n/(w_n+l_n)*100, 1) if (w_n+l_n) > 0 else 0
    inv_rate = round(inv_n/len(weekly_alerts)*100, 1) if weekly_alerts else 0
    net_pnl = sum(estimate_pnl(a)[0] for a in triggered)
    a2_trig = sum(1 for a in weekly_alerts if a.get('entry_alert_sent'))
    a2_inv  = sum(1 for a in weekly_alerts if a.get('invalidation_email_sent'))

    rrs = []
    for a in wins_list:
        try:
            entry = float(str(a.get('entry','0')).split('-')[0].strip() or 0)
            sl = float(a.get('sl',0) or 0); tp1 = float(a.get('tp1',0) or 0)
            bias = a.get('bias','')
            if entry > 0 and sl > 0 and tp1 > 0:
                rp = (entry-sl) if bias=="LONG" else (sl-entry)
                rw = (tp1-entry) if bias=="LONG" else (entry-tp1)
                if rp > 0: rrs.append(rw/rp)
        except Exception: pass
    avg_rr = round(sum(rrs)/len(rrs), 2) if rrs else 0.0
    avg_cw = round(sum(a.get('confidence_score',0) for a in wins_list)/len(wins_list), 1) if wins_list else 0
    avg_cl = round(sum(a.get('confidence_score',0) for a in losses_list)/len(losses_list), 1) if losses_list else 0

    def safe(key): return (analysis.get(key,'—') or '—') if analysis else '—'

    ws_s.merge_cells('A1:B1')
    ws_s['A1'].value = f"Weekly Trading Review — {ist_date}"
    ws_s['A1'].font = Font(name='Arial', bold=True, color=C_HDR_FG, size=13)
    ws_s['A1'].fill = bg(C_HDR_BG); ws_s['A1'].alignment = al('center')
    ws_s.row_dimensions[1].height = 28

    rows = [
        ("PERFORMANCE",                    None,              C_SEC_BG),
        ("Review Period",                   review_period_label, None),
        ("Total Zone Alerts",               len(weekly_alerts), None),
        ("Triggered Trades (W + L)",        w_n + l_n,          None),
        ("Wins",                            w_n,                None),
        ("Losses",                          l_n,                None),
        ("Win Rate (triggered only)",       f"{wr}%",           None),
        ("Not Triggered (entry not reached)", nt_n,             None),
        ("Invalidated (setup cancelled)",   inv_n,              None),
        ("Invalidation Rate",               f"{inv_rate}%",     None),
        ("Entry Alerts Sent",               a2_trig,            None),
        ("Invalidation Emails Sent",        a2_inv,             None),
        ("Still Pending",                   pend_n,             None),
        ("Geo-Flagged Alerts",              len([a for a in weekly_alerts if a.get('geo_flag')]), None),
        ("News-Active Alerts",              len([a for a in weekly_alerts if a.get('news_flag','none').lower()!='none']), None),
        ("Net Estimated P&L (USD)",         f"${net_pnl:+.2f}", None),
        ("QUALITY METRICS",                None,              C_SEC_BG),
        ("Avg R:R on Wins",                 f"{avg_rr}x",       None),
        ("Avg Score — Wins",                f"{avg_cw}/10",     None),
        ("Avg Score — Losses",              f"{avg_cl}/10",     None),
        ("ANALYSIS",                       None,              C_SEC_BG),
        ("Overall Grade",                   safe('overall_grade'), None),
        ("Grade Comment",                   safe('grade_comment'), None),
        ("Best Pair",                       safe('best_pair'),  None),
        ("Weakest Pair",                    safe('worst_pair'), None),
        ("Pattern Insight",                 safe('pattern_insight'), None),
        ("Confidence Calibration",          safe('confidence_calibration'), None),
        ("Session Summary",                 safe('session_summary'), None),
        ("Session Recommendation",          safe('session_recommendation'), None),
        ("Timing Observation",              safe('timing_observation'), None),
        ("Timing Recommendation",           safe('timing_recommendation'), None),
        ("Asset Class Insight",             safe('asset_class_insight'), None),
        ("Geo-Flagged Insight",             safe('geo_insight'), None),
        ("Not-Triggered Insight",           safe('not_triggered_insight'), None),
        ("Streak Summary",                  safe('streak_summary'), None),
        ("Drawdown Flag",                   safe('drawdown_flag'), None),
        ("Improvement Suggestion",          safe('improvement_suggestion'), None),
    ]

    ri = 2
    for label, value, sec in rows:
        if sec:
            ws_s.merge_cells(f'A{ri}:B{ri}')
            ws_s[f'A{ri}'].value = label
            ws_s[f'A{ri}'].font = hfont(10)
            ws_s[f'A{ri}'].fill = bg(sec)
            ws_s[f'A{ri}'].alignment = al()
            ws_s.row_dimensions[ri].height = 18
        else:
            ws_s[f'A{ri}'].value = label
            ws_s[f'B{ri}'].value = value
            ws_s[f'A{ri}'].font = cfont(bold=True, size=10)
            ws_s[f'B{ri}'].font = cfont(size=10)
            rbg = C_ALT if ri % 2 == 0 else "FFFFFF"
            ws_s[f'A{ri}'].fill = bg(rbg); ws_s[f'B{ri}'].fill = bg(rbg)
            ws_s[f'A{ri}'].alignment = al(); ws_s[f'B{ri}'].alignment = al('left')
            ws_s.row_dimensions[ri].height = 20
        bc(ws_s[f'A{ri}']); bc(ws_s[f'B{ri}'])
        ri += 1

    # Zone flags
    flags = safe('zone_fatigue_flags')
    if isinstance(flags, list) and flags:
        ws_s.merge_cells(f'A{ri}:B{ri}')
        ws_s[f'A{ri}'].value = "ZONE DEPLETION FLAGS"
        ws_s[f'A{ri}'].font = hfont(10); ws_s[f'A{ri}'].fill = bg(C_SEC_BG)
        ws_s[f'A{ri}'].alignment = al()
        bc(ws_s[f'A{ri}']); bc(ws_s[f'B{ri}']); ri += 1
        for flag in flags:
            ws_s.merge_cells(f'A{ri}:B{ri}')
            ws_s[f'A{ri}'].value = f"⚠ {flag}"
            ws_s[f'A{ri}'].font = cfont(color="E67E22", size=10)
            ws_s[f'A{ri}'].fill = bg("FFF3CD"); ws_s[f'A{ri}'].alignment = al()
            bc(ws_s[f'A{ri}']); bc(ws_s[f'B{ri}']); ri += 1

    # ── TAB 2: TRADE JOURNAL ──────────────────────────────────────────────────
    ws = wb.create_sheet("Trade Journal")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A3'

    columns = [
        ("Date (IST)",           16, C_ID),    ("Pair",              10, C_ID),
        ("Asset Class",          12, C_ID),    ("Session",           12, C_ID),
        ("Direction",            11, C_ID),
        ("Score",                10, C_SETUP), ("Key Confluences",   40, C_SETUP),
        ("Missing",              32, C_SETUP), ("Trigger Condition", 40, C_SETUP),
        ("Entry",                13, C_RISK),  ("Stop Loss",         13, C_RISK),
        ("TP1",                  13, C_RISK),  ("TP2",               13, C_RISK),
        ("R:R (TP1)",            12, C_RISK),  ("Risk (USD)",        11, C_RISK),
        ("Geo Flag",             10, C_TRIG),  ("News Flag",         18, C_TRIG),
        ("Outcome",              14, C_OUT),   ("Outcome Price",     14, C_OUT),
        ("Points +/-",           12, C_OUT),   ("Est. P&L",          12, C_OUT),
        ("Entry Alert Sent",     14, C_LEARN), ("System Note",       36, C_LEARN),
    ]

    groups = [
        ("IDENTITY",      1,  5,  C_ID),    ("SETUP",    6,  9,  C_SETUP),
        ("RISK LEVELS",   10, 15, C_RISK),  ("CONTEXT",  16, 17, C_TRIG),
        ("OUTCOME",       18, 21, C_OUT),   ("TRACKING", 22, 23, C_LEARN),
    ]
    for grp, cs, ce, gc in groups:
        sl_l = get_column_letter(cs); el_l = get_column_letter(ce)
        ws.merge_cells(f'{sl_l}1:{el_l}1')
        cell = ws[f'{sl_l}1']
        cell.value = grp; cell.font = hfont(10); cell.fill = bg(C_HDR_BG)
        cell.alignment = al('center'); bc(cell)
    ws.row_dimensions[1].height = 22

    for ci, (cn, cw, cc) in enumerate(columns, 1):
        lt = get_column_letter(ci); cell = ws[f'{lt}2']
        cell.value = cn; cell.font = Font(name='Arial', bold=True, color="1A1A2E", size=9)
        cell.fill = bg(cc); cell.alignment = al('center'); bc(cell)
        ws.column_dimensions[lt].width = cw
    ws.row_dimensions[2].height = 32

    dr = 3
    for a in weekly_alerts:
        outcome = a.get('outcome', 'pending')
        if   outcome == 'win_tp1':     rbg = C_WIN
        elif outcome == 'loss':        rbg = C_LOSS
        elif outcome == 'invalidated': rbg = C_INVALID
        elif outcome == 'not_triggered': rbg = C_NOT_TRIG
        else:                          rbg = C_PENDING

        pnl_usd, pts_str = estimate_pnl(a)
        rr_str = "—"
        try:
            entry = float(str(a.get('entry','0')).split('-')[0].strip() or 0)
            sl_v = float(a.get('sl',0) or 0); tp1_v = float(a.get('tp1',0) or 0)
            bias = a.get('bias','')
            if entry>0 and sl_v>0 and tp1_v>0:
                rp = (entry-sl_v) if bias=="LONG" else (sl_v-entry)
                rw = (tp1_v-entry) if bias=="LONG" else (entry-tp1_v)
                if rp>0: rr_str = f"{round(rw/rp,2)}:1"
        except Exception: pass

        try:
            utc_dt = datetime.strptime(a['timestamp_utc'], "%Y-%m-%d %H:%M")
            ist_dt = utc_dt + timedelta(hours=5, minutes=30)
            date_str = ist_dt.strftime("%d %b %Y %H:%M")
            session = get_session(utc_dt.hour)
        except Exception:
            date_str = a.get('timestamp_utc',''); session = "—"

        conf_str = "; ".join(a.get('confluences', []))
        miss_raw = a.get('missing', [])
        miss_str = "; ".join(m.get('item','') if isinstance(m,dict) else str(m) for m in miss_raw) if isinstance(miss_raw, list) else str(miss_raw)

        outcome_label = {'win_tp1':'WIN','loss':'LOSS','invalidated':'INVALIDATED',
                         'not_triggered':'NOT TRIGGERED','pending':'PENDING'}.get(outcome, outcome.upper())
        op = a.get('outcome_price')
        op_str = str(round(float(op), 5)) if op is not None else "—"
        pnl_d = f"${pnl_usd:+.2f}" if pnl_usd != 0.0 else "—"
        pair_type = a.get('pair_type', get_pair_type(a.get('pair','')))
        news_f = a.get('news_flag', 'none')
        a2_sent = "Yes" if a.get('entry_alert_sent') else "No"

        row_vals = [
            date_str, a.get('pair',''), pair_type.title(), session, a.get('bias',''),
            a.get('confidence_score',0), conf_str, miss_str, a.get('trigger',''),
            a.get('entry',''), a.get('sl',''), a.get('tp1',''), a.get('tp2',''),
            rr_str, f"${RISK_PER_TRADE:.0f}",
            "YES" if a.get('geo_flag') else "NO", news_f if news_f.lower()!='none' else "—",
            outcome_label, op_str, pts_str, pnl_d,
            a2_sent, a.get('confidence_reason', ''),
        ]

        for ci, val in enumerate(row_vals, 1):
            lt = get_column_letter(ci); cell = ws[f'{lt}{dr}']
            cell.value = val; cell.font = cfont(size=9)
            cell.fill = bg(rbg); cell.alignment = al(); bc(cell)
            if ci == 6:
                c = a.get('confidence_score', 0)
                cell.font = cfont(bold=True, size=9, color="27AE60" if c>=8 else "F39C12" if c>=7 else "E74C3C")
            if ci == 18:
                clr = {"WIN":"1E8449","LOSS":"C0392B","INVALIDATED":"E67E22",
                       "NOT TRIGGERED":"3498DB","PENDING":"888888"}
                cell.font = cfont(bold=True, size=9, color=clr.get(outcome_label,"000000"))
            if ci == 20:
                cell.font = cfont(bold=True, size=9, color="1E8449" if pnl_usd>0 else "C0392B" if pnl_usd<0 else "888888")
        ws.row_dimensions[dr].height = 50
        dr += 1

    # Legend
    dr += 1
    ws.merge_cells(f'A{dr}:D{dr}')
    ws[f'A{dr}'].value = "COLOUR LEGEND"; ws[f'A{dr}'].font = hfont(9)
    ws[f'A{dr}'].fill = bg(C_HDR_BG); ws[f'A{dr}'].alignment = al(); dr += 1

    for lbg, lt in [(C_WIN,"Green — Win (TP1 hit). Counted in win rate."),
                     (C_LOSS,"Red — Loss (SL hit). Counted in win rate."),
                     (C_NOT_TRIG,"Blue — Not triggered (entry never reached). Excluded from win rate."),
                     (C_INVALID,"Orange — Invalidated (setup cancelled). Excluded from win rate."),
                     (C_PENDING,"White — Still pending resolution.")]:
        ws.merge_cells(f'A{dr}:F{dr}')
        ws[f'A{dr}'].value = lt; ws[f'A{dr}'].fill = bg(lbg)
        ws[f'A{dr}'].font = cfont(size=9); ws[f'A{dr}'].alignment = al()
        ws.row_dimensions[dr].height = 15; dr += 1

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    return buf.read()


# ── Email HTML ────────────────────────────────────────────────────────────────
def insight_card(title, color, content):
    if not content or content == '—':
        return ""  # Hide empty sections
    return (f'<div style="background:#f8f9fa;padding:12px 14px;border-radius:10px;'
            f'margin-bottom:10px;border-left:4px solid {color};">'
            f'<p style="font-size:10px;color:{color};margin:0 0 3px;text-transform:uppercase;'
            f'font-weight:bold;letter-spacing:0.5px;">{title}</p>'
            f'<p style="font-size:12px;color:#333;margin:0;line-height:1.5;">{content}</p></div>')

def build_weekly_email_html(data, weekly_alerts, wins, losses,
                            invalidated_count, not_triggered_count, pending, win_rate,
                            analysis_run_time_ist="—"):
    ist_date = ist_now().strftime("%A, %d %b %Y")
    grade    = (data.get('overall_grade','—') or '—') if data else '—'
    grade_c  = "#27ae60" if grade=="A" else "#f39c12" if grade in ("B","C") else "#e74c3c"
    total    = len(weekly_alerts)
    net_pnl  = sum(estimate_pnl(a)[0] for a in weekly_alerts if a.get('outcome') in ('win_tp1','loss'))
    pnl_c    = "#27ae60" if net_pnl >= 0 else "#e74c3c"

    def safe(key): return (data.get(key,'—') or '—') if data else '—'

    drawdown = safe('drawdown_flag')
    ddhtml   = (f'<div style="background:#fef0f0;padding:10px 14px;border-radius:8px;'
                f'border-left:4px solid #e74c3c;margin-bottom:14px;">'
                f'<p style="margin:0;font-size:12px;color:#c0392b;">'
                f'<b>⚠ DRAWDOWN:</b> {drawdown}</p></div>'
                if drawdown and drawdown.lower() != 'none' else "")

    flags = safe('zone_fatigue_flags')
    flags_html = ""
    if isinstance(flags, list) and flags:
        flags_html = "".join([f'<li style="font-size:11px;color:#e67e22;margin-bottom:3px;">⚠ {z}</li>' for z in flags])
    else:
        flags_html = '<li style="font-size:11px;color:#aaa;">No depletion flags this week.</li>'

    # Geo section always shown
    geo_card = insight_card("GEO-FLAGGED vs CLEAN TECHNICAL", "#e67e22", safe('geo_insight'))
    if not geo_card:
        geo_card = insight_card("GEO-FLAGGED vs CLEAN TECHNICAL", "#e67e22",
                                "No geo-flagged trades this week. All resolved trades were clean technical setups.")

    return f"""<!DOCTYPE html>
<html>
<body style="font-family:Arial,sans-serif;background:#f0f2f5;padding:12px;margin:0;">
<div style="max-width:620px;margin:auto;background:white;border-radius:14px;
     overflow:hidden;box-shadow:0 4px 20px rgba(0,0,0,0.12);">

  <div style="background:#1a1a2e;padding:16px 20px;">
    <h2 style="color:white;margin:0;font-size:16px;">Weekly Trading Review</h2>
    <p style="color:#8899bb;margin:4px 0 0;font-size:11px;">{ist_date}</p>
    <p style="color:#8899bb;margin:2px 0 0;font-size:11px;">Review period: {review_period_label}</p>
    <p style="color:#445566;margin:2px 0 0;font-size:10px;">Generated: {analysis_run_time_ist}</p>
  </div>

  <div style="display:flex;background:#f8f9fa;border-bottom:1px solid #eee;flex-wrap:wrap;">
    <div style="flex:1;min-width:70px;padding:12px 8px;text-align:center;border-right:1px solid #eee;">
      <p style="margin:0;font-size:9px;color:#888;text-transform:uppercase;">Grade</p>
      <p style="margin:3px 0 0;font-size:24px;font-weight:bold;color:{grade_c};">{grade}</p>
    </div>
    <div style="flex:1;min-width:70px;padding:12px 8px;text-align:center;border-right:1px solid #eee;">
      <p style="margin:0;font-size:9px;color:#888;text-transform:uppercase;">Win Rate</p>
      <p style="margin:3px 0 0;font-size:20px;font-weight:bold;color:#1a1a2e;">{win_rate:.0f}%</p>
      <p style="margin:1px 0 0;font-size:8px;color:#aaa;">triggered</p>
    </div>
    <div style="flex:1;min-width:70px;padding:12px 8px;text-align:center;border-right:1px solid #eee;">
      <p style="margin:0;font-size:9px;color:#888;text-transform:uppercase;">W / L</p>
      <p style="margin:3px 0 0;font-size:18px;font-weight:bold;">
        <span style="color:#27ae60;">{wins}</span> /
        <span style="color:#e74c3c;">{losses}</span></p>
    </div>
    <div style="flex:1;min-width:70px;padding:12px 8px;text-align:center;border-right:1px solid #eee;">
      <p style="margin:0;font-size:9px;color:#888;text-transform:uppercase;">Not Trig</p>
      <p style="margin:3px 0 0;font-size:18px;font-weight:bold;color:#3498db;">{not_triggered_count}</p>
    </div>
    <div style="flex:1;min-width:70px;padding:12px 8px;text-align:center;border-right:1px solid #eee;">
      <p style="margin:0;font-size:9px;color:#888;text-transform:uppercase;">Invalidated</p>
      <p style="margin:3px 0 0;font-size:18px;font-weight:bold;color:#e67e22;">{invalidated_count}</p>
    </div>
    <div style="flex:1;min-width:70px;padding:12px 8px;text-align:center;">
      <p style="margin:0;font-size:9px;color:#888;text-transform:uppercase;">Est. P&L</p>
      <p style="margin:3px 0 0;font-size:18px;font-weight:bold;color:{pnl_c};">${net_pnl:+.0f}</p>
    </div>
  </div>

  <div style="padding:16px 20px;">
    {ddhtml}

    <p style="font-size:11px;color:#2980b9;background:#e8f4fd;padding:8px 12px;
       border-radius:8px;margin:0 0 16px;">
      Full trade journal attached — Excel workbook with scorecard breakdown, P&L, and outcomes.</p>

    {insight_card("PATTERN INTELLIGENCE", "#3498db", safe('pattern_insight'))}
    {insight_card("CONFIDENCE CALIBRATION", "#27ae60", safe('confidence_calibration'))}
    {insight_card("ASSET CLASS PERFORMANCE", "#9b59b6", safe('asset_class_insight'))}
    {insight_card("SESSION WIN RATE", "#3498db", safe('session_summary'))}
    {insight_card("SESSION RECOMMENDATION", "#27ae60", safe('session_recommendation'))}
    {insight_card("TIMING CLUSTERS", "#f39c12", safe('timing_observation'))}
    {insight_card("TIMING RECOMMENDATION", "#e67e22", safe('timing_recommendation'))}
    {geo_card}
    {insight_card("NOT-TRIGGERED ANALYSIS", "#3498db", safe('not_triggered_insight'))}
    {insight_card("ENTRY ALERT PIPELINE", "#27ae60", safe('entry_alert_observation'))}
    {insight_card("STREAK AWARENESS", "#9b59b6", safe('streak_summary'))}

    <div style="background:#fff8f0;padding:12px 14px;border-radius:10px;
         margin-bottom:10px;border-left:4px solid #e67e22;">
      <p style="font-size:10px;color:#e67e22;margin:0 0 4px;text-transform:uppercase;
         font-weight:bold;letter-spacing:0.5px;">ZONE DEPLETION FLAGS</p>
      <ul style="list-style:none;padding:0;margin:0;">{flags_html}</ul>
    </div>

    <div style="background:#1a1a2e;padding:14px 16px;border-radius:8px;">
      <p style="color:#8899bb;font-size:9px;margin:0 0 4px;text-transform:uppercase;letter-spacing:1px;">SYSTEM IMPROVEMENT</p>
      <p style="color:white;font-size:12px;margin:0;line-height:1.5;">{safe('improvement_suggestion')}</p>
    </div>
  </div>
</div>
</body>
</html>"""

# ── Email senders ─────────────────────────────────────────────────────────────
def send_status_email(message):
    ist_date = ist_now().strftime("%d %b")
    subject  = f"Weekly Review | {ist_date}"
    html = (f'<!DOCTYPE html><html><body style="font-family:Arial,sans-serif;background:#f0f2f5;padding:16px;">'
            f'<div style="max-width:500px;margin:auto;background:white;border-radius:12px;overflow:hidden;">'
            f'<div style="background:#1a1a2e;padding:16px 20px;">'
            f'<h2 style="color:white;margin:0;font-size:15px;">Weekly Review</h2></div>'
            f'<div style="padding:20px;"><p style="font-size:13px;color:#333;">{message}</p></div>'
            f'</div></body></html>')
    recipients = config["account"].get("alert_emails", [ALERT_EMAIL])
    for r in recipients:
        msg = MIMEMultipart("mixed"); msg["Subject"] = subject
        msg["From"] = GMAIL_ADDRESS; msg["To"] = r
        msg.attach(MIMEText(html, "html"))
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(GMAIL_ADDRESS, GMAIL_PASS); s.sendmail(GMAIL_ADDRESS, r, msg.as_string())

def send_weekly_email(html_body, excel_bytes, total, wins, losses,
                      invalidated_count, not_triggered_count, win_rate):
    ist_dt   = ist_now()
    filename = f"Trading_Journal_{ist_dt.strftime('%d_%b_%Y')}.xlsx"
    subject  = (f"Weekly Review | {total} alerts | "
                f"{wins}W {losses}L {invalidated_count}INV {not_triggered_count}NT | "
                f"{win_rate:.0f}% WR | {ist_dt.strftime('%d %b')}")
    recipients = config["account"].get("alert_emails", [ALERT_EMAIL])
    for r in recipients:
        msg = MIMEMultipart("mixed"); msg["Subject"] = subject
        msg["From"] = GMAIL_ADDRESS; msg["To"] = r
        msg.attach(MIMEText(html_body, "html"))
        part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        part.set_payload(excel_bytes); encoders.encode_base64(part)
        part.add_header("Content-Disposition", "attachment", filename=filename)
        msg.attach(part)
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as s:
            s.login(GMAIL_ADDRESS, GMAIL_PASS); s.sendmail(GMAIL_ADDRESS, r, msg.as_string())
        print(f"  Weekly review sent to {r}")

# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
analysis_run_time_ist = ist_now().strftime("%H:%M IST, %d %b %Y")
print(f"Weekly review started {analysis_run_time_ist}")
print("  Running outcome update (entry-gated — 0 Gemini calls)...")
update_outcomes()

weekly_alerts_all = get_weekly_alerts()

# Filter: minimum score from config
weekly_alerts_below = [a for a in weekly_alerts_all if a.get('confidence_score', 0) < MIN_SCORE]
weekly_alerts = [a for a in weekly_alerts_all if a not in weekly_alerts_below]
below_count = len(weekly_alerts_below)

if below_count > 0:
    print(f"  WARNING: {below_count} alert(s) below min score {MIN_SCORE} excluded.")

if not weekly_alerts:
    send_status_email(
        f"No qualifying zone alerts (score >= {MIN_SCORE}) found for {review_period_label}. "
        "The system is monitoring — alerts will appear once the week is captured."
    )
    exit(0)

wins              = sum(1 for a in weekly_alerts if a.get('outcome') == 'win_tp1')
losses            = sum(1 for a in weekly_alerts if a.get('outcome') == 'loss')
invalidated_count = sum(1 for a in weekly_alerts if a.get('outcome') == 'invalidated')
not_triggered     = sum(1 for a in weekly_alerts if a.get('outcome') == 'not_triggered')
pending           = sum(1 for a in weekly_alerts if a.get('outcome') == 'pending')
alert2_triggered  = sum(1 for a in weekly_alerts if a.get('entry_alert_sent'))
alert2_invalidated = sum(1 for a in weekly_alerts if a.get('invalidation_email_sent'))
win_rate          = (wins/(wins+losses)*100) if (wins+losses) > 0 else 0

# Pair stats
pair_stats = {}
for a in weekly_alerts:
    p = a['pair']
    if p not in pair_stats:
        pair_stats[p] = {'alerts':0,'wins':0,'losses':0,'invalidated':0,
                         'not_triggered':0,'pending':0,'entry_alerts':0}
    pair_stats[p]['alerts'] += 1
    o = a.get('outcome','pending')
    if   o == 'win_tp1':      pair_stats[p]['wins']          += 1
    elif o == 'loss':         pair_stats[p]['losses']        += 1
    elif o == 'invalidated':  pair_stats[p]['invalidated']   += 1
    elif o == 'not_triggered':pair_stats[p]['not_triggered'] += 1
    else:                     pair_stats[p]['pending']       += 1
    if a.get('entry_alert_sent'): pair_stats[p]['entry_alerts'] += 1

# Asset class stats
asset_class_stats = {}
for a in weekly_alerts:
    pt = a.get('pair_type', get_pair_type(a.get('pair', '')))
    if pt not in asset_class_stats:
        asset_class_stats[pt] = {'alerts':0,'wins':0,'losses':0,'not_triggered':0}
    asset_class_stats[pt]['alerts'] += 1
    o = a.get('outcome','')
    if   o == 'win_tp1':      asset_class_stats[pt]['wins']          += 1
    elif o == 'loss':         asset_class_stats[pt]['losses']        += 1
    elif o == 'not_triggered':asset_class_stats[pt]['not_triggered'] += 1
for pt, s in asset_class_stats.items():
    w, l = s['wins'], s['losses']
    s['win_rate'] = round(w/(w+l)*100, 1) if (w+l) > 0 else None

print("  Calling Gemini for analysis...")
analysis = build_weekly_analysis(
    weekly_alerts, wins, losses, invalidated_count, not_triggered, pending,
    win_rate, pair_stats, asset_class_stats, alert2_triggered, alert2_invalidated
)

if not analysis:
    send_status_email(
        f"Found {len(weekly_alerts)} alerts but Gemini analysis failed. "
        f"Raw: {wins}W / {losses}L / {invalidated_count}INV / {not_triggered}NT / {pending}P."
    )
    exit(0)

print("  Building Excel journal...")
excel_bytes = build_excel_journal(weekly_alerts, analysis)

print("  Building email...")
html = build_weekly_email_html(
    analysis, weekly_alerts, wins, losses, invalidated_count, not_triggered,
    pending, win_rate, analysis_run_time_ist
)

send_weekly_email(html, excel_bytes, len(weekly_alerts),
                  wins, losses, invalidated_count, not_triggered, win_rate)
print("  Done.")
